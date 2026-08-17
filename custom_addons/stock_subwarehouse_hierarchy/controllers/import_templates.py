import json
from io import BytesIO

from odoo.http import Controller, content_disposition, request, route
from werkzeug.utils import redirect


class StockSubwarehouseImportTemplateController(Controller):
    @route(
        "/stock_subwarehouse_hierarchy/product_encoder/template.xlsx",
        type="http",
        auth="user",
    )
    def download_product_encoder_template(self, **kwargs):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as error:
            raise ImportError("Product encoder templates require openpyxl") from error

        columns = [
            ("产品名称", "单板滑雪鞋"),
            ("厂家代码/名称", "15"),
            ("生产年月(YYMM)", "2410"),
            ("产品类型代码/名称", "S1"),
            ("成品类型(M/F)", "M"),
            ("成人儿童(A/K)", "A"),
            ("硬度", "7"),
            ("颜色", "黑"),
            ("尺码", "250"),
        ]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "批量编码"
        sheet.append([header for header, _sample in columns])
        sheet.append([sample for _header, sample in columns])
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        for index, column in enumerate(columns, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = max(len(column[0]) + 4, 16)
        note_sheet = workbook.create_sheet("编码说明")
        note_sheet.append(["组成部分", "说明"])
        for row in [
            ("厂家", "可填 01-15 厂家代码，或填写已配置的厂家名称，例如 熙堃。"),
            ("生产年月", "四位 YYMM，例如 2410。"),
            ("产品类型", "可填 S1、S2、T5 等代码，也可填写单板滑雪鞋、五指手套等名称。"),
            ("成品类型", "M=成品，F=非成品。"),
            ("成人儿童", "A=成人，K=儿童/青少年。"),
            ("硬度", "无硬度可留空，编码为 000；其他字段不可留空。"),
            ("颜色", "可填中文颜色组合或四位颜色代码。"),
            ("尺码", "可填数字、XS/S/M/L/XL，通码会编码为 ###。"),
            ("名称自动匹配", "厂家、产品类型、颜色会与完整编码规则按共同字符数匹配；并列或无可靠结果的行会被拒绝。"),
            ("失败行", "空白或未匹配导致无法生成编码的行不会成为可使用结果，并会在页面和结果 Excel 中注明原行号及原因。"),
        ]:
            note_sheet.append(row)
        for cell in note_sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        note_sheet.column_dimensions["A"].width = 18
        note_sheet.column_dimensions["B"].width = 70
        output = BytesIO()
        workbook.save(output)
        headers = [
            ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            ("Content-Disposition", content_disposition("product_quick_encoder_template_zh.xlsx")),
        ]
        return request.make_response(output.getvalue(), headers=headers)

    @route(
        "/stock_subwarehouse_hierarchy/product_encoder/result.xlsx",
        type="http",
        auth="user",
    )
    def download_product_encoder_result(self, wizard_id=None, **kwargs):
        if not str(wizard_id or "").isdigit():
            return request.not_found()
        wizard = request.env["stock.subwarehouse.product.quick.encoder.wizard"].search([
            ("id", "=", int(wizard_id)),
            ("create_uid", "=", request.env.user.id),
        ], limit=1)
        if not wizard:
            return request.not_found()
        content = wizard._generate_result_xlsx()
        headers = [
            ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            ("Content-Disposition", content_disposition("product_quick_encoder_result.xlsx")),
        ]
        return request.make_response(content, headers=headers)

    @route(
        "/stock_subwarehouse_hierarchy/storefront",
        type="http",
        auth="user",
    )
    def open_separated_storefront(self, **kwargs):
        host = request.httprequest.host.split(":", 1)[0]
        scheme = request.httprequest.scheme
        return redirect(f"{scheme}://{host}:8070/", code=302)

    def _parse_ids(self, ids):
        return [
            int(record_id)
            for record_id in (ids or "").split(",")
            if record_id.strip().isdigit()
        ]

    def _records_from_request(self, model_name, ids="", domain="[]"):
        Model = request.env[model_name]
        record_ids = self._parse_ids(ids)
        if record_ids:
            return Model.browse(record_ids).exists()
        try:
            parsed_domain = json.loads(domain or "[]")
        except json.JSONDecodeError:
            parsed_domain = []
        return Model.search(parsed_domain)

    @route(
        "/stock_subwarehouse_hierarchy/import_template/mrp_production.xlsx",
        type="http",
        auth="user",
    )
    def download_mrp_production_import_template(self, **kwargs):
        content = request.env["mrp.production"]._generate_dynamic_import_template_xlsx()
        headers = [
            (
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            (
                "Content-Disposition",
                content_disposition("manufacturing_import_template_zh.xlsx"),
            ),
        ]
        return request.make_response(content, headers=headers)

    @route(
        "/stock_subwarehouse_hierarchy/import_template/mrp_bom.xlsx",
        type="http",
        auth="user",
    )
    def download_mrp_bom_import_template(self, **kwargs):
        content = request.env["mrp.bom"]._generate_bom_import_template_xlsx()
        headers = [
            (
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            (
                "Content-Disposition",
                content_disposition("bom_import_template_zh.xlsx"),
            ),
        ]
        return request.make_response(content, headers=headers)

    @route(
        "/stock_subwarehouse_hierarchy/import_template/product_template.xlsx",
        type="http",
        auth="user",
    )
    def download_product_template_import_template(self, **kwargs):
        content = request.env["product.template"]._generate_dynamic_product_import_template_xlsx()
        headers = [
            (
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            (
                "Content-Disposition",
                content_disposition("product_import_template_zh.xlsx"),
            ),
        ]
        return request.make_response(content, headers=headers)

    @route(
        "/stock_subwarehouse_hierarchy/import_template/sale_order.xlsx",
        type="http",
        auth="user",
    )
    def download_sale_order_import_template(self, **kwargs):
        content = request.env["sale.order"]._generate_sale_order_import_template_xlsx()
        headers = [
            (
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            (
                "Content-Disposition",
                content_disposition("sale_order_import_template_zh.xlsx"),
            ),
        ]
        return request.make_response(content, headers=headers)

    @route(
        "/stock_subwarehouse_hierarchy/import_template/distributor.xlsx",
        type="http",
        auth="user",
    )
    def download_distributor_import_template(self, **kwargs):
        content = request.env[
            "res.partner"
        ]._generate_partner_channel_import_template_xlsx("distributor")
        headers = [
            (
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            (
                "Content-Disposition",
                content_disposition("distributor_import_template_zh.xlsx"),
            ),
        ]
        return request.make_response(content, headers=headers)

    @route(
        "/stock_subwarehouse_hierarchy/import_template/supplier.xlsx",
        type="http",
        auth="user",
    )
    def download_supplier_import_template(self, **kwargs):
        content = request.env[
            "res.partner"
        ]._generate_partner_channel_import_template_xlsx("supplier")
        headers = [
            (
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            (
                "Content-Disposition",
                content_disposition("supplier_import_template_zh.xlsx"),
            ),
        ]
        return request.make_response(content, headers=headers)

    @route(
        "/stock_subwarehouse_hierarchy/export/partner_channel.xlsx",
        type="http",
        auth="user",
    )
    def export_partner_channel_import_format(
        self,
        channel="",
        domain="[]",
        **kwargs,
    ):
        if channel not in ("distributor", "supplier"):
            return request.not_found()
        records = self._records_from_request(
            "res.partner",
            domain=domain,
        )
        if channel == "distributor":
            records = records.filtered(
                lambda partner: partner.x_is_distributor and partner.is_company
            )
            filename = "distributor_export_import_format_zh.xlsx"
        else:
            records = records.filtered(
                lambda partner: partner.supplier_rank > 0 and partner.is_company
            )
            filename = "supplier_export_import_format_zh.xlsx"
        content = records._generate_partner_channel_export_xlsx(channel)
        headers = [
            (
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            (
                "Content-Disposition",
                content_disposition(filename),
            ),
        ]
        return request.make_response(content, headers=headers)

    @route(
        "/stock_subwarehouse_hierarchy/export/product_template.xlsx",
        type="http",
        auth="user",
    )
    def export_product_template_import_format(self, ids="", domain="[]", **kwargs):
        records = self._records_from_request("product.template", ids=ids, domain=domain)
        content = records._generate_dynamic_product_export_xlsx()
        headers = [
            (
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            (
                "Content-Disposition",
                content_disposition("product_export_import_format_zh.xlsx"),
            ),
        ]
        return request.make_response(content, headers=headers)

    @route(
        "/stock_subwarehouse_hierarchy/export/mrp_production.xlsx",
        type="http",
        auth="user",
    )
    def export_mrp_production_import_format(self, ids="", domain="[]", **kwargs):
        records = self._records_from_request("mrp.production", ids=ids, domain=domain)
        content = records._generate_dynamic_export_xlsx()
        headers = [
            (
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            (
                "Content-Disposition",
                content_disposition("manufacturing_export_import_format_zh.xlsx"),
            ),
        ]
        return request.make_response(content, headers=headers)

    @route(
        "/stock_subwarehouse_hierarchy/export/mrp_bom.xlsx",
        type="http",
        auth="user",
    )
    def export_mrp_bom_import_format(self, ids="", domain="[]", **kwargs):
        records = self._records_from_request("mrp.bom", ids=ids, domain=domain)
        content = records._generate_bom_export_xlsx()
        headers = [
            (
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            (
                "Content-Disposition",
                content_disposition("bom_export_import_format_zh.xlsx"),
            ),
        ]
        return request.make_response(content, headers=headers)

    @route(
        "/stock_subwarehouse_hierarchy/export/sale_order.xlsx",
        type="http",
        auth="user",
    )
    def export_sale_order_import_format(self, ids="", domain="[]", **kwargs):
        records = self._records_from_request("sale.order", ids=ids, domain=domain)
        content = records._generate_sale_order_export_xlsx()
        headers = [
            (
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            (
                "Content-Disposition",
                content_disposition("sale_order_export_import_format_zh.xlsx"),
            ),
        ]
        return request.make_response(content, headers=headers)
