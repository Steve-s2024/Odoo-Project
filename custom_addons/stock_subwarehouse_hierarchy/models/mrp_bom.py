from io import BytesIO
from urllib.parse import urlencode

from odoo import _, api, fields, models
from odoo.exceptions import UserError
from odoo.fields import Command


BOM_IMPORT_TEMPLATE_ROUTE = "/stock_subwarehouse_hierarchy/import_template/mrp_bom.xlsx"
BOM_IMPORT_COMPONENT_SLOT_COUNT = 20
BOM_COMPONENT_PRODUCT_FIELD = "x_import_bom_component_product"
BOM_COMPONENT_QTY_FIELD = "x_import_bom_component_qty"
BOM_COMPONENT_UOM_FIELD = "x_import_bom_component_uom"


class MrpBom(models.Model):
    _inherit = "mrp.bom"

    x_import_bom_component_product = fields.Char(string="导入组件产品", copy=False)
    x_import_bom_component_qty = fields.Float(string="导入组件数量", copy=False)
    x_import_bom_component_uom = fields.Char(string="导入组件单位", copy=False)
    x_reusable_template_id = fields.Many2one(
        "stock.subwarehouse.bom.template",
        string="可复用BOM模板",
        copy=False,
        check_company=True,
    )

    # These fields keep previously downloaded 20-slot workbooks importable.
    for _slot_number in range(1, BOM_IMPORT_COMPONENT_SLOT_COUNT + 1):
        locals()[f"x_import_bom_component_product_{_slot_number}"] = fields.Char(
            string=f"导入组件产品 {_slot_number}", copy=False
        )
        locals()[f"x_import_bom_component_qty_{_slot_number}"] = fields.Float(
            string=f"导入组件数量 {_slot_number}", copy=False
        )
        locals()[f"x_import_bom_component_uom_{_slot_number}"] = fields.Char(
            string=f"导入组件单位 {_slot_number}", copy=False
        )
    del _slot_number

    @api.model
    def get_import_templates(self):
        return [{
            "label": _("物料清单导入模板"),
            "template": BOM_IMPORT_TEMPLATE_ROUTE,
        }]

    @staticmethod
    def _get_bom_import_component_field_names():
        field_names = {
            BOM_COMPONENT_PRODUCT_FIELD,
            BOM_COMPONENT_QTY_FIELD,
            BOM_COMPONENT_UOM_FIELD,
        }
        for slot_number in range(1, BOM_IMPORT_COMPONENT_SLOT_COUNT + 1):
            field_names.update({
                f"x_import_bom_component_product_{slot_number}",
                f"x_import_bom_component_qty_{slot_number}",
                f"x_import_bom_component_uom_{slot_number}",
            })
        return field_names

    def load(self, import_fields, import_data):
        component_fields = self._get_bom_import_component_field_names()
        if not any(field_name in import_fields for field_name in component_fields):
            return super().load(import_fields, import_data)

        kept_indexes = [
            index
            for index, field_name in enumerate(import_fields)
            if field_name not in component_fields
        ]
        cleaned_fields = [import_fields[index] for index in kept_indexes]
        groups = self._group_bom_import_rows(import_fields, import_data, kept_indexes)
        result = super().load(cleaned_fields, [group["parent"] for group in groups])
        if result.get("ids"):
            for bom, group in zip(self.browse(result["ids"]), groups):
                if group["components"]:
                    bom._write_imported_bom_components(group["components"])
        return result

    def _group_bom_import_rows(self, import_fields, import_data, kept_indexes):
        identity_fields = [
            field_name
            for field_name in ("id", "code", "product_id", "product_tmpl_id")
            if field_name in import_fields
        ]
        groups = []
        current_group = None
        current_key = None
        for row in import_data:
            key = tuple(
                str(row[import_fields.index(field_name)] or "").strip()
                for field_name in identity_fields
            )
            has_identity = any(key)
            if current_group is None or (has_identity and key != current_key):
                current_group = {
                    "parent": [row[index] for index in kept_indexes],
                    "components": [],
                }
                groups.append(current_group)
                current_key = key
            current_group["components"].extend(
                self._extract_bom_components_from_row(import_fields, row)
            )
        return groups

    def _extract_bom_component_import_rows(self, import_fields, import_data):
        return [
            self._extract_bom_components_from_row(import_fields, row)
            for row in import_data
        ]

    def _extract_bom_components_from_row(self, import_fields, row):
        components = []
        if BOM_COMPONENT_PRODUCT_FIELD in import_fields:
            product_ref = str(
                row[import_fields.index(BOM_COMPONENT_PRODUCT_FIELD)] or ""
            ).strip()
            if product_ref:
                components.append((
                    product_ref,
                    self._component_quantity(import_fields, row, BOM_COMPONENT_QTY_FIELD),
                    self._component_uom(import_fields, row, BOM_COMPONENT_UOM_FIELD),
                ))

        for slot_number in range(1, BOM_IMPORT_COMPONENT_SLOT_COUNT + 1):
            product_field = f"x_import_bom_component_product_{slot_number}"
            qty_field = f"x_import_bom_component_qty_{slot_number}"
            uom_field = f"x_import_bom_component_uom_{slot_number}"
            if product_field not in import_fields:
                continue
            product_ref = str(row[import_fields.index(product_field)] or "").strip()
            if product_ref:
                components.append((
                    product_ref,
                    self._component_quantity(import_fields, row, qty_field),
                    self._component_uom(import_fields, row, uom_field),
                ))
        return components

    @staticmethod
    def _component_quantity(import_fields, row, qty_field):
        if qty_field not in import_fields:
            return 1.0
        return float(row[import_fields.index(qty_field)] or 0.0) or 1.0

    @staticmethod
    def _component_uom(import_fields, row, uom_field):
        if uom_field not in import_fields:
            return ""
        return str(row[import_fields.index(uom_field)] or "").strip()

    def _write_imported_bom_components(self, component_values):
        commands = [Command.clear()]
        missing_products = []
        missing_uoms = []
        for product_ref, quantity, uom_ref in component_values:
            product = self._find_bom_product(product_ref)
            if not product:
                missing_products.append(product_ref)
                continue
            uom = self._find_bom_uom(uom_ref) if uom_ref else product.uom_id
            if not uom:
                missing_uoms.append(uom_ref)
                continue
            commands.append(Command.create({
                "product_id": product.id,
                "product_qty": quantity,
                "product_uom_id": uom.id,
            }))
        if missing_products or missing_uoms:
            messages = []
            if missing_products:
                messages.append(_("找不到以下BOM组件产品：%s") % ", ".join(sorted(set(missing_products))))
            if missing_uoms:
                messages.append(_("找不到以下BOM组件单位：%s") % ", ".join(sorted(set(missing_uoms))))
            raise UserError("\n".join(messages))
        self.write({"bom_line_ids": commands})

    def _find_bom_product(self, product_ref):
        Product = self.env["product.product"]
        product = Product.search([("default_code", "=", product_ref)], limit=1)
        if product:
            return product
        matches = Product.name_search(name=product_ref, operator="=", limit=1)
        return Product.browse(matches[0][0]) if matches else Product

    def _find_bom_uom(self, uom_ref):
        if not uom_ref:
            return self.env["uom.uom"]
        Uom = self.env["uom.uom"]
        uom = Uom.search([("name", "=", uom_ref)], limit=1)
        if uom:
            return uom
        matches = Uom.name_search(name=uom_ref, operator="=", limit=1)
        return Uom.browse(matches[0][0]) if matches else Uom

    def _get_bom_import_template_columns(self):
        return [
            ("product_tmpl_id", "成品模板"),
            ("product_id", "成品变体"),
            ("product_qty", "成品数量"),
            ("product_uom_id", "成品单位"),
            ("type", "清单类型"),
            ("code", "物料清单编号"),
            ("company_id", "公司"),
            (BOM_COMPONENT_PRODUCT_FIELD, "组件产品"),
            (BOM_COMPONENT_QTY_FIELD, "组件数量"),
            (BOM_COMPONENT_UOM_FIELD, "组件单位"),
        ]

    def _generate_bom_import_template_xlsx(self):
        workbook = self._create_bom_workbook(include_records=False)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def action_export_import_template_format(self):
        ids = ",".join(str(record_id) for record_id in self.ids)
        return {
            "type": "ir.actions.act_url",
            "url": f"/stock_subwarehouse_hierarchy/export/mrp_bom.xlsx?{urlencode({'ids': ids})}",
            "target": "self",
        }

    def _generate_bom_export_xlsx(self):
        workbook = self._create_bom_workbook(include_records=True)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _create_bom_workbook(self, include_records=False):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl.workbook.defined_name import DefinedName
            from openpyxl.worksheet.datavalidation import DataValidation
        except ImportError as error:
            raise ImportError(_("生成模板需要安装 openpyxl。")) from error

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "物料清单导入" if not include_records else "物料清单导出"
        columns = self._get_bom_import_template_columns()
        sheet.append([field_name for field_name, _label in columns])
        sheet.append([label for _field_name, label in columns])

        if include_records:
            for bom in self:
                for row in bom._get_bom_export_rows(columns):
                    sheet.append(row)
        else:
            sheet.append([
                "示例成品", "", 1, "件", "normal", "BOM-EXAMPLE-001",
                self.env.company.display_name, "COMPONENT-001", 2, "件",
            ])
            sheet.append(["", "", "", "", "", "", "", "COMPONENT-002", 1, "件"])

        product_sheet = workbook.create_sheet("产品列表")
        product_sheet.append(["产品ID", "产品名称", "内部编号", "物料类型", "单位"])
        for product in self.env["product.product"].search([], order="default_code, name, id"):
            product_sheet.append([
                product.id,
                product.display_name,
                product.default_code or "",
                product.product_tmpl_id.x_material_type or "",
                product.uom_id.display_name,
            ])

        component_range = f"'产品列表'!$C$2:$C${max(product_sheet.max_row, 2)}"
        workbook.defined_names.add(DefinedName("bom_component_refs", attr_text=component_range))
        validation = DataValidation(type="list", formula1="=bom_component_refs", allow_blank=True)
        validation.error = "请选择产品列表中的内部编号。"
        validation.errorTitle = "组件产品无效"
        validation.prompt = "可从下拉列表选择，也可输入现有产品内部编号。"
        validation.promptTitle = "选择组件产品"
        sheet.add_data_validation(validation)
        component_column = next(
            index + 1
            for index, (field_name, _label) in enumerate(columns)
            if field_name == BOM_COMPONENT_PRODUCT_FIELD
        )
        column_letter = get_column_letter(component_column)
        validation.add(f"{column_letter}3:{column_letter}5000")

        field_sheet = workbook.create_sheet("导入字段")
        field_sheet.append(["字段", "中文说明"])
        for field_name, label in columns:
            field_sheet.append([field_name, label])

        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A3" if worksheet == sheet else "A2"
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
            if worksheet == sheet and worksheet.max_row >= 2:
                for cell in worksheet[2]:
                    cell.font = Font(italic=True)
                    cell.fill = PatternFill("solid", fgColor="E2F0D9")
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 45)
        return workbook

    def _get_bom_export_rows(self, columns):
        parent_values = {
            "product_tmpl_id": self.product_tmpl_id.display_name,
            "product_id": self.product_id.display_name if self.product_id else "",
            "product_qty": self.product_qty,
            "product_uom_id": self.product_uom_id.display_name,
            "type": self.type,
            "code": self.code or "",
            "company_id": self.company_id.display_name if self.company_id else "",
        }
        rows = []
        for line_index, line in enumerate(list(self.bom_line_ids) or [False]):
            values = dict(parent_values) if line_index == 0 else {}
            if line:
                values.update({
                    BOM_COMPONENT_PRODUCT_FIELD: line.product_id.default_code or line.product_id.display_name,
                    BOM_COMPONENT_QTY_FIELD: line.product_qty,
                    BOM_COMPONENT_UOM_FIELD: line.product_uom_id.display_name,
                })
            rows.append([values.get(field_name, "") for field_name, _label in columns])
        return rows

    def action_apply_reusable_template(self):
        for bom in self:
            template = bom.x_reusable_template_id
            if not template:
                raise UserError(_("请先选择一个可复用BOM模板。"))
            bom.bom_line_ids = [Command.clear()] + [
                Command.create({
                    "product_id": line.product_id.id,
                    "product_qty": line.product_qty,
                    "product_uom_id": line.product_uom_id.id,
                })
                for line in template.line_ids
            ]
        return True

    def action_save_as_reusable_template(self):
        self.ensure_one()
        if not self.bom_line_ids:
            raise UserError(_("当前物料清单没有组件，无法保存为模板。"))
        template = self.env["stock.subwarehouse.bom.template"].create({
            "name": _("%s BOM模板") % (self.code or self.product_tmpl_id.display_name),
            "code": self.code or False,
            "company_id": self.company_id.id or self.env.company.id,
            "line_ids": [
                Command.create({
                    "sequence": line.sequence,
                    "product_id": line.product_id.id,
                    "product_qty": line.product_qty,
                    "product_uom_id": line.product_uom_id.id,
                })
                for line in self.bom_line_ids
            ],
        })
        self.x_reusable_template_id = template
        return {
            "type": "ir.actions.act_window",
            "name": _("可复用BOM模板"),
            "res_model": "stock.subwarehouse.bom.template",
            "res_id": template.id,
            "view_mode": "form",
            "target": "current",
        }
