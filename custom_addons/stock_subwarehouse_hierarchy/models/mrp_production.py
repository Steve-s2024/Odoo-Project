from io import BytesIO
from urllib.parse import urlencode

from odoo import _, api, fields, models


IMPORT_TEMPLATE_ROUTE = "/stock_subwarehouse_hierarchy/import_template/mrp_production.xlsx"


class MrpProduction(models.Model):
    _inherit = "mrp.production"

    @api.model
    def load(self, import_fields, import_data):
        filtered = self._filter_unique_mrp_import_rows(import_fields, import_data)
        result = (
            super().load(filtered["fields"], filtered["data"])
            if filtered["data"]
            else {"ids": [], "messages": [], "nextrow": 0}
        )
        result["x_business_import_failures"] = filtered["failures"]
        if filtered["applied"]:
            result["nextrow"] = (
                filtered["source_window_size"]
                if filtered["has_more_source_rows"]
                else 0
            )
            self._record_mrp_import_results(result, filtered)
        return result

    @api.model
    def _filter_unique_mrp_import_rows(self, import_fields, import_data):
        result = {
            "fields": import_fields,
            "data": import_data,
            "kept_source_indexes": list(range(len(import_data))),
            "accepted_orders": [],
            "failures": [],
            "applied": False,
            "source_window_size": len(import_data),
            "has_more_source_rows": False,
        }
        if not self.env.context.get("import_file"):
            return result

        import_limit = self.env.context.get("_import_limit")
        window_size = min(len(import_data), import_limit) if import_limit else len(import_data)
        source_data = import_data[:window_size]
        result.update({
            "data": source_data,
            "kept_source_indexes": list(range(window_size)),
            "applied": True,
            "source_window_size": window_size,
            "has_more_source_rows": window_size < len(import_data),
        })
        source_offset = self.env.context.get("business_import_source_offset", 0)
        header_offset = 2 if self.env.context.get("business_import_has_headers") else 1
        if "name" not in import_fields:
            result["data"] = []
            result["kept_source_indexes"] = []
            result["failures"] = [
                {
                    "source_row": source_offset + index + header_offset,
                    "identifier": "",
                    "reason": "必须映射制造单号（name）字段。",
                }
                for index, _row in enumerate(source_data)
            ]
            return result

        name_index = import_fields.index("name")
        self.env.cr.execute(
            "SELECT pg_advisory_xact_lock(hashtext(%s))",
            ["stock_subwarehouse_hierarchy.mrp_name_import"],
        )
        self.flush_model(["name"])
        self.env.cr.execute("""
            SELECT lower(btrim(name))
              FROM mrp_production
             WHERE name IS NOT NULL AND btrim(name) != ''
        """)
        existing = {row[0] for row in self.env.cr.fetchall()}
        accepted = set()
        prepared_fields, continuation_field_map = self._prepare_grouped_mrp_import_fields(
            import_fields
        )
        groups = []
        current_group = []
        for source_index, row in enumerate(source_data):
            starts_new_order = not self._is_empty_mrp_import_value(row[name_index])
            if current_group and starts_new_order:
                groups.append(current_group)
                current_group = []
            current_group.append((source_index, row))
        if current_group:
            groups.append(current_group)

        kept_rows = []
        kept_indexes = []
        accepted_orders = []
        failures = []
        for group in groups:
            first_source_index, first_row = group[0]
            display_name = str(first_row[name_index] or "").strip()
            normalized_name = display_name.casefold()
            if not normalized_name:
                reason = "制造单号不能为空。"
            elif normalized_name in existing:
                reason = "制造单号已存在于 ERP。"
            elif normalized_name in accepted:
                reason = "制造单号在本次导入文件中重复。"
            else:
                accepted.add(normalized_name)
                for group_index, (source_index, row) in enumerate(group):
                    prepared_row = list(row) + [""] * (len(prepared_fields) - len(row))
                    if group_index:
                        for source_field_index, target_field_index in continuation_field_map.items():
                            if self._is_empty_mrp_import_value(prepared_row[target_field_index]):
                                prepared_row[target_field_index] = prepared_row[source_field_index]
                        # Odoo starts a new parent record whenever any parent
                        # value is populated.  A continuation row represents an
                        # additional finished product, so only its nested
                        # by-product values may remain populated.
                        for field_index, field_name in enumerate(prepared_fields):
                            if not field_name.startswith("move_byproduct_ids/"):
                                prepared_row[field_index] = ""
                    kept_rows.append(prepared_row)
                    kept_indexes.append(source_index)
                accepted_orders.append({
                    "source_index": first_source_index,
                    "identifier": display_name,
                })
                continue
            failures.append({
                "source_row": source_offset + first_source_index + header_offset,
                "identifier": display_name,
                "reason": reason,
            })
        result.update({
            "fields": prepared_fields,
            "data": kept_rows,
            "kept_source_indexes": kept_indexes,
            "accepted_orders": accepted_orders,
            "failures": failures,
        })
        return result

    @api.model
    def _prepare_grouped_mrp_import_fields(self, import_fields):
        """Add nested finished-product fields used by continuation rows.

        The spreadsheet deliberately keeps the familiar flat columns
        ``product_id``, ``product_qty`` and ``product_uom_id`` on every product
        row.  Only the first row is the MO's main product.  Values on following
        blank-name rows are redirected to ``move_byproduct_ids`` before Odoo's
        standard importer runs.
        """
        prepared_fields = list(import_fields)
        continuation_field_map = {}
        field_targets = []
        for source_field_index, field_name in enumerate(import_fields):
            if field_name == "product_qty":
                target_field = "move_byproduct_ids/product_uom_qty"
            elif field_name == "product_id" or field_name.startswith("product_id/"):
                target_field = (
                    "move_byproduct_ids/product_id" + field_name[len("product_id"):]
                )
            elif field_name == "product_uom_id" or field_name.startswith("product_uom_id/"):
                target_field = (
                    "move_byproduct_ids/product_uom"
                    + field_name[len("product_uom_id"):]
                )
            else:
                continue
            field_targets.append((source_field_index, target_field))

        for source_field_index, target_field in field_targets:
            if target_field not in prepared_fields:
                prepared_fields.append(target_field)
            continuation_field_map[source_field_index] = prepared_fields.index(target_field)
        return prepared_fields, continuation_field_map

    @api.model
    def _is_empty_mrp_import_value(self, value):
        return value is None or (isinstance(value, str) and not value.strip())

    @api.model
    def _record_mrp_import_results(self, result, filtered):
        batch_id = self.env.context.get("business_import_result_batch_id")
        if not batch_id:
            return
        source_offset = self.env.context.get("business_import_source_offset", 0)
        header_offset = 2 if self.env.context.get("business_import_has_headers") else 1
        values = []
        for accepted, record_id in zip(filtered["accepted_orders"], result.get("ids") or []):
            values.append({
                "batch_id": batch_id,
                "source_row": source_offset + accepted["source_index"] + header_offset,
                "status": "success",
                "identifier": accepted["identifier"],
                "record_ref": f"mrp.production,{record_id}",
                "reason": "导入成功。",
            })
        values.extend({
            "batch_id": batch_id,
            "source_row": failure["source_row"],
            "status": "failed",
            "identifier": failure["identifier"],
            "reason": failure["reason"],
        } for failure in filtered["failures"])
        if values:
            self.env["stock.subwarehouse.business.import.result.line"].create(values)

    def _get_subwarehouse_manufacturing_location(self):
        location_id = self.env.context.get("subwarehouse_manufacturing_location_id")
        if not location_id:
            return self.env["stock.location"]
        location = self.env["stock.location"].browse(location_id).exists()
        return location if location.usage == "internal" else self.env["stock.location"]

    def default_get(self, fields_list):
        defaults = super().default_get(fields_list)
        location = self._get_subwarehouse_manufacturing_location()
        if location:
            if "location_src_id" in fields_list:
                defaults["location_src_id"] = location.id
            if "location_dest_id" in fields_list:
                defaults["location_dest_id"] = location.id
        return defaults

    @api.depends("picking_type_id")
    def _compute_locations(self):
        super()._compute_locations()
        location = self._get_subwarehouse_manufacturing_location()
        if location:
            for production in self:
                production.location_src_id = location
                production.location_dest_id = location

    def _get_default_bom_for_product(self, product, picking_type=False, company_id=False):
        if not product:
            return self.env["mrp.bom"]
        domain = [
            ("type", "=", "normal"),
            ("product_tmpl_id", "=", product.product_tmpl_id.id),
            "|",
            ("product_id", "=", product.id),
            ("product_id", "=", False),
            "|",
            ("company_id", "=", company_id or self.env.company.id),
            ("company_id", "=", False),
        ]
        if picking_type:
            domain += [
                "|",
                ("picking_type_id", "=", picking_type.id),
                ("picking_type_id", "=", False),
            ]
        candidate_boms = self.env["mrp.bom"].with_context(active_test=True).search(
            domain,
            order="write_date desc, id desc",
        )
        product_boms = candidate_boms.filtered(lambda bom: bom.product_id == product)
        template_boms = candidate_boms.filtered(lambda bom: not bom.product_id)
        for boms in (
            product_boms.filtered("bom_line_ids"),
            template_boms.filtered("bom_line_ids"),
            product_boms,
            template_boms,
        ):
            if boms:
                return boms[0]
        boms_by_product = self.env["mrp.bom"].with_context(active_test=True)._bom_find(
            product,
            picking_type=picking_type,
            company_id=company_id or self.env.company.id,
            bom_type="normal",
        )
        return boms_by_product.get(product, self.env["mrp.bom"])

    @api.onchange("product_id", "picking_type_id", "company_id")
    def _onchange_product_id_use_default_bom_components(self):
        for production in self:
            if not production.product_id:
                continue
            bom = production._get_default_bom_for_product(
                production.product_id,
                picking_type=production.picking_type_id,
                company_id=production.company_id.id or self.env.company.id,
            )
            if bom and (
                not production.bom_id
                or production.bom_id.product_tmpl_id != production.product_tmpl_id
                or (not production.bom_id.bom_line_ids and bom.bom_line_ids)
                or (production.bom_id.product_id and production.bom_id.product_id != production.product_id)
            ):
                production.bom_id = bom
            if production.bom_id:
                production._compute_move_raw_ids()

    @api.model_create_multi
    def create(self, vals_list):
        location = self._get_subwarehouse_manufacturing_location()
        for vals in vals_list:
            if vals.get("product_id") and not vals.get("bom_id"):
                product = self.env["product.product"].browse(vals["product_id"])
                picking_type = (
                    vals.get("picking_type_id")
                    and self.env["stock.picking.type"].browse(vals["picking_type_id"])
                )
                bom = self._get_default_bom_for_product(
                    product,
                    picking_type=picking_type,
                    company_id=vals.get("company_id") or self.env.company.id,
                )
                if bom:
                    vals["bom_id"] = bom.id
            if location:
                vals.setdefault("location_src_id", location.id)
                vals.setdefault("location_dest_id", location.id)
        return super().create(vals_list)

    @api.model
    def get_import_templates(self):
        return [{
            "label": _("制造单导入模板（当前产品属性）"),
            "template": IMPORT_TEMPLATE_ROUTE,
        }]

    @api.model
    def _get_dynamic_import_template_columns(self):
        return [
            ("name", _("制造单号")),
            ("product_id", _("产品")),
            ("product_qty", _("数量")),
            ("product_uom_id", _("计量单位")),
            ("bom_id", _("物料清单")),
            ("origin", _("源单据")),
            ("date_start", _("计划日期")),
            ("location_src_id", _("组件库位")),
            ("location_dest_id", _("成品库位")),
            ("picking_type_id", _("作业类型")),
            ("company_id", _("公司")),
            ("never_product_template_attribute_value_ids", _("产品属性值")),
        ]

    @api.model
    def _generate_dynamic_import_template_xlsx(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as error:
            raise ImportError(_("生成导入模板需要安装 openpyxl。")) from error

        workbook = Workbook()
        import_sheet = workbook.active
        import_sheet.title = "制造单导入"
        field_columns = self._get_dynamic_import_template_columns()
        headers = [field_name for field_name, _label in field_columns]
        import_sheet.append(headers)
        import_sheet.append([label for _field_name, label in field_columns])
        import_sheet.append([
            "MO-IMPORT-001",
            "产品显示名称或外部 ID",
            1,
            "件",
            "",
            "示例制造单",
            fields.Datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "WH/Stock",
            "WH/Stock",
            "制造",
            self.env.company.display_name,
            "如需要，使用逗号分隔产品属性值",
        ])
        import_sheet.append([
            "",
            "同一制造单的第二个产品（制造单号留空）",
            1,
            "件",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ])

        attribute_sheet = workbook.create_sheet("产品属性")
        attribute_sheet.append(["属性ID", "属性", "变体创建方式", "值ID", "值"])
        for attribute in self.env["product.attribute"].search([], order="sequence, id"):
            if attribute.value_ids:
                for value in attribute.value_ids.sorted(lambda record: (record.sequence, record.id)):
                    attribute_sheet.append([
                        attribute.id,
                        attribute.display_name,
                        attribute.create_variant,
                        value.id,
                        value.display_name,
                    ])
            else:
                attribute_sheet.append([
                    attribute.id,
                    attribute.display_name,
                    attribute.create_variant,
                    "",
                    "",
                ])

        field_sheet = workbook.create_sheet("导入字段")
        field_sheet.append(["字段", "标签", "类型", "关联模型"])
        fields_get = self.fields_get([field_name for field_name, _label in field_columns])
        for field_name, label in field_columns:
            metadata = fields_get.get(field_name, {})
            field_sheet.append([
                field_name,
                metadata.get("string") or label,
                metadata.get("type", ""),
                metadata.get("relation", ""),
            ])

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A3" if sheet == import_sheet else "A2"
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
            if sheet == import_sheet and sheet.max_row >= 2:
                for cell in sheet[2]:
                    cell.font = Font(italic=True)
                    cell.fill = PatternFill("solid", fgColor="E2F0D9")
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 45)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def action_export_import_template_format(self):
        ids = ",".join(str(record_id) for record_id in self.ids)
        return {
            "type": "ir.actions.act_url",
            "url": f"/stock_subwarehouse_hierarchy/export/mrp_production.xlsx?{urlencode({'ids': ids})}",
            "target": "self",
        }

    def _generate_dynamic_export_xlsx(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as error:
            raise ImportError(_("生成导出文件需要安装 openpyxl。")) from error

        workbook = Workbook()
        export_sheet = workbook.active
        export_sheet.title = "制造单导出"
        field_columns = self._get_dynamic_import_template_columns()
        export_sheet.append([field_name for field_name, _label in field_columns])
        export_sheet.append([label for _field_name, label in field_columns])
        for production in self:
            for row in production._get_dynamic_export_rows(field_columns):
                export_sheet.append(row)

        attribute_sheet = workbook.create_sheet("产品属性")
        attribute_sheet.append(["属性ID", "属性", "变体创建方式", "值ID", "值"])
        for attribute in self.env["product.attribute"].search([], order="sequence, id"):
            if attribute.value_ids:
                for value in attribute.value_ids.sorted(lambda record: (record.sequence, record.id)):
                    attribute_sheet.append([
                        attribute.id,
                        attribute.display_name,
                        attribute.create_variant,
                        value.id,
                        value.display_name,
                    ])
            else:
                attribute_sheet.append([
                    attribute.id,
                    attribute.display_name,
                    attribute.create_variant,
                    "",
                    "",
                ])

        field_sheet = workbook.create_sheet("导出字段")
        field_sheet.append(["字段", "标签", "类型", "关联模型"])
        fields_get = self.fields_get([field_name for field_name, _label in field_columns])
        for field_name, label in field_columns:
            metadata = fields_get.get(field_name, {})
            field_sheet.append([
                field_name,
                metadata.get("string") or label,
                metadata.get("type", ""),
                metadata.get("relation", ""),
            ])

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A3" if sheet == export_sheet else "A2"
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
            if sheet == export_sheet and sheet.max_row >= 2:
                for cell in sheet[2]:
                    cell.font = Font(italic=True)
                    cell.fill = PatternFill("solid", fgColor="E2F0D9")
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 45)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _get_dynamic_export_row(self, field_columns):
        return [self._get_mrp_export_value(field_name) for field_name, _label in field_columns]

    def _get_dynamic_export_rows(self, field_columns):
        self.ensure_one()
        rows = [self._get_dynamic_export_row(field_columns)]
        for move in self.move_byproduct_ids.sorted("id"):
            row = []
            for field_name, _label in field_columns:
                if field_name == "product_id":
                    row.append(move.product_id.default_code or move.product_id.display_name)
                elif field_name == "product_qty":
                    row.append(move.product_uom_qty)
                elif field_name == "product_uom_id":
                    row.append(move.product_uom.display_name)
                else:
                    row.append("")
            rows.append(row)
        return rows

    def _get_mrp_export_value(self, field_name):
        if field_name == "product_id":
            return self.product_id.default_code or self.product_id.display_name
        if field_name == "never_product_template_attribute_value_ids":
            return ", ".join(self.never_product_template_attribute_value_ids.mapped("display_name"))
        if field_name not in self._fields:
            return ""
        value = self[field_name]
        field = self._fields[field_name]
        if field.type == "many2one":
            return value.display_name if value else ""
        if field.type in ("many2many", "one2many"):
            return ", ".join(value.mapped("display_name"))
        if field.type == "datetime" and value:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if field.type == "date" and value:
            return value.strftime("%Y-%m-%d")
        return value
