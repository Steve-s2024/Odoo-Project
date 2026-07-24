import base64
from io import BytesIO

from odoo import _, fields, models
from odoo.exceptions import UserError


class ProductInternationalMappingImportWizard(models.TransientModel):
    _name = "stock.subwarehouse.product.international.mapping.import.wizard"
    _description = "Import Product International Mapping"

    import_file = fields.Binary(string="\u4ef7\u683c\u8868\u6587\u4ef6", required=True)
    import_filename = fields.Char(string="\u6587\u4ef6\u540d")

    @staticmethod
    def _normalized_name(value):
        return "".join(str(value or "").split()).casefold()

    @staticmethod
    def _normalized_flex(value):
        normalized = ProductInternationalMappingImportWizard._normalized_name(value).replace(
            "\u786c\u5ea6", ""
        ).replace("flex", "")
        return "" if normalized in {"", "000", "\u65e0", "none", "noflex", "n/a", "notapplicable", "\u672a\u8bc6\u522b", "\u9ed8\u8ba4"} else normalized

    def action_import_mapping(self):
        self.ensure_one()
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise UserError(_("\u5bfc\u5165\u4ef7\u683c\u8868\u9700\u8981\u5b89\u88c5 openpyxl\u3002")) from error

        try:
            workbook = load_workbook(
                BytesIO(base64.b64decode(self.import_file)),
                read_only=True,
                data_only=True,
            )
        except Exception as error:
            raise UserError(_("\u65e0\u6cd5\u8bfb\u53d6Excel\u4ef7\u683c\u8868\uff0c\u8bf7\u4e0a\u4f20 .xlsx \u6587\u4ef6\u3002")) from error

        worksheet = workbook.active
        header_row = None
        headers = {}
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            normalized_cells = {
                self._normalized_name(value): index
                for index, value in enumerate(row)
                if value not in (None, "")
            }
            if (
                "\u540d\u79f0" in normalized_cells
                and "flex" in normalized_cells
                and "product" in normalized_cells
            ):
                header_row = row_number
                headers = normalized_cells
                break
            if row_number >= 10:
                break

        if header_row is None:
            raise UserError(_("\u672a\u627e\u5230\u4ef7\u683c\u8868\u6807\u9898\u884c\u3002\u6587\u4ef6\u5fc5\u987b\u5305\u542b\u201c\u540d\u79f0\u201d\u3001\u201cflex\u201d\u3001\u201cPRODUCT\u201d\u548c\u201cRETAIL PRICE (USD)\u201d\u5217\u3002"))

        price_index = next(
            (index for name, index in headers.items() if "retailprice" in name and "usd" in name),
            None,
        )
        if price_index is None:
            raise UserError(_("\u672a\u627e\u5230\u201cRETAIL PRICE (USD)\u201d\u4ef7\u683c\u5217\u3002"))

        name_index = headers["\u540d\u79f0"]
        flex_index = headers["flex"]
        english_name_index = headers["product"]
        Product = self.env["product.template"]
        products_by_key = {}
        for product in Product.search([]):
            key = (
                self._normalized_name(product.name),
                self._normalized_flex(product._get_website_mapping_flex()),
            )
            products_by_key[key] = products_by_key.get(key, Product) | product

        updated_products = Product
        unmatched_names = []
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            chinese_name = row[name_index] if len(row) > name_index else False
            flex = row[flex_index] if len(row) > flex_index else False
            english_name = row[english_name_index] if len(row) > english_name_index else False
            price = row[price_index] if len(row) > price_index else False
            mapping_key = (self._normalized_name(chinese_name), self._normalized_flex(flex))
            if not mapping_key[0]:
                continue
            products = products_by_key.get(mapping_key, Product)
            if not products:
                unmatched_names.append(
                    "%s / %s" % (str(chinese_name).strip(), str(flex or "").strip())
                )
                continue
            try:
                usd_price = float(price)
            except (TypeError, ValueError):
                raise UserError(_("\u4ea7\u54c1\u201c%s\u201d\u7684\u7f8e\u5143\u4ef7\u683c\u4e0d\u662f\u6709\u6548\u6570\u5b57\u3002") % chinese_name)
            products.write({
                "x_website_english_name": str(english_name or "").strip(),
                "x_website_mapping_flex": str(flex or "").strip(),
                "x_website_usd_price": usd_price,
            })
            updated_products |= products

        message = _("\u5df2\u66f4\u65b0 %s \u4e2a\u4ea7\u54c1\u7684\u82f1\u6587\u540d\u79f0\u548c\u7f8e\u5143\u4ef7\u683c\u3002") % len(updated_products)
        if unmatched_names:
            message += "\n" + _("\u672a\u5339\u914d %s \u4e2a\u4e2d\u6587\u540d\u79f0\uff1a%s") % (
                len(unmatched_names), ", ".join(unmatched_names[:12]),
            )
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("\u82f1\u6587\u540d\u79f0\u4e0e\u7f8e\u5143\u4ef7\u683c\u5bfc\u5165\u5b8c\u6210"),
                "message": message,
                "type": "warning" if unmatched_names else "success",
                "sticky": bool(unmatched_names),
            },
        }
