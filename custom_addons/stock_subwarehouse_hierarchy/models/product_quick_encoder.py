import base64
import re
import unicodedata
from collections import Counter
from io import BytesIO

from odoo import _, api, fields, models
from odoo.exceptions import UserError, ValidationError

from .product_template import _ProductCodeTrie


ENCODER_INPUT_COLUMNS = [
    ("product_name", "产品名称"),
    ("manufacturer", "厂家代码/名称"),
    ("production_yymm", "生产年月(YYMM)"),
    ("product_type", "产品类型代码/名称"),
    ("finished_type", "成品类型(M/F)"),
    ("audience", "成人儿童(A/K)"),
    ("flex", "硬度"),
    ("color", "颜色"),
    ("size", "尺码"),
]

MANUFACTURER_CODES = {
    "丰昂": "01", "鼎宏": "02", "嘉瑞": "03", "立晟": "04",
    "猛犸": "05", "曼琳": "06", "山峰与海": "07", "鹏亿发": "08",
    "航瑞": "09", "奥博": "10", "凡夫": "11", "魔诺": "12",
    "闪亮": "13", "起点": "13", "米泰": "14", "熙堃": "15",
}

MANUFACTURER_NAMES = {
    "01": "丰昂", "02": "鼎宏", "03": "嘉瑞", "04": "立晟",
    "05": "猛犸", "06": "曼琳", "07": "山峰与海", "08": "鹏亿发",
    "09": "航瑞", "10": "奥博", "11": "凡夫", "12": "魔诺",
    "13": "闪亮/起点", "14": "米泰", "15": "熙堃",
}

PRODUCT_TYPE_CODES = {
    "单板鞋": "S1", "单板滑雪鞋": "S1", "双板鞋": "S2", "双板滑雪鞋": "S2",
    "快穿单板鞋": "SE", "4+2双板鞋": "S", "加厚内靴": "L",
    "单板": "D", "全能板": "D", "追逐板": "Dz", "追逐版": "Dz", "双板": "X",
    "单板包": "B1", "双板包": "B", "纯色上衣": "Y", "混色上衣": "Yx",
    "单板裤": "Yb1", "单板裤子": "Yb1", "双板裤": "Yb", "双板裤子": "Yb",
    "雪杖": "Z", "雪仗": "Z", "无背板固定器": "N", "快穿固定器": "E",
    "无框雪镜": "G", "翻盖雪镜": "G1", "全脸护脸": "Pa", "半脸护脸": "Pb",
    "五指手套": "T5", "焖子手套": "T1", "连指手套": "T1", "头盔": "H",
    "HD护臀": "K1", "PU护臀": "K2", "护膝1": "KK1", "护膝2": "KK2",
    "雪袜": "W", "鞋垫": "Q", "小防冻贴": "F1", "大防冻贴": "F2",
}

PRODUCT_TYPE_NAMES = {
    "S1": "单板鞋", "S2": "3+3双板鞋", "SE": "快穿单板鞋", "S": "4+2双板鞋",
    "L": "加厚内靴", "D": "全能板", "Dz": "追逐版", "X": "双板",
    "B1": "单板包", "B": "双板包", "Y": "纯色上衣", "Yx": "混色上衣",
    "Yb1": "单板裤子", "Yb": "双板裤子", "Z": "雪仗",
    "N": "无背板固定器", "E": "快穿固定器", "G": "无框雪镜",
    "G1": "翻盖雪镜", "Pa": "全脸护脸", "Pb": "半脸护脸",
    "T5": "五指手套", "T1": "焖子手套", "H": "头盔",
    "K1": "HD护臀", "K2": "PU护臀", "KK1": "护膝1", "KK2": "护膝2",
    "W": "雪袜", "Q": "鞋垫", "F1": "小防冻贴", "F2": "大防冻贴",
}

# Aliases retain the exact entries from the encoding-rules workbook and add
# established ERP/website wording so imports do not depend on one spelling.
PRODUCT_TYPE_CODES.update({
    "3+3双板鞋": "S2", "4+2双板鞋": "S", "快穿雪鞋": "SE",
    "无背板": "N", "无框款": "G", "翻盖款": "G1",
    "全脸": "Pa", "半脸": "Pb", "五指": "T5", "焖子": "T1",
    "小防冻贴": "F1", "大防冻贴": "F2",
})

COLOR_TOKENS = {
    "蓝": "B", "黑": "H", "白": "W", "红": "R", "绿": "G", "灰": "A",
    "粉": "P", "黄": "Y", "橙": "O", "棕": "Z", "青": "CY", "彩": "X",
    "紫": "U", "玫红": "Q",
}

COLOR_NAMES = {
    "B": "蓝", "H": "黑", "W": "白", "R": "红", "G": "绿", "A": "灰",
    "P": "粉", "Y": "黄", "O": "橙", "Z": "棕", "CY": "青", "X": "彩",
    "U": "紫", "Q": "玫红",
}

HEADER_ALIASES = {
    "产品名称": "product_name", "name": "product_name", "product_name": "product_name",
    "厂家": "manufacturer", "厂家代码": "manufacturer", "厂家代码/名称": "manufacturer",
    "manufacturer": "manufacturer",
    "生产年月": "production_yymm", "生产年月(yymm)": "production_yymm",
    "production_yymm": "production_yymm",
    "产品类型": "product_type", "产品类型代码": "product_type", "产品类型代码/名称": "product_type",
    "type": "product_type", "product_type": "product_type",
    "成品类型": "finished_type", "成品类型(m/f)": "finished_type", "finished_type": "finished_type",
    "成人儿童": "audience", "成人儿童(a/k)": "audience", "audience": "audience",
    "硬度": "flex", "flex": "flex", "颜色": "color", "colour": "color", "color": "color",
    "尺码": "size", "size": "size",
}


def _clean(value):
    # Odoo returns False for an empty Char field. Treat it like an empty Excel
    # cell instead of turning it into the literal string "False".
    if value is None or value is False:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _match_text(value):
    value = unicodedata.normalize("NFKC", _clean(value)).casefold()
    return re.sub(r"[^0-9a-z\u4e00-\u9fff+]", "", value)


def _common_character_score(query, candidate):
    """Rank names primarily by the number of matching characters.

    Coverage, subsequence order, and length difference only break ties. This
    follows the user's requested matching rule while keeping ambiguous names
    (for example, 滑雪手套) blocked instead of guessed.
    """
    query = _match_text(query)
    candidate = _match_text(candidate)
    if not query or not candidate:
        return None
    common = sum((Counter(query) & Counter(candidate)).values())
    if not common:
        return None
    query_coverage = common / len(query)
    candidate_coverage = common / len(candidate)
    if query_coverage + candidate_coverage < 1.0:
        return None
    subsequence = 0
    candidate_index = 0
    for character in query:
        found_at = candidate.find(character, candidate_index)
        if found_at >= 0:
            subsequence += 1
            candidate_index = found_at + 1
    containment = int(query in candidate or candidate in query)
    return (
        common,
        containment,
        round((query_coverage + candidate_coverage) * 1000),
        subsequence,
        -abs(len(query) - len(candidate)),
    )


def _match_named_rule(value, aliases, canonical_names, valid_codes=None):
    """Return a safe exact/fuzzy match without guessing tied rule codes."""
    original = _clean(value)
    folded = _match_text(original)
    codes = valid_codes or canonical_names.keys()
    direct_codes = {_match_text(code): code for code in codes}
    rental_requested = folded.endswith("r")
    base_folded = folded[:-1] if rental_requested else folded
    if base_folded in direct_codes:
        code = direct_codes[base_folded]
        return code, canonical_names[code], "code", ""

    exact_codes = {
        code for alias, code in aliases.items() if _match_text(alias) == folded
    }
    if len(exact_codes) == 1:
        code = exact_codes.pop()
        return code, canonical_names[code], "exact", ""
    if len(exact_codes) > 1:
        names = "、".join(f"{canonical_names[code]}({code})" for code in sorted(exact_codes))
        return "", "", "", f"匹配结果不唯一：{names}"

    best_by_code = {}
    best_alias_by_code = {}
    for alias, code in aliases.items():
        score = _common_character_score(original, alias)
        if score is not None and (code not in best_by_code or score > best_by_code[code]):
            best_by_code[code] = score
            best_alias_by_code[code] = alias
    if not best_by_code:
        return "", "", "", "未在编码规则中找到相似名称"
    best_score = max(best_by_code.values())
    winners = [code for code, score in best_by_code.items() if score == best_score]
    if len(winners) != 1:
        names = "、".join(f"{canonical_names[code]}({code})" for code in sorted(winners))
        return "", "", "", f"相似字符数相同，无法唯一匹配：{names}"
    code = winners[0]
    return code, canonical_names[code], "fuzzy", ""


def _match_display(original, canonical="", code="", method="", error=""):
    source = _clean(original) or "（空白）"
    if error:
        return f"{source} → 匹配失败：{error}"
    suffix = "【相似字符自动匹配】" if method == "fuzzy" else ""
    return f"{source} → {canonical}（{code}）{suffix}"


class ProductQuickEncoderWizard(models.TransientModel):
    _name = "stock.subwarehouse.product.quick.encoder.wizard"
    _description = "产品批量编码"

    import_file = fields.Binary(string="Excel 文件", attachment=False)
    import_filename = fields.Char(string="文件名")
    line_ids = fields.One2many(
        "stock.subwarehouse.product.quick.encoder.line",
        "wizard_id",
        string="编码明细",
        copy=False,
    )
    line_count = fields.Integer(string="总行数", compute="_compute_counts")
    ready_count = fields.Integer(string="可使用", compute="_compute_counts")
    warning_count = fields.Integer(string="需处理", compute="_compute_counts")
    failed_entries = fields.Text(string="失败条目", compute="_compute_counts")

    @api.depends("line_ids.status")
    def _compute_counts(self):
        for wizard in self:
            wizard.line_count = len(wizard.line_ids)
            wizard.ready_count = len(wizard.line_ids.filtered(lambda line: line.status == "ready"))
            wizard.warning_count = len(wizard.line_ids.filtered(lambda line: line.status in ("exists", "duplicate", "error")))
            failed = wizard.line_ids.filtered(lambda line: line.status == "error").sorted(
                lambda line: (line.sequence, line.id)
            )
            wizard.failed_entries = "\n".join(
                f"Excel 第 {line.source_row or '-'} 行："
                f"{line.product_name or '未命名产品'} — {line.remark}"
                for line in failed
            )

    def _reload_action(self):
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "name": _("批量编码"),
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    def action_download_template(self):
        return {
            "type": "ir.actions.act_url",
            "url": "/stock_subwarehouse_hierarchy/product_encoder/template.xlsx",
            "target": "self",
        }

    def action_import_excel(self):
        self.ensure_one()
        if not self.import_file:
            raise UserError(_("请先选择 Excel 文件。"))
        if self.import_filename and not self.import_filename.lower().endswith(".xlsx"):
            raise ValidationError(_("批量编码目前仅支持 .xlsx 文件。"))
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise UserError(_("读取 Excel 需要安装 openpyxl。")) from error

        try:
            workbook = load_workbook(BytesIO(base64.b64decode(self.import_file)), data_only=True)
        except Exception as error:
            raise ValidationError(_("无法读取 Excel 文件：%s") % error) from error
        sheet = workbook.active
        headers = [_clean(cell.value).casefold() for cell in sheet[1]]
        field_by_column = [HEADER_ALIASES.get(header) for header in headers]
        if not any(field_by_column):
            raise ValidationError(_("未找到可识别的表头。请先下载并使用批量编码模板。"))

        commands = [fields.Command.clear()]
        sequence = 10
        for source_row, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            values = {}
            for index, field_name in enumerate(field_by_column):
                if field_name and index < len(cells):
                    values[field_name] = _clean(cells[index])
            if not any(values.values()):
                continue
            # Preserve blanks: incomplete rows must be rejected visibly rather
            # than receiving hidden defaults during Excel import.
            finished_input = values.get("finished_type", "")
            audience_input = values.get("audience", "")
            values["finished_type"] = {
                "成品": "M", "非成品": "F", "组件": "F", "零件": "F",
            }.get(finished_input, finished_input.upper()) or False
            values["audience"] = {
                "成人": "A", "儿童": "K", "青少年": "K",
            }.get(audience_input, audience_input.upper()) or False
            values.update({"sequence": sequence, "source_row": source_row})
            commands.append(fields.Command.create(values))
            sequence += 10
        if len(commands) == 1:
            raise ValidationError(_("Excel 文件中没有可编码的数据行。"))
        self.line_ids = commands
        self.action_generate_codes()
        return self._reload_action()

    def action_add_line(self):
        self.ensure_one()
        next_sequence = (max(self.line_ids.mapped("sequence"), default=0) or 0) + 10
        self.env["stock.subwarehouse.product.quick.encoder.line"].create({
            "wizard_id": self.id,
            "sequence": next_sequence,
            "manufacturer": "15",
            "finished_type": "M",
            "audience": "A",
        })
        return self._reload_action()

    def action_generate_codes(self):
        self.ensure_one()
        if not self.line_ids:
            raise UserError(_("请先添加一行或导入 Excel 文件。"))
        self.env.cr.execute(
            "SELECT pg_advisory_xact_lock(hashtext(%s))",
            ["stock_subwarehouse_hierarchy.product_quick_encoder"],
        )
        self.env["product.product"].flush_model(["default_code"])
        self.env.cr.execute("""
            SELECT default_code
              FROM product_product
             WHERE default_code IS NOT NULL
               AND btrim(default_code) != ''
        """)
        existing_codes = _ProductCodeTrie(
            _clean(row[0]).casefold() for row in self.env.cr.fetchall()
        )
        generated_codes = _ProductCodeTrie()
        for line in self.line_ids.sorted(lambda item: (item.sequence, item.id)):
            code, error, matches = line._build_product_code()
            normalized = code.casefold()
            if error:
                status, remark = "error", error
                code = ""
            elif existing_codes.contains(normalized):
                status, remark = "exists", _("该产品编码已存在于 ERP；请确认是否为重复产品。")
            elif generated_codes.contains(normalized):
                status, remark = "duplicate", _("该产品编码在本批次中重复。")
            else:
                generated_codes.add(normalized)
                status, remark = "ready", ""
            line.write({
                "generated_code": code,
                "status": status,
                "remark": remark,
                **matches,
            })
        return self._reload_action()

    def action_download_result(self):
        self.ensure_one()
        self.action_generate_codes()
        return {
            "type": "ir.actions.act_url",
            "url": f"/stock_subwarehouse_hierarchy/product_encoder/result.xlsx?wizard_id={self.id}",
            "target": "self",
        }

    def _generate_result_xlsx(self):
        self.ensure_one()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as error:
            raise UserError(_("生成 Excel 需要安装 openpyxl。")) from error
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "批量编码结果"
        headers = [label for _field, label in ENCODER_INPUT_COLUMNS]
        headers += ["输入与规则匹配结果", "生成的产品编码", "状态", "备注", "原Excel行号"]
        sheet.append(headers)
        status_labels = dict(self.env["stock.subwarehouse.product.quick.encoder.line"]._fields["status"].selection)
        for line in self.line_ids.sorted(lambda item: (item.sequence, item.id)):
            sheet.append([
                line.product_name, line.manufacturer, line.production_yymm, line.product_type,
                line.finished_type, line.audience, line.flex, line.color, line.size,
                line.match_summary, line.generated_code, status_labels.get(line.status, line.status), line.remark,
                line.source_row or "",
            ])
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        for row in range(2, sheet.max_row + 1):
            status = sheet.cell(row, 12).value
            if status == "可使用":
                sheet.cell(row, 12).fill = PatternFill("solid", fgColor="C6EFCE")
            elif status:
                sheet.cell(row, 12).fill = PatternFill("solid", fgColor="FFC7CE")
        for column_cells in sheet.columns:
            width = min(max(max(len(_clean(cell.value)) for cell in column_cells) + 2, 12), 48)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()


class ProductQuickEncoderLine(models.TransientModel):
    _name = "stock.subwarehouse.product.quick.encoder.line"
    _description = "产品批量编码明细"
    _order = "sequence, id"

    wizard_id = fields.Many2one(
        "stock.subwarehouse.product.quick.encoder.wizard",
        required=True,
        ondelete="cascade",
    )
    sequence = fields.Integer(default=10)
    source_row = fields.Integer(string="原Excel行号", readonly=True)
    product_name = fields.Char(string="产品名称")
    manufacturer = fields.Char(string="厂家代码/名称", default="15")
    production_yymm = fields.Char(
        string="生产年月",
        default=lambda self: fields.Date.context_today(self).strftime("%y%m"),
        help="四位 YYMM，例如 2410。",
    )
    product_type = fields.Char(string="产品类型", help="可填 S1、T5 等代码，也可填单板滑雪鞋、五指手套等名称。")
    finished_type = fields.Selection([("M", "成品 M"), ("F", "非成品 F")], string="成品类型", default="M")
    audience = fields.Selection([("A", "成人 A"), ("K", "儿童/青少年 K")], string="成人/儿童", default="A")
    flex = fields.Char(string="硬度", help="无硬度可留空，编码为 000。")
    color = fields.Char(string="颜色", help="可填黑、白黑、玫红白等，也可直接填四位颜色代码。")
    size = fields.Char(string="尺码", help="可填 230、XS、XL 或通码。")
    generated_code = fields.Char(string="生成的产品编码", readonly=True)
    manufacturer_match = fields.Char(string="厂家：输入 → 匹配", readonly=True)
    product_type_match = fields.Char(string="产品类型：输入 → 匹配", readonly=True)
    color_match = fields.Char(string="颜色：输入 → 匹配", readonly=True)
    finished_type_match = fields.Char(string="成品类型：输入 → 匹配", readonly=True)
    audience_match = fields.Char(string="成人/儿童：输入 → 匹配", readonly=True)
    match_summary = fields.Text(string="输入与规则匹配结果", compute="_compute_match_summary")
    status = fields.Selection([
        ("pending", "未生成"), ("ready", "可使用"), ("exists", "ERP已存在"),
        ("duplicate", "本批次重复"), ("error", "无法生成"),
    ], string="状态", default="pending", readonly=True)
    remark = fields.Char(string="备注", readonly=True)

    @api.depends(
        "manufacturer_match", "product_type_match", "color_match",
        "finished_type_match", "audience_match",
    )
    def _compute_match_summary(self):
        for line in self:
            line.match_summary = "\n".join(filter(None, [
                line.manufacturer_match,
                line.product_type_match,
                line.color_match,
                line.finished_type_match,
                line.audience_match,
            ]))

    def _normalize_manufacturer(self):
        value = _clean(self.manufacturer)
        match = re.match(r"^(\d{1,2})", value)
        if match:
            code = match.group(1).zfill(2)
            if code in MANUFACTURER_NAMES:
                return code, _match_display(value, MANUFACTURER_NAMES[code], code, "code")
        code, name, method, error = _match_named_rule(
            value, MANUFACTURER_CODES, MANUFACTURER_NAMES
        )
        return code, _match_display(value, name, code, method, error)

    def _normalize_yymm(self):
        value = re.sub(r"\D", "", _clean(self.production_yymm))
        if len(value) == 6 and value.startswith("20"):
            value = value[2:]
        if len(value) != 4:
            return ""
        month = int(value[2:]) if value[2:].isdigit() else 0
        return value if 1 <= month <= 12 else ""

    def _normalize_product_type(self):
        value = _clean(self.product_type)
        source = value or _clean(self.product_name)
        code, name, method, error = _match_named_rule(
            source,
            PRODUCT_TYPE_CODES,
            PRODUCT_TYPE_NAMES,
            valid_codes=PRODUCT_TYPE_NAMES.keys(),
        )
        rental = False
        if code and source.casefold().endswith("r"):
            rental = True
        if code and ("租赁" in source or "租用" in source):
            rental = True
        if rental:
            code = f"{code}r"
            name = _("%s（租赁特化）") % name
        return code, _match_display(source, name, code, method, error)

    def _normalize_flex(self):
        value = _clean(self.flex)
        if not value or value in ("无", "无硬度", "none", "None"):
            return "000"
        digits = re.sub(r"\D", "", value)
        if not digits or len(digits) > 3:
            return ""
        return digits.zfill(3)

    def _normalize_color(self):
        value = _clean(self.color).upper().replace("色", "")
        if re.fullmatch(r"[A-Z]{1,4}\d{0,3}", value) and len(value) == 4:
            return value, _match_display(self.color, _("颜色代码"), value, "code")
        value = re.sub(r"[\s/、+\-]+", "", value)
        tokens = ""
        matched_names = []
        index = 0
        while index < len(value):
            if value[index:index + 2] == "玫红":
                tokens += "Q"
                matched_names.append("玫红")
                index += 2
                continue
            token = COLOR_TOKENS.get(value[index])
            if not token:
                code, name, method, error = _match_named_rule(
                    value, COLOR_TOKENS, COLOR_NAMES
                )
                if not code:
                    return "", _match_display(self.color, error=error)
                tokens = code
                matched_names = [name]
                break
            tokens += token
            matched_names.append(COLOR_NAMES[token])
            index += 1
        if not tokens or len(tokens) > 4:
            return "", _match_display(self.color, error=_("无法形成四位颜色编码"))
        # The colour segment is always exactly four characters: H001, CY01,
        # WRB1, or a four-token colour such as WRBG.
        if len(tokens) == 4:
            code = tokens
        else:
            code = tokens + ("0" * (3 - len(tokens))) + "1"
        name = "+".join(matched_names)
        return code, _match_display(self.color, name, code, "exact")

    def _normalize_size(self):
        value = _clean(self.size).upper()
        if value in ("通码", "均码", "ONE SIZE", "ONESIZE", "OS"):
            return "###"
        if not value or len(value) > 3:
            return ""
        return value.rjust(3, "#")

    def _build_product_code(self):
        self.ensure_one()
        manufacturer, manufacturer_match = self._normalize_manufacturer()
        yymm = self._normalize_yymm()
        product_type, product_type_match = self._normalize_product_type()
        finished_type = _clean(self.finished_type).upper()
        audience = _clean(self.audience).upper()
        flex = self._normalize_flex()
        color, color_match = self._normalize_color()
        size = self._normalize_size()
        finished_names = {"M": "成品", "F": "非成品"}
        audience_names = {"A": "成人", "K": "儿童/青少年"}
        matches = {
            "manufacturer_match": manufacturer_match,
            "product_type_match": product_type_match,
            "color_match": color_match,
            "finished_type_match": _match_display(
                self.finished_type, finished_names.get(finished_type, ""), finished_type,
                "code", "" if finished_type in finished_names else _("未匹配 M/F 规则"),
            ),
            "audience_match": _match_display(
                self.audience, audience_names.get(audience, ""), audience,
                "code", "" if audience in audience_names else _("未匹配 A/K 规则"),
            ),
        }
        errors = []
        for value, label in [
            (manufacturer, "厂家"), (yymm, "生产年月"), (product_type, "产品类型"),
            (finished_type, "成品类型"), (audience, "成人/儿童"),
            (flex, "硬度"), (color, "颜色"), (size, "尺码"),
        ]:
            if not value:
                errors.append(label)
        if finished_type not in ("M", "F"):
            errors.append("成品类型")
        if audience not in ("A", "K"):
            errors.append("成人/儿童")
        if errors:
            return "", _("该行已阻止生成；空白或未匹配：%s") % "、".join(dict.fromkeys(errors)), matches
        return (
            f"{manufacturer}{yymm}{product_type}-{finished_type}{audience}{flex}-{color}{size}",
            "",
            matches,
        )
