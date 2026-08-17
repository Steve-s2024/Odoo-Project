import re
from io import BytesIO
from urllib.parse import urlencode

from odoo import _, api, fields, models
from odoo.exceptions import UserError
from odoo.fields import Command


PRODUCT_IMPORT_TEMPLATE_ROUTE = "/stock_subwarehouse_hierarchy/import_template/product_template.xlsx"
IMPORT_CUSTOM_ATTRIBUTE_SLOT_COUNT = 20


class _ProductCodeTrie:
    """In-memory exact lookup with O(length(code)) membership checks."""

    _END = object()

    def __init__(self, values=()):
        self.root = {}
        for value in values:
            self.add(value)

    def add(self, value):
        node = self.root
        for character in value:
            node = node.setdefault(character, {})
        node[self._END] = True

    def contains(self, value):
        node = self.root
        for character in value:
            node = node.get(character)
            if node is None:
                return False
        return self._END in node


class ProductAttribute(models.Model):
    _inherit = "product.attribute"

    x_apply_to_all_products = fields.Boolean(
        string="应用到所有产品",
        help="通过全局产品属性工具创建的属性会添加到所有现有和未来创建的产品。",
    )
    x_default_custom_value = fields.Char(
        string="默认自定义值",
        help="应用到产品的此全局自定义属性默认自由文本值。",
    )


class ProductTemplateCustomAttributeValue(models.Model):
    _name = "product.template.custom.attribute.value"
    _description = "产品自定义属性值"
    _order = "attribute_id, id"

    product_tmpl_id = fields.Many2one(
        "product.template",
        required=True,
        ondelete="cascade",
        index=True,
    )
    attribute_id = fields.Many2one(
        "product.attribute",
        required=True,
        domain=[("x_apply_to_all_products", "=", True)],
        ondelete="cascade",
        index=True,
    )
    value_text = fields.Char(string="值")

    _unique_product_attribute = models.Constraint(
        "UNIQUE(product_tmpl_id, attribute_id)",
        "每个自定义属性在同一产品中只能出现一次。",
    )


class ProductTemplate(models.Model):
    _inherit = "product.template"

    x_material_type = fields.Selection(
        selection=[
            ("finished", "\u6210\u54c1"),
            ("semi_finished", "\u534a\u6210\u54c1"),
            ("component", "\u90e8\u4ef6"),
            ("raw_material", "\u539f\u6599"),
            ("packaging", "\u5305\u88c5\u6750\u6599"),
        ],
        string="\u7269\u6599\u7c7b\u578b",
        default="finished",
        copy=True,
        index=True,
    )
    x_component_material = fields.Char(
        string="\u6750\u6599",
        help="\u90e8\u4ef6\u7684\u6750\u6599\u6216\u6750\u6599\u724c\u53f7\uff0c\u5141\u8bb8\u8f93\u5165\u4efb\u610f\u6587\u672c\u3002",
        copy=True,
        index=True,
    )
    x_component_specification = fields.Char(
        string="\u5c3a\u5bf8\u89c4\u683c",
        help="\u90e8\u4ef6\u7684\u5c3a\u5bf8\u3001\u89c4\u683c\u6216\u516c\u5dee\u63cf\u8ff0\uff0c\u5141\u8bb8\u8f93\u5165\u4efb\u610f\u6587\u672c\u3002",
        copy=True,
        index=True,
    )
    x_component_color = fields.Char(
        string="\u989c\u8272",
        help="\u90e8\u4ef6\u989c\u8272\u6216\u8868\u9762\u5904\u7406\u63cf\u8ff0\uff0c\u5141\u8bb8\u8f93\u5165\u4efb\u610f\u6587\u672c\u3002",
        copy=True,
        index=True,
    )

    x_custom_attribute_value_ids = fields.One2many(
        "product.template.custom.attribute.value",
        "product_tmpl_id",
        string="自定义属性",
        copy=True,
    )
    x_shop_group_variant_count = fields.Integer(
        string="同名网店规格数",
        compute="_compute_x_shop_group_variant_count",
    )
    x_shop_group_cover = fields.Boolean(
        string="同名商品网店封面",
        help="设置后，同名商品组的网店商品卡将使用此产品的主图。每个同名商品组只能有一个封面。",
        copy=False,
        index=True,
    )
    x_website_english_name = fields.Char(
        string="英文网站名称",
        help="面向英文网站访客显示的产品名称，不会改变中文ERP产品名称或库存编码。",
        copy=True,
        index=True,
    )
    x_website_description_zh = fields.Html(
        string="中文网站描述",
        help="中文网站商品页显示的产品描述。",
        sanitize_overridable=True,
        sanitize_attributes=False,
        sanitize_form=False,
        copy=True,
    )
    x_website_description_en = fields.Html(
        string="英文网站描述",
        help="英文网站商品页显示的产品描述。",
        sanitize_overridable=True,
        sanitize_attributes=False,
        sanitize_form=False,
        copy=True,
    )
    x_website_code_mapping_id = fields.Many2one(
        "stock.subwarehouse.product.website.code.mapping",
        string="英文网站编号规则",
        readonly=True,
        copy=False,
        ondelete="set null",
    )
    x_website_mapping_flex = fields.Char(
        string="英文映射硬度",
        help="用于匹配国际价格表的 flex 值。留空时会使用产品现有的硬度属性或编码解码结果。",
        copy=True,
        index=True,
    )
    x_website_usd_price = fields.Float(
        string="英文网站价格 (USD)",
        digits="Product Price",
        help="面向英文网站访客显示的美元零售价。该字段不会改变中文销售价格。",
        copy=True,
    )

    # Unique import-only columns. They avoid repeated one2many field paths, which
    # Odoo can pair incorrectly during spreadsheet imports.
    x_import_custom_attribute_1 = fields.Char(string="导入自定义属性 1", copy=False)
    x_import_custom_attribute_value_1 = fields.Char(string="导入自定义属性值 1", copy=False)
    x_import_custom_attribute_2 = fields.Char(string="导入自定义属性 2", copy=False)
    x_import_custom_attribute_value_2 = fields.Char(string="导入自定义属性值 2", copy=False)
    x_import_custom_attribute_3 = fields.Char(string="导入自定义属性 3", copy=False)
    x_import_custom_attribute_value_3 = fields.Char(string="导入自定义属性值 3", copy=False)
    x_import_custom_attribute_4 = fields.Char(string="导入自定义属性 4", copy=False)
    x_import_custom_attribute_value_4 = fields.Char(string="导入自定义属性值 4", copy=False)
    x_import_custom_attribute_5 = fields.Char(string="导入自定义属性 5", copy=False)
    x_import_custom_attribute_value_5 = fields.Char(string="导入自定义属性值 5", copy=False)
    x_import_custom_attribute_6 = fields.Char(string="导入自定义属性 6", copy=False)
    x_import_custom_attribute_value_6 = fields.Char(string="导入自定义属性值 6", copy=False)
    x_import_custom_attribute_7 = fields.Char(string="导入自定义属性 7", copy=False)
    x_import_custom_attribute_value_7 = fields.Char(string="导入自定义属性值 7", copy=False)
    x_import_custom_attribute_8 = fields.Char(string="导入自定义属性 8", copy=False)
    x_import_custom_attribute_value_8 = fields.Char(string="导入自定义属性值 8", copy=False)
    x_import_custom_attribute_9 = fields.Char(string="导入自定义属性 9", copy=False)
    x_import_custom_attribute_value_9 = fields.Char(string="导入自定义属性值 9", copy=False)
    x_import_custom_attribute_10 = fields.Char(string="导入自定义属性 10", copy=False)
    x_import_custom_attribute_value_10 = fields.Char(string="导入自定义属性值 10", copy=False)
    x_import_custom_attribute_11 = fields.Char(string="导入自定义属性 11", copy=False)
    x_import_custom_attribute_value_11 = fields.Char(string="导入自定义属性值 11", copy=False)
    x_import_custom_attribute_12 = fields.Char(string="导入自定义属性 12", copy=False)
    x_import_custom_attribute_value_12 = fields.Char(string="导入自定义属性值 12", copy=False)
    x_import_custom_attribute_13 = fields.Char(string="导入自定义属性 13", copy=False)
    x_import_custom_attribute_value_13 = fields.Char(string="导入自定义属性值 13", copy=False)
    x_import_custom_attribute_14 = fields.Char(string="导入自定义属性 14", copy=False)
    x_import_custom_attribute_value_14 = fields.Char(string="导入自定义属性值 14", copy=False)
    x_import_custom_attribute_15 = fields.Char(string="导入自定义属性 15", copy=False)
    x_import_custom_attribute_value_15 = fields.Char(string="导入自定义属性值 15", copy=False)
    x_import_custom_attribute_16 = fields.Char(string="导入自定义属性 16", copy=False)
    x_import_custom_attribute_value_16 = fields.Char(string="导入自定义属性值 16", copy=False)
    x_import_custom_attribute_17 = fields.Char(string="导入自定义属性 17", copy=False)
    x_import_custom_attribute_value_17 = fields.Char(string="导入自定义属性值 17", copy=False)
    x_import_custom_attribute_18 = fields.Char(string="导入自定义属性 18", copy=False)
    x_import_custom_attribute_value_18 = fields.Char(string="导入自定义属性值 18", copy=False)
    x_import_custom_attribute_19 = fields.Char(string="导入自定义属性 19", copy=False)
    x_import_custom_attribute_value_19 = fields.Char(string="导入自定义属性值 19", copy=False)
    x_import_custom_attribute_20 = fields.Char(string="导入自定义属性 20", copy=False)
    x_import_custom_attribute_value_20 = fields.Char(string="导入自定义属性值 20", copy=False)

    def init(self):
        super().init()
        self.env.cr.execute("""
            UPDATE product_template
               SET x_material_type = 'finished'
             WHERE x_material_type IS NULL
        """)
        # Preserve descriptions entered before bilingual fields were introduced.
        # The existing website description is translated JSON in Odoo 19; prefer
        # its Chinese value, with English as a last-resort migration source.
        self.env.cr.execute("""
            UPDATE product_template
               SET x_website_description_zh = COALESCE(
                       description_ecommerce->>'zh_CN',
                       description_ecommerce->>'en_US'
                   )
             WHERE x_website_description_zh IS NULL
               AND description_ecommerce IS NOT NULL
        """)
        # The ERP product name is the canonical Chinese business name.  English
        # storefront copy belongs in x_website_english_name and must not replace
        # the product name used by inventory, sales, manufacturing, and imports.
        # Keep the two language keys aligned for legacy rows; the separated shop
        # reads its English label from the dedicated website field/API payload.
        self.env.cr.execute("""
            UPDATE product_template
               SET name = jsonb_set(
                       jsonb_set(
                           COALESCE(name, '{}'::jsonb),
                           '{zh_CN}',
                           to_jsonb(COALESCE(name->>'zh_CN', name->>'en_US')),
                           TRUE
                       ),
                       '{en_US}',
                       to_jsonb(COALESCE(name->>'zh_CN', name->>'en_US')),
                       TRUE
                   )
             WHERE name IS NOT NULL
               AND name->>'en_US' IS DISTINCT FROM COALESCE(name->>'zh_CN', name->>'en_US')
        """)

    def _get_safe_english_website_name_for_slug(self):
        self.ensure_one()
        english_name = " ".join((self.x_website_english_name or "").split())
        if not english_name or not re.search(r"[A-Za-z0-9]", english_name):
            return ""
        if not self.env["ir.http"]._slugify(english_name):
            return ""
        return english_name

    def _sync_english_website_name_translation(self):
        for product in self:
            # Read the persisted map because the language-aware ORM cache can
            # temporarily expose en_US as the zh_CN fallback during imports.
            product.flush_recordset(["name"])
            self.env.cr.execute(
                "SELECT name FROM product_template WHERE id = %s",
                [product.id],
            )
            translations = self.env.cr.fetchone()[0] or {}
            chinese_name = (
                translations.get("zh_CN")
                or translations.get("en_US")
                or product.name
            )
            if not chinese_name:
                continue
            chinese_product = product.with_context(lang="zh_CN")
            super(ProductTemplate, chinese_product).write({"name": chinese_name})
            product.invalidate_recordset(["name"])
            # Do not copy x_website_english_name into `name`: doing so makes
            # backend product lists and exports appear English.  The dedicated
            # field remains the authoritative English storefront label.
            english_product = product.with_context(lang="en_US")
            super(ProductTemplate, english_product).write({"name": chinese_name})

    def _normalize_shop_group_name(self):
        self.ensure_one()
        base_name = self.with_context(lang="zh_CN").name or self.name or ""
        return " ".join(base_name.split()).casefold()

    def _get_website_display_name(self, is_english=False):
        self.ensure_one()
        if is_english and self.x_website_english_name:
            return self.x_website_english_name
        if not is_english:
            return self.with_context(lang="zh_CN").name
        return self.with_context(lang="en_US").name

    def _get_website_display_price_label(self, is_english=False):
        self.ensure_one()
        if is_english:
            return f"${self.x_website_usd_price:,.2f}" if self.x_website_usd_price else "Price on request"
        return f"￥{self.list_price:,.2f}"

    @staticmethod
    def _normalize_website_mapping_flex(value):
        normalized = "".join(str(value or "").split()).casefold()
        for marker in ("硬度", "flex"):
            normalized = normalized.replace(marker, "")
        if normalized in {
            "", "000", "\u65e0", "no", "none", "noflex", "n/a", "notapplicable",
            "notspecified", "\u672a\u8bc6\u522b", "\u9ed8\u8ba4",
        }:
            return ""
        return normalized

    def _get_website_mapping_flex(self):
        self.ensure_one()
        return self.x_website_mapping_flex or self._get_shop_variant_display_values().get("flex") or ""

    def _get_shop_grouped_products(self):
        ProductTemplate = self.env["product.template"]
        grouped_products = {}
        ordered_keys = []
        for product in self:
            key = product._normalize_shop_group_name()
            if key not in grouped_products:
                grouped_products[key] = ProductTemplate
                ordered_keys.append(key)
            grouped_products[key] |= product

        representative_ids = []
        for key in ordered_keys:
            products = grouped_products[key]
            selected_cover = products.filtered(
                lambda product: product.x_shop_group_cover and product.image_1920
            )[:1]
            representative_ids.append((selected_cover or products[:1]).id)
        return ProductTemplate.browse(representative_ids)

    def _get_all_shop_group_siblings(self):
        self.ensure_one()
        base_product = self.with_context(lang="zh_CN")
        normalized_name = " ".join((base_product.name or self.name or "").split())
        if not normalized_name:
            return self
        return base_product.search([
            ("name", "=ilike", normalized_name),
        ], order="default_code, id")

    def action_set_shop_group_cover(self):
        self.ensure_one()
        if not self.sale_ok or not self.website_published:
            raise UserError(_("请先将此产品设为可销售并发布到网店。"))
        if not self.image_1920:
            raise UserError(_("此产品没有主图，请先在“产品图片”页签上传主图。"))

        siblings = self._get_all_shop_group_siblings()
        siblings.filtered("x_shop_group_cover").write({"x_shop_group_cover": False})
        self.write({"x_shop_group_cover": True})
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("网店封面已设置"),
                "message": _("同名商品卡现在使用此产品的主图。"),
                "type": "success",
            },
        }

    def action_clear_shop_group_cover(self):
        self.ensure_one()
        self._get_all_shop_group_siblings().filtered("x_shop_group_cover").write({
            "x_shop_group_cover": False,
        })
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("已恢复自动封面"),
                "message": _("同名商品卡将使用原有的自动选择规则。"),
                "type": "success",
            },
        }

    def _get_shop_product_family(self):
        self.ensure_one()
        name = f"{self.name or ''} {self.display_name or ''} {self.default_code or ''}".casefold()
        product_code = self._get_shop_product_code_from_default_code()

        if any(term in name for term in ("单板鞋", "单板", "snowboard")):
            return "snowboard"
        if any(term in name for term in ("双板鞋", "双板", "雪杖", "滑雪杖", "ski boot", "ski pole")):
            return "ski"

        if product_code in {"S1", "SE"} or product_code.startswith(("D", "DC", "DZ")):
            return "snowboard"
        if product_code in {"S", "S2"} or product_code.startswith(("X", "Z")):
            return "ski"
        return "other"

    def _get_shop_product_code_from_default_code(self):
        self.ensure_one()
        first_code_part = (self.default_code or "").split("-", 1)[0]
        match = re.match(r"\[?(\d{6})([A-Za-z0-9]+)", first_code_part.strip())
        return match.group(2).upper() if match else ""

    def _filter_shop_products_by_family(self, family):
        if family == "other":
            return self.filtered(lambda product: product._get_shop_product_family() == "other")
        return self.filtered(lambda product: product._get_shop_product_family() == family)

    @api.model
    def _get_item_page_products(self, family):
        products = self.sudo().search([
            ("sale_ok", "=", True),
            ("website_published", "=", True),
        ], order="name, default_code, id")
        return products._filter_shop_products_by_family(family)._get_shop_grouped_products()

    def _get_shop_group_siblings(self):
        self.ensure_one()
        base_product = self.with_context(lang="zh_CN")
        normalized_name = " ".join((base_product.name or self.name or "").split())
        if not normalized_name:
            return self
        sibling_ids = base_product.search([
            ("name", "=ilike", normalized_name),
            ("sale_ok", "=", True),
            ("website_published", "=", True),
        ], order="default_code, id").ids
        return self.browse(sibling_ids)

    def _get_shop_group_variant_rows(self, is_english=False):
        self.ensure_one()
        rows = []
        for sibling in self._get_shop_group_siblings():
            available_quantity = sibling._get_shop_available_quantity()
            rows.append({
                "product": sibling,
                "values": sibling._get_shop_variant_display_values(is_english=is_english),
                "available_quantity": available_quantity,
                "is_available": sibling._is_shop_available(),
            })
        return rows

    def _is_shop_boot_product(self):
        self.ensure_one()
        normalized_name = " ".join((self.name or "").lower().split())
        return any(term in normalized_name for term in (
            "双板鞋",
            "单板鞋",
            "滑雪鞋",
            "ski boot",
            "snowboard boot",
        ))

    def _get_shop_variant_option_sort_key(self, key, value):
        normalized_value = " ".join((value or "").strip().split())
        if key == "size":
            number_match = re.search(r"\d+(?:\.\d+)?", normalized_value.replace(",", "."))
            if number_match:
                return (0, float(number_match.group()), normalized_value.casefold())
            letter_sizes = {
                "xxs": 0,
                "xs": 1,
                "s": 2,
                "m": 3,
                "l": 4,
                "xl": 5,
                "xxl": 6,
                "xxxl": 7,
            }
            letter_rank = letter_sizes.get(normalized_value.casefold())
            if letter_rank is not None:
                return (1, letter_rank, normalized_value.casefold())
        return (2, 0, normalized_value.casefold())

    def _get_shop_group_variant_option_groups(self, is_english=False):
        self.ensure_one()
        rows = self._get_shop_group_variant_rows(is_english=is_english)
        option_specs = [
            ("type_color", "Type / Color" if is_english else "类型 / 颜色"),
            ("size", "Size" if is_english else "尺码"),
            ("flex", "Flex" if is_english else "硬度"),
        ]
        groups = []
        for key, label in option_specs:
            if key == "flex" and not any(
                self._normalize_website_mapping_flex(row["values"].get(key))
                for row in rows
            ):
                continue
            values = []
            seen_values = set()
            image_products = {}
            for row in rows:
                value = row["values"].get(key) or "未识别"
                if value in seen_values:
                    current_image_product = image_products.get(value)
                    if (
                        key == "type_color"
                        and current_image_product
                        and not current_image_product.image_1920
                        and row["product"].image_1920
                    ):
                        image_products[value] = row["product"]
                    continue
                seen_values.add(value)
                values.append(value)
                if key == "type_color":
                    image_products[value] = row["product"]
            values.sort(key=lambda value: self._get_shop_variant_option_sort_key(key, value))
            groups.append({
                "key": key,
                "label": label,
                "values": values,
                "image_products": image_products,
            })
        return groups

    def _get_shop_available_quantity(self):
        self.ensure_one()
        product_variant = self.product_variant_id
        if not product_variant or not product_variant.is_storable:
            return 1.0
        quants = self.env["stock.quant"].sudo().search([
            ("product_id", "=", product_variant.id),
            ("location_id.usage", "=", "internal"),
        ])
        quant_available = sum(quants.mapped("available_quantity"))
        stock_available = max(product_variant.free_qty, product_variant.qty_available, 0.0)
        return max(quant_available, stock_available)

    def _is_shop_available(self):
        self.ensure_one()
        product_variant = self.product_variant_id
        if not product_variant or not product_variant.is_storable:
            return True
        return self._get_shop_available_quantity() > 0

    def _get_visible_website_attribute_lines(self):
        self.ensure_one()
        return self.valid_product_template_attribute_line_ids.filtered(
            lambda line: not line.attribute_id.x_apply_to_all_products
        )

    def _compute_x_shop_group_variant_count(self):
        for product in self:
            product.x_shop_group_variant_count = len(product._get_shop_group_siblings())

    def _get_custom_attribute_value(self, aliases):
        self.ensure_one()
        normalized_aliases = {"".join(alias.lower().split()) for alias in aliases}
        fallback_value = ""
        for custom_value in self.x_custom_attribute_value_ids:
            attribute_name = "".join((custom_value.attribute_id.name or "").lower().split())
            if attribute_name in normalized_aliases:
                value_text = custom_value.value_text or ""
                if value_text and value_text.lower() not in {"default", "默认"}:
                    return value_text
                fallback_value = fallback_value or value_text
        return fallback_value

    def _is_missing_shop_variant_value(self, value):
        return not value or value.strip().lower() in {"default", "默认"}

    def _decode_shop_variant_values_from_default_code(self):
        self.ensure_one()
        code_parts = (self.default_code or "").split("-")
        decoded = {
            "color": "未识别",
            "size": "未识别",
            "flex": "未识别",
            "audience": "未识别",
        }
        if len(code_parts) < 3:
            return decoded

        spec_code = code_parts[-2].strip().upper()
        color_size_code = code_parts[-1].strip().upper()

        if len(spec_code) >= 2:
            decoded["audience"] = {
                "A": "成人",
                "K": "儿童/青少年",
            }.get(spec_code[1], decoded["audience"])
        if len(spec_code) >= 5:
            flex_code = spec_code[2:5]
            decoded["flex"] = "无硬度" if flex_code == "000" else flex_code.lstrip("0") or flex_code

        if len(color_size_code) >= 4:
            decoded["color"] = self._decode_shop_color_code(color_size_code[:4])
        if len(color_size_code) >= 3:
            size_code = color_size_code[-3:]
            decoded["size"] = "通码" if size_code == "###" else size_code.replace("#", "") or "通码"

        return decoded

    def _decode_shop_color_code(self, color_code):
        color_names = {
            "CY": "青",
            "B": "蓝",
            "H": "黑",
            "W": "白",
            "R": "红",
            "G": "绿",
            "A": "灰",
            "P": "粉",
            "Y": "黄",
            "O": "橙",
            "Z": "棕",
        }
        colors = []
        index = 0
        while index < len(color_code):
            two_char_token = color_code[index:index + 2]
            one_char_token = color_code[index]
            if two_char_token in color_names:
                colors.append(color_names[two_char_token])
                index += 2
            elif one_char_token in color_names:
                colors.append(color_names[one_char_token])
                index += 1
            else:
                index += 1
        return "".join(colors) or "未识别"

    def _shop_variant_value_or_decoded(self, custom_value, decoded_value):
        if self._is_missing_shop_variant_value(custom_value):
            return decoded_value
        return custom_value

    @staticmethod
    def _get_english_shop_variant_value(key, value):
        value = str(value or "").strip()
        if not value:
            return "Not specified"
        if key == "color":
            color_names = {
                "\u9ed1": "Black", "\u767d": "White", "\u7ea2": "Red", "\u84dd": "Blue",
                "\u7eff": "Green", "\u7070": "Gray", "\u7d2b": "Purple", "\u7c89": "Pink",
                "\u9ec4": "Yellow", "\u6a59": "Orange", "\u68d5": "Brown", "\u91d1": "Gold",
                "\u94f6": "Silver",
            }
            if value in {"\u672a\u8bc6\u522b", "\u9ed8\u8ba4"}:
                return "Not specified"
            translated = [color_names.get(character, character) for character in value]
            return " / ".join(translated)
        if key == "size":
            if value in {"\u901a\u7801", "One size"}:
                return "One size"
            if value in {"\u672a\u8bc6\u522b", "\u9ed8\u8ba4"}:
                return "Not specified"
            return value
        if key == "flex":
            normalized = ProductTemplate._normalize_website_mapping_flex(value)
            if normalized in {"", "000", "\u65e0"}:
                return "No flex"
            if normalized in {"\u672a\u8bc6\u522b", "\u9ed8\u8ba4"}:
                return "Not specified"
            return f"{normalized} flex"
        if key == "audience":
            return {
                "\u6210\u4eba": "Adult",
                "\u513f\u7ae5/\u9752\u5c11\u5e74": "Kids / Youth",
                "\u5b69\u5b50": "Kids",
                "\u672a\u8bc6\u522b": "Not specified",
            }.get(value, value)
        return value

    def _get_shop_variant_display_values(self, is_english=False):
        self.ensure_one()
        decoded_values = self._decode_shop_variant_values_from_default_code()
        color = self._get_custom_attribute_value(["颜色", "颜色分类", "colour", "color"])
        size = self._get_custom_attribute_value(["尺码", "尺寸", "size"])
        flex = self._get_custom_attribute_value(["硬度", "款型", "flex"])
        audience = self._get_custom_attribute_value(["成人儿童", "成人/儿童", "适用人群", "人群", "kids/adult", "kid/adult"])
        values = {
            "default_code": self.default_code or "",
            "color": self._shop_variant_value_or_decoded(color, decoded_values["color"]),
            "size": self._shop_variant_value_or_decoded(size, decoded_values["size"]),
            "flex": self._shop_variant_value_or_decoded(flex, decoded_values["flex"]),
            "audience": self._shop_variant_value_or_decoded(audience, decoded_values["audience"]),
        }
        if is_english:
            values.update({
                key: self._get_english_shop_variant_value(key, value)
                for key, value in values.items()
                if key != "default_code"
            })
        code_parts = (self.default_code or "").strip().split("-")
        first_block = code_parts[0].strip().upper() if code_parts else ""
        last_block = code_parts[-1].strip().upper() if code_parts else ""
        type_code = first_block[-2:] if len(first_block) >= 2 else first_block
        color_code = last_block[:4]
        type_label = type_code or ("Unknown type" if is_english else "未知类型")
        color_label = values["color"]
        values.update({
            "type_code": type_code,
            "color_code": color_code,
            "type_color": f"{type_label} · {color_label}",
        })
        return values

    def _get_shop_group_summary(self):
        self.ensure_one()
        siblings = self._get_shop_group_siblings()
        if len(siblings) <= 1:
            return ""
        sibling_values = [sibling._get_shop_variant_display_values() for sibling in siblings]
        colors = sorted({values["color"] for values in sibling_values if values["color"]})
        sizes = sorted({values["size"] for values in sibling_values if values["size"]})
        parts = [f"{len(siblings)}个规格"]
        if colors:
            parts.append("颜色：" + " / ".join(colors[:4]) + ("..." if len(colors) > 4 else ""))
        if sizes:
            parts.append("尺码：" + " / ".join(sizes[:4]) + ("..." if len(sizes) > 4 else ""))
        return "，".join(parts)

    def action_publish_to_shop(self):
        self.write({
            "sale_ok": True,
            "website_published": True,
        })

    def action_unpublish_from_shop(self):
        self.write({"website_published": False})

    @api.model
    def _get_global_custom_attributes(self):
        return self.env["product.attribute"].search([
            ("x_apply_to_all_products", "=", True),
        ], order="sequence, id")

    @api.model
    def _get_global_attribute_line_commands(self):
        commands = []
        for attribute in self._get_global_custom_attributes():
            value = attribute.value_ids.sorted(lambda record: (record.sequence, record.id))[:1]
            if value:
                commands.append(Command.create({
                    "attribute_id": attribute.id,
                    "value_ids": [Command.set(value.ids)],
                }))
        return commands

    @api.model
    def _get_global_custom_value_commands(self):
        return [
            Command.create({
                "attribute_id": attribute.id,
                "value_text": attribute.x_default_custom_value or "",
            })
            for attribute in self._get_global_custom_attributes()
        ]

    @api.model
    def default_get(self, fields_list):
        defaults = super().default_get(fields_list)
        if "taxes_id" in fields_list:
            defaults["taxes_id"] = [Command.clear()]
        if "supplier_taxes_id" in fields_list:
            defaults["supplier_taxes_id"] = [Command.clear()]
        if "attribute_line_ids" in fields_list:
            defaults["attribute_line_ids"] = (
                defaults.get("attribute_line_ids", [])
                + self._get_global_attribute_line_commands()
            )
        if "x_custom_attribute_value_ids" in fields_list:
            defaults["x_custom_attribute_value_ids"] = (
                defaults.get("x_custom_attribute_value_ids", [])
                + self._get_global_custom_value_commands()
            )
        return defaults

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            vals.setdefault("taxes_id", [Command.clear()])
            vals.setdefault("supplier_taxes_id", [Command.clear()])
        products = super().create(vals_list)
        products._ensure_global_attribute_lines()
        products._sync_english_website_name_translation()
        return products

    def write(self, vals):
        result = super().write(vals)
        if "x_website_english_name" in vals:
            self._sync_english_website_name_translation()
        return result

    @api.model
    def load(self, fields, data):
        import_result = self._filter_unique_product_import_rows(fields, data)
        fields = import_result["fields"]
        data = import_result["data"]
        fields, data = self._ensure_inventory_product_import_storable(fields, data)
        custom_pairs = self._extract_custom_attribute_import_pairs(fields, data)
        if not custom_pairs:
            result = super().load(fields, data) if data else self._empty_product_import_result()
            return self._complete_unique_product_import_result(result, import_result)

        custom_field_names = {
            "x_custom_attribute_value_ids/attribute_id",
            "x_custom_attribute_value_ids/value_text",
        }
        for slot_number in range(1, IMPORT_CUSTOM_ATTRIBUTE_SLOT_COUNT + 1):
            custom_field_names.add(f"x_import_custom_attribute_{slot_number}")
            custom_field_names.add(f"x_import_custom_attribute_value_{slot_number}")

        kept_indexes = [
            index
            for index, field_name in enumerate(fields)
            if field_name not in custom_field_names
        ]
        cleaned_fields = [fields[index] for index in kept_indexes]
        cleaned_data = [
            [row[index] for index in kept_indexes]
            for row in data
        ]

        result = super().load(cleaned_fields, cleaned_data) if cleaned_data else self._empty_product_import_result()
        if result.get("ids"):
            products = self.browse(result["ids"])
            for product, row_pairs in zip(products, custom_pairs):
                product._write_imported_custom_attribute_values(row_pairs)
        return self._complete_unique_product_import_result(result, import_result)

    @api.model
    def _empty_product_import_result(self):
        return {"ids": [], "messages": [], "nextrow": 0}

    @api.model
    def _normalize_import_product_code(self, value):
        return str(value or "").strip().casefold()

    @api.model
    def _filter_unique_product_import_rows(self, import_fields, import_data):
        original_import_data = import_data
        result = {
            "fields": import_fields,
            "data": import_data,
            "kept_source_indexes": list(range(len(import_data))),
            "failures": [],
            "applied": False,
            "source_window_size": len(import_data),
            "has_more_source_rows": False,
        }
        if not self.env.context.get("import_file"):
            return result

        original_data_length = len(original_import_data)
        import_limit = self.env.context.get("_import_limit")
        source_window_size = (
            min(len(import_data), import_limit)
            if import_limit
            else len(import_data)
        )
        import_data = original_import_data[:source_window_size]
        result.update({
            "data": import_data,
            "kept_source_indexes": list(range(len(import_data))),
            "applied": True,
            "source_window_size": source_window_size,
            "has_more_source_rows": source_window_size < original_data_length,
        })

        name_index = import_fields.index("name") if "name" in import_fields else None
        source_offset = self.env.context.get("product_import_source_offset", 0)
        header_offset = 2 if self.env.context.get("product_import_has_headers") else 1
        if "default_code" not in import_fields:
            result["data"] = []
            result["kept_source_indexes"] = []
            result["failures"] = [
                {
                    "source_index": source_index,
                    "source_row": source_offset + source_index + header_offset,
                    "product_name": str(row[name_index] or "").strip() if name_index is not None else "",
                    "default_code": "",
                    "reason": "必须映射产品编码（内部编号）字段。",
                }
                for source_index, row in enumerate(import_data)
            ]
            return result

        code_index = import_fields.index("default_code")
        # Serialize product imports so two concurrent files cannot both pass
        # the pre-check before either transaction commits.
        self.env.cr.execute(
            "SELECT pg_advisory_xact_lock(hashtext(%s))",
            ["stock_subwarehouse_hierarchy.product_default_code_import"],
        )
        self.env["product.product"].flush_model(["default_code"])
        # Build once from indexed database rows, then perform every membership
        # lookup in O(length(code)); do not issue one ORM query per input row.
        self.env.cr.execute("""
            SELECT default_code
              FROM product_product
             WHERE default_code IS NOT NULL
               AND btrim(default_code) != ''
        """)
        existing_codes = _ProductCodeTrie(
            self._normalize_import_product_code(row[0])
            for row in self.env.cr.fetchall()
        )
        accepted_codes = _ProductCodeTrie()
        kept_rows = []
        kept_source_indexes = []
        failures = []
        for source_index, row in enumerate(import_data):
            display_code = str(row[code_index] or "").strip()
            normalized_code = self._normalize_import_product_code(display_code)
            product_name = (
                str(row[name_index] or "").strip()
                if name_index is not None
                else ""
            )
            if not normalized_code:
                reason = "产品编码不能为空。"
            elif existing_codes.contains(normalized_code):
                reason = "产品编码已存在于 ERP。"
            elif accepted_codes.contains(normalized_code):
                reason = "产品编码在本次导入文件中重复。"
            else:
                accepted_codes.add(normalized_code)
                kept_rows.append(row)
                kept_source_indexes.append(source_index)
                continue

            failures.append({
                "source_index": source_index,
                "source_row": source_offset + source_index + header_offset,
                "product_name": product_name,
                "default_code": display_code,
                "reason": reason,
            })

        return {
            "fields": import_fields,
            "data": kept_rows,
            "kept_source_indexes": kept_source_indexes,
            "failures": failures,
            "applied": True,
            "source_window_size": source_window_size,
            "has_more_source_rows": result["has_more_source_rows"],
        }

    @api.model
    def _complete_unique_product_import_result(self, result, import_result):
        failures = import_result["failures"]
        result["x_product_import_failures"] = [
            {
                "source_row": failure["source_row"],
                "default_code": failure["default_code"],
                "reason": failure["reason"],
            }
            for failure in failures
        ]
        if import_result.get("applied"):
            result["nextrow"] = (
                import_result["source_window_size"]
                if import_result["has_more_source_rows"]
                else 0
            )

        batch_id = self.env.context.get("product_import_result_batch_id")
        if batch_id:
            self._record_product_import_results(result, import_result, batch_id)
        return result

    @api.model
    def _record_product_import_results(self, result, import_result, batch_id):
        line_model = self.env["stock.subwarehouse.product.import.result.line"]
        source_offset = self.env.context.get("product_import_source_offset", 0)
        header_offset = 2 if self.env.context.get("product_import_has_headers") else 1
        fields_list = import_result["fields"]
        name_index = fields_list.index("name") if "name" in fields_list else None
        code_index = fields_list.index("default_code") if "default_code" in fields_list else None
        created_ids = list(result.get("ids") or [])
        success_values = []
        for result_index, (source_index, product_id) in enumerate(zip(
            import_result["kept_source_indexes"],
            created_ids,
        )):
            row = import_result["data"][result_index]
            success_values.append({
                "batch_id": batch_id,
                "source_row": source_offset + source_index + header_offset,
                "status": "success",
                "product_name": str(row[name_index] or "").strip() if name_index is not None else "",
                "default_code": str(row[code_index] or "").strip() if code_index is not None else "",
                "product_tmpl_id": product_id,
                "reason": "导入成功。",
            })
        failed_values = [
            {
                "batch_id": batch_id,
                "source_row": failure["source_row"],
                "status": "failed",
                "product_name": failure["product_name"],
                "default_code": failure["default_code"],
                "reason": failure["reason"],
            }
            for failure in import_result["failures"]
        ]
        if success_values or failed_values:
            line_model.create(success_values + failed_values)

    @api.model
    def _ensure_inventory_product_import_storable(self, import_fields, import_data):
        if "is_storable" in import_fields:
            return import_fields, import_data
        uses_inventory_template = (
            "type" in import_fields
            or any(field_name.startswith("x_import_custom_attribute_") for field_name in import_fields)
            or "x_custom_attribute_value_ids/attribute_id" in import_fields
        )
        if not uses_inventory_template or "is_storable" not in self._fields:
            return import_fields, import_data
        return (
            [*import_fields, "is_storable"],
            [[*row, "1"] for row in import_data],
        )

    @api.model
    def _extract_custom_attribute_import_pairs(self, import_fields, import_data):
        legacy_attribute_field = "x_custom_attribute_value_ids/attribute_id"
        legacy_value_field = "x_custom_attribute_value_ids/value_text"
        legacy_attribute_indexes = [
            index
            for index, field_name in enumerate(import_fields)
            if field_name == legacy_attribute_field
        ]
        legacy_value_indexes = [
            index
            for index, field_name in enumerate(import_fields)
            if field_name == legacy_value_field
        ]

        slot_pairs = []
        value_only_slots = []
        global_attributes = self._get_global_custom_attributes()
        for slot_number in range(1, IMPORT_CUSTOM_ATTRIBUTE_SLOT_COUNT + 1):
            attribute_slot = f"x_import_custom_attribute_{slot_number}"
            value_slot = f"x_import_custom_attribute_value_{slot_number}"
            if attribute_slot in import_fields and value_slot in import_fields:
                slot_pairs.append((
                    import_fields.index(attribute_slot),
                    import_fields.index(value_slot),
                ))
            elif value_slot in import_fields and len(global_attributes) >= slot_number:
                value_only_slots.append((
                    global_attributes[slot_number - 1].display_name,
                    import_fields.index(value_slot),
                ))

        if (
            (not legacy_attribute_indexes or not legacy_value_indexes)
            and not slot_pairs
            and not value_only_slots
        ):
            return []

        pairs_by_row = []
        for row in import_data:
            pairs = []
            for attribute_index, value_index in zip(legacy_attribute_indexes, legacy_value_indexes):
                attribute_name = str(row[attribute_index] or "").strip()
                value_text = str(row[value_index] or "").strip()
                if attribute_name:
                    pairs.append((attribute_name, value_text))
            for attribute_index, value_index in slot_pairs:
                attribute_name = str(row[attribute_index] or "").strip()
                value_text = str(row[value_index] or "").strip()
                if attribute_name:
                    pairs.append((attribute_name, value_text))
            for attribute_name, value_index in value_only_slots:
                pairs.append((attribute_name, str(row[value_index] or "").strip()))
            pairs_by_row.append(pairs)
        return pairs_by_row

    def _write_imported_custom_attribute_values(self, row_pairs):
        ProductAttribute = self.env["product.attribute"]
        CustomValue = self.env["product.template.custom.attribute.value"]
        for attribute_name, value_text in row_pairs:
            attribute = ProductAttribute.search([
                ("x_apply_to_all_products", "=", True),
                ("name", "=", attribute_name),
            ], limit=1)
            if not attribute:
                matches = ProductAttribute.name_search(
                    name=attribute_name,
                    domain=[("x_apply_to_all_products", "=", True)],
                    operator="=",
                    limit=1,
                )
                attribute = ProductAttribute.browse(matches[0][0]) if matches else ProductAttribute
            if not attribute:
                continue

            custom_value = self.x_custom_attribute_value_ids.filtered(
                lambda record: record.attribute_id == attribute
            )
            if custom_value:
                custom_value[:1].value_text = value_text
            else:
                CustomValue.create({
                    "product_tmpl_id": self.id,
                    "attribute_id": attribute.id,
                    "value_text": value_text,
                })

    def _ensure_global_attribute_lines(self):
        AttributeLine = self.env["product.template.attribute.line"]
        CustomValue = self.env["product.template.custom.attribute.value"]
        global_attributes = self._get_global_custom_attributes()
        for product in self:
            existing_attributes = product.attribute_line_ids.attribute_id
            for attribute in global_attributes - existing_attributes:
                value = attribute.value_ids.sorted(lambda record: (record.sequence, record.id))[:1]
                if value:
                    AttributeLine.create({
                        "product_tmpl_id": product.id,
                        "attribute_id": attribute.id,
                        "value_ids": [Command.set(value.ids)],
                    })
            existing_custom_attributes = product.x_custom_attribute_value_ids.attribute_id
            for attribute in global_attributes - existing_custom_attributes:
                CustomValue.create({
                    "product_tmpl_id": product.id,
                    "attribute_id": attribute.id,
                    "value_text": attribute.x_default_custom_value or "",
                })

    def _remove_global_custom_attribute(self, attribute):
        self.mapped("x_custom_attribute_value_ids").filtered(
            lambda value: value.attribute_id == attribute
        ).unlink()
        self.mapped("attribute_line_ids").filtered(
            lambda line: line.attribute_id == attribute
        ).unlink()

    @api.model
    def get_import_templates(self):
        return [{
            "label": _("产品导入模板（当前自定义属性）"),
            "template": PRODUCT_IMPORT_TEMPLATE_ROUTE,
        }]

    @api.model
    def _get_dynamic_product_import_columns(self):
        columns = [
            ("name", "\u4ea7\u54c1\u540d\u79f0"),
            ("x_website_english_name", "\u82f1\u6587\u7f51\u7ad9\u540d\u79f0"),
            ("x_website_mapping_flex", "\u82f1\u6587\u6620\u5c04\u786c\u5ea6"),
            ("x_website_usd_price", "\u82f1\u6587\u7f51\u7ad9\u4ef7\u683c (USD)"),
            ("default_code", "\u5185\u90e8\u7f16\u53f7"),
            ("type", "\u4ea7\u54c1\u7c7b\u578b"),
            ("x_material_type", "\u7269\u6599\u7c7b\u578b"),
            ("x_component_material", "\u6750\u6599"),
            ("x_component_specification", "\u5c3a\u5bf8\u89c4\u683c"),
            ("x_component_color", "\u989c\u8272"),
            ("is_storable", "\u53ef\u5e93\u5b58"),
            ("categ_id", "\u4ea7\u54c1\u7c7b\u522b"),
            ("list_price", "\u9500\u552e\u4ef7\u683c"),
            ("standard_price", "\u6210\u672c"),
            ("uom_id", "\u8ba1\u91cf\u5355\u4f4d"),
            ("uom_po_id", "\u91c7\u8d2d\u5355\u4f4d"),
            ("sale_ok", "\u53ef\u9500\u552e"),
            ("purchase_ok", "\u53ef\u91c7\u8d2d"),
            ("barcode", "\u6761\u7801"),
        ]
        for slot_number, _attribute in enumerate(
            self._get_global_custom_attributes()[:IMPORT_CUSTOM_ATTRIBUTE_SLOT_COUNT],
            start=1,
        ):
            columns += [
                (f"x_import_custom_attribute_value_{slot_number}", _attribute.display_name),
            ]
        return columns

    @api.model
    def _generate_dynamic_product_import_template_xlsx(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as error:
            raise ImportError(_("生成导入模板需要安装 openpyxl。")) from error

        workbook = Workbook()
        import_sheet = workbook.active
        import_sheet.title = "\u4ea7\u54c1\u5bfc\u5165"
        columns = self._get_dynamic_product_import_columns()
        import_sheet.append([field_name for field_name, _label in columns])
        import_sheet.append([label for _field_name, label in columns])

        sample_row = [
            "\u5f39\u7c27\u5957\u7ba1",
            "Spring Sleeve",
            "",
            0,
            "DBX-19",
            "consu",
            "component",
            "\u4e0d\u9508\u94a2",
            "\u03c68.5*8.1*52",
            "\u672c\u8272",
            "1",
            "\u5168\u90e8",
            0,
            0,
            "\u4ef6",
            "\u4ef6",
            True,
            True,
            "",
        ]
        for attribute in self._get_global_custom_attributes()[:IMPORT_CUSTOM_ATTRIBUTE_SLOT_COUNT]:
            sample_row += [
                attribute.x_default_custom_value or "\u4efb\u610f\u6587\u672c",
            ]
        import_sheet.append(sample_row)

        attribute_sheet = workbook.create_sheet("\u81ea\u5b9a\u4e49\u5c5e\u6027\u5217\u8868")
        attribute_sheet.append(["\u5c5e\u6027ID", "\u5c5e\u6027", "\u9ed8\u8ba4\u503c", "\u5141\u8bb8\u4efb\u610f\u503c"])
        for attribute in self._get_global_custom_attributes():
            attribute_sheet.append([
                attribute.id,
                attribute.display_name,
                attribute.x_default_custom_value or "",
                "\u662f",
            ])

        field_sheet = workbook.create_sheet("\u5bfc\u5165\u5b57\u6bb5")
        field_sheet.append(["\u5b57\u6bb5", "\u4e2d\u6587\u8bf4\u660e"])
        for field_name, label in columns:
            field_sheet.append([field_name, label])

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 45)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def action_export_import_template_format(self):
        ids = ",".join(str(record_id) for record_id in self.ids)
        return {
            "type": "ir.actions.act_url",
            "url": f"/stock_subwarehouse_hierarchy/export/product_template.xlsx?{urlencode({'ids': ids})}",
            "target": "self",
        }

    def action_configure_product_bom(self):
        self.ensure_one()
        action = self.env["ir.actions.actions"]._for_xml_id("mrp.mrp_bom_form_action")
        form_view = self.env.ref("mrp.mrp_bom_form_view")
        boms = self.env["mrp.bom"].search([
            ("product_tmpl_id", "=", self.id),
            ("type", "=", "normal"),
        ], order="write_date desc, id desc")
        bom = boms.filtered("bom_line_ids")[:1] or boms[:1]

        action.update({
            "name": _("配置物料清单"),
            "domain": [("product_tmpl_id", "=", self.id)],
            "context": {
                "default_product_tmpl_id": self.id,
                "default_product_qty": 1.0,
                "default_product_uom_id": self.uom_id.id,
                "default_type": "normal",
                "default_company_id": self.company_id.id or self.env.company.id,
            },
            "views": [(form_view.id, "form")],
            "view_mode": "form",
            "target": "current",
        })
        if bom:
            action["res_id"] = bom.id
        else:
            action.pop("res_id", None)
        return action

    def _generate_dynamic_product_export_xlsx(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as error:
            raise ImportError(_("生成导出文件需要安装 openpyxl。")) from error

        workbook = Workbook()
        export_sheet = workbook.active
        export_sheet.title = "\u4ea7\u54c1\u5bfc\u51fa"
        columns = self._get_dynamic_product_import_columns()
        export_sheet.append([field_name for field_name, _label in columns])
        export_sheet.append([label for _field_name, label in columns])
        for product in self:
            export_sheet.append(product._get_dynamic_product_export_row(columns))

        attribute_sheet = workbook.create_sheet("\u81ea\u5b9a\u4e49\u5c5e\u6027\u5217\u8868")
        attribute_sheet.append(["\u5c5e\u6027ID", "\u5c5e\u6027", "\u9ed8\u8ba4\u503c", "\u5141\u8bb8\u4efb\u610f\u503c"])
        for attribute in self._get_global_custom_attributes():
            attribute_sheet.append([
                attribute.id,
                attribute.display_name,
                attribute.x_default_custom_value or "",
                "\u662f",
            ])

        field_sheet = workbook.create_sheet("\u5bfc\u51fa\u5b57\u6bb5")
        field_sheet.append(["\u5b57\u6bb5", "\u4e2d\u6587\u8bf4\u660e"])
        for field_name, label in columns:
            field_sheet.append([field_name, label])

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A3" if sheet == export_sheet else "A2"
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
            if sheet == export_sheet and sheet.max_row >= 2:
                for cell in sheet[2]:
                    cell.font = Font(italic=True)
                    cell.fill = PatternFill("solid", fgColor="E2F0D9")
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 45)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _get_dynamic_product_export_row(self, columns):
        custom_values = {
            value.attribute_id.id: value.value_text
            for value in self.x_custom_attribute_value_ids
        }
        row = []
        for field_name, _label in columns:
            if field_name.startswith("x_import_custom_attribute_value_"):
                slot_number = int(field_name.rsplit("_", 1)[1])
                attributes = self._get_global_custom_attributes()
                attribute = attributes[slot_number - 1] if len(attributes) >= slot_number else self.env["product.attribute"]
                row.append(custom_values.get(attribute.id, "") if attribute else "")
            elif field_name.startswith("x_import_custom_attribute_"):
                slot_number = int(field_name.rsplit("_", 1)[1])
                attributes = self._get_global_custom_attributes()
                attribute = attributes[slot_number - 1] if len(attributes) >= slot_number else self.env["product.attribute"]
                row.append(attribute.display_name if attribute else "")
            else:
                row.append(self._get_product_export_value(field_name))
        return row

    def _get_product_export_value(self, field_name):
        if field_name == "type" and field_name not in self._fields:
            return "consu"
        if field_name not in self._fields:
            return ""
        value = self[field_name]
        field = self._fields[field_name]
        if field.type == "many2one":
            return value.display_name if value else ""
        if field.type in ("many2many", "one2many"):
            return ", ".join(value.mapped("display_name"))
        return value
