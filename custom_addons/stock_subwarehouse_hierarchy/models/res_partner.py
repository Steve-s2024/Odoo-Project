from io import BytesIO

from odoo import _, api, fields, models
from odoo.exceptions import ValidationError


DISTRIBUTOR_IMPORT_TEMPLATE_ROUTE = (
    "/stock_subwarehouse_hierarchy/import_template/distributor.xlsx"
)
SUPPLIER_IMPORT_TEMPLATE_ROUTE = (
    "/stock_subwarehouse_hierarchy/import_template/supplier.xlsx"
)


class ResPartner(models.Model):
    _inherit = "res.partner"

    x_is_distributor = fields.Boolean(string="Is a Distributor", index=True)
    x_distributor_code = fields.Char(
        string="Distributor Code",
        copy=False,
        index=True,
        readonly=True,
    )
    x_distributor_status = fields.Selection(
        selection=[
            ("draft", "Draft"),
            ("review", "Under Review"),
            ("active", "Active"),
            ("suspended", "Suspended"),
            ("terminated", "Terminated"),
        ],
        string="Distributor Status",
        default="draft",
        required=True,
    )
    x_distributor_level = fields.Selection(
        selection=[
            ("standard", "Standard"),
            ("silver", "Silver"),
            ("gold", "Gold"),
            ("strategic", "Strategic"),
        ],
        string="Distributor Level",
        default="standard",
        required=True,
    )
    x_distributor_territory = fields.Char(string="Sales Territory")
    x_distributor_exclusive = fields.Boolean(string="Exclusive Territory")
    x_distributor_agreement_start = fields.Date(string="Agreement Start")
    x_distributor_agreement_end = fields.Date(string="Agreement Expiry")
    x_distributor_annual_target = fields.Monetary(
        string="Annual Purchase Target",
        currency_field="x_distributor_currency_id",
    )
    x_distributor_currency_id = fields.Many2one(
        comodel_name="res.currency",
        string="Target Currency",
        default=lambda self: self.env.company.currency_id,
    )
    x_distributor_default_warehouse_id = fields.Many2one(
        comodel_name="stock.warehouse",
        string="Default Delivery Warehouse",
        check_company=True,
    )
    x_distributor_notes = fields.Html(string="Distributor Notes")
    x_distributor_sale_order_count = fields.Integer(
        string="Sales Orders",
        compute="_compute_distributor_document_counts",
    )
    x_distributor_delivery_count = fields.Integer(
        string="Deliveries",
        compute="_compute_distributor_document_counts",
    )
    x_supplier_code = fields.Char(
        string="Supplier Code",
        copy=False,
        index=True,
        readonly=True,
    )
    x_supplier_status = fields.Selection(
        selection=[
            ("prospective", "Prospective"),
            ("approved", "Approved"),
            ("active", "Active"),
            ("blocked", "Blocked"),
            ("inactive", "Inactive"),
        ],
        string="Supplier Status",
        default="prospective",
        required=True,
    )
    x_supplier_channel_type = fields.Selection(
        selection=[
            ("manufacturer", "Manufacturer"),
            ("wholesaler", "Wholesaler"),
            ("agent", "Agent"),
            ("importer", "Importer"),
            ("other", "Other"),
        ],
        string="Supply Channel",
        default="manufacturer",
        required=True,
    )
    x_supplier_preferred = fields.Boolean(string="Preferred Supplier")
    x_supplier_lead_days = fields.Integer(string="Expected Lead Time (Days)")
    x_supplier_default_warehouse_id = fields.Many2one(
        comodel_name="stock.warehouse",
        string="Default Receipt Warehouse",
        check_company=True,
    )
    x_supplier_notes = fields.Html(string="Supplier Notes")
    x_supplier_outstanding_bill_count = fields.Integer(
        string="Outstanding Bills",
        compute="_compute_supplier_financial_totals",
        compute_sudo=True,
        groups="account.group_account_invoice",
    )
    x_supplier_debt_amount = fields.Monetary(
        string="Amount Payable",
        compute="_compute_supplier_financial_totals",
        compute_sudo=True,
        currency_field="x_supplier_company_currency_id",
        groups="account.group_account_invoice",
    )
    x_supplier_company_currency_id = fields.Many2one(
        comodel_name="res.currency",
        compute="_compute_supplier_company_currency",
        groups="account.group_account_invoice",
    )
    x_supplier_expected_receipt_count = fields.Integer(
        string="Expected Receipts",
        compute="_compute_supplier_receipt_totals",
        compute_sudo=True,
        groups="stock.group_stock_user",
    )
    x_supplier_next_receipt_date = fields.Datetime(
        string="Next Expected Arrival",
        compute="_compute_supplier_receipt_totals",
        compute_sudo=True,
        groups="stock.group_stock_user",
    )

    @api.constrains(
        "x_distributor_agreement_start",
        "x_distributor_agreement_end",
    )
    def _check_distributor_agreement_dates(self):
        for partner in self:
            if (
                partner.x_distributor_agreement_start
                and partner.x_distributor_agreement_end
                and partner.x_distributor_agreement_end
                < partner.x_distributor_agreement_start
            ):
                raise ValidationError(
                    _("The distributor agreement expiry cannot precede its start date.")
                )

    def _ensure_distributor_code(self):
        for partner in self.filtered(
            lambda record: record.x_is_distributor and not record.x_distributor_code
        ):
            code = self.env["ir.sequence"].next_by_code(
                "stock.subwarehouse.distributor"
            )
            super(ResPartner, partner).write({
                "x_distributor_code": code or _("New"),
            })

    def _ensure_supplier_code(self):
        for partner in self.filtered(
            lambda record: record.supplier_rank > 0 and not record.x_supplier_code
        ):
            code = self.env["ir.sequence"].next_by_code(
                "stock.subwarehouse.supplier"
            )
            super(ResPartner, partner).write({
                "x_supplier_code": code or _("New"),
            })

    @api.model
    def action_backfill_supplier_codes(self):
        self.search([
            ("supplier_rank", ">", 0),
            ("x_supplier_code", "=", False),
        ])._ensure_supplier_code()
        return True

    @api.model_create_multi
    def create(self, vals_list):
        partners = super().create(vals_list)
        partners._ensure_distributor_code()
        partners._ensure_supplier_code()
        return partners

    def write(self, vals):
        result = super().write(vals)
        if vals.get("x_is_distributor"):
            self._ensure_distributor_code()
        if vals.get("supplier_rank"):
            self._ensure_supplier_code()
        return result

    @api.depends_context("company")
    def _compute_supplier_company_currency(self):
        for partner in self:
            partner.x_supplier_company_currency_id = (
                partner.company_id.currency_id or self.env.company.currency_id
            )

    def _compute_supplier_financial_totals(self):
        AccountMove = self.env["account.move"]
        for partner in self:
            bills = AccountMove.search([
                ("partner_id", "child_of", partner.commercial_partner_id.id),
                ("move_type", "in", ("in_invoice", "in_refund")),
                ("state", "=", "posted"),
                ("amount_residual", "!=", 0),
            ])
            partner.x_supplier_outstanding_bill_count = len(bills)
            partner.x_supplier_debt_amount = -sum(
                bills.mapped("amount_residual_signed")
            )

    def _compute_supplier_receipt_totals(self):
        StockPicking = self.env["stock.picking"]
        for partner in self:
            receipts = StockPicking.search([
                ("partner_id", "child_of", partner.commercial_partner_id.id),
                ("picking_type_id.code", "=", "incoming"),
                ("state", "not in", ("done", "cancel")),
            ])
            partner.x_supplier_expected_receipt_count = len(receipts)
            partner.x_supplier_next_receipt_date = (
                min(receipts.mapped("scheduled_date")) if receipts else False
            )

    def _compute_distributor_document_counts(self):
        SaleOrder = self.env["sale.order"]
        StockPicking = self.env["stock.picking"]
        for partner in self:
            commercial_partner = partner.commercial_partner_id
            partner_domain = [("partner_id", "child_of", commercial_partner.id)]
            partner.x_distributor_sale_order_count = SaleOrder.search_count(
                partner_domain
            )
            partner.x_distributor_delivery_count = StockPicking.search_count(
                partner_domain
            )

    def action_view_distributor_sale_orders(self):
        self.ensure_one()
        action = self.env["ir.actions.actions"]._for_xml_id("sale.action_orders")
        action["domain"] = [
            ("partner_id", "child_of", self.commercial_partner_id.id),
        ]
        action["context"] = {
            "default_partner_id": self.commercial_partner_id.id,
        }
        return action

    def action_view_distributor_deliveries(self):
        self.ensure_one()
        action = self.env["ir.actions.actions"]._for_xml_id(
            "stock.action_picking_tree_all"
        )
        action["domain"] = [
            ("partner_id", "child_of", self.commercial_partner_id.id),
        ]
        return action

    def action_view_supplier_outstanding_bills(self):
        self.ensure_one()
        action = self.env["ir.actions.actions"]._for_xml_id(
            "account.action_move_in_invoice_type"
        )
        action["name"] = _("Outstanding Bills")
        action["domain"] = [
            ("partner_id", "child_of", self.commercial_partner_id.id),
            ("move_type", "in", ("in_invoice", "in_refund")),
            ("state", "=", "posted"),
            ("amount_residual", "!=", 0),
        ]
        action["context"] = {
            "default_move_type": "in_invoice",
            "default_partner_id": self.commercial_partner_id.id,
        }
        return action

    def action_view_supplier_expected_receipts(self):
        self.ensure_one()
        action = self.env["ir.actions.actions"]._for_xml_id(
            "stock.action_picking_tree_all"
        )
        action["name"] = _("Expected Receipts")
        action["domain"] = [
            ("partner_id", "child_of", self.commercial_partner_id.id),
            ("picking_type_id.code", "=", "incoming"),
            ("state", "not in", ("done", "cancel")),
        ]
        return action

    @api.model
    def _get_distributor_import_columns(self):
        return [
            ("name", "公司名称", "是", "填写公司全称"),
            ("company_type", "联系人类型", "是", "固定填写：公司（company）"),
            ("customer_rank", "客户等级标记", "否", "客户标记填写 1"),
            ("email", "电子邮箱", "否", "填写有效的电子邮箱地址"),
            ("phone", "电话", "否", "填写联系电话，可包含国家/地区代码"),
            ("street", "地址", "否", "填写街道及门牌地址"),
            ("city", "城市", "否", "填写城市名称"),
            ("country_id", "国家/地区", "否", "填写系统中已有的国家/地区名称"),
            (
                "x_distributor_status",
                "经销商状态",
                "是",
                "草稿（draft）/ 待审核（review）/ 生效（active）/ 暂停（suspended）/ 终止（terminated）",
            ),
            (
                "x_distributor_level",
                "经销商等级",
                "是",
                "标准（standard）/ 银级（silver）/ 金级（gold）/ 战略（strategic）",
            ),
            ("x_distributor_territory", "销售区域", "否", "填写销售区域名称"),
            ("x_distributor_exclusive", "独家区域", "否", "填 1 表示是，填 0 表示否"),
            ("x_distributor_agreement_start", "协议开始日期", "否", "按 YYYY-MM-DD 格式填写，如 2026-01-01"),
            ("x_distributor_agreement_end", "协议到期日期", "否", "按 YYYY-MM-DD 格式填写，如 2026-12-31"),
            ("x_distributor_annual_target", "年度采购目标", "否", "填写数字，不要包含货币符号或千位分隔符"),
            ("x_distributor_currency_id", "目标币种", "否", "填写币种代码，如 CNY"),
            ("x_distributor_default_warehouse_id", "默认发货仓库", "否", "填写系统中已有的仓库名称"),
            ("user_id", "负责人", "否", "填写系统中已有的用户名称"),
            ("property_product_pricelist", "销售价目表", "否", "填写系统中已有的价目表名称"),
            ("property_payment_term_id", "客户付款条款", "否", "填写系统中已有的付款条款名称"),
            ("x_distributor_notes", "经销商备注", "否", "填写普通文本或 HTML 内容"),
        ]

    @api.model
    def _get_supplier_import_columns(self):
        return [
            ("name", "公司名称", "是", "文本"),
            ("supplier_rank", "供应商等级标记", "是", "1"),
            ("company_type", "联系人类型", "是", "company"),
            ("email", "电子邮箱", "否", "文本"),
            ("phone", "电话", "否", "文本"),
            ("street", "地址", "否", "文本"),
            ("city", "城市", "否", "文本"),
            ("country_id", "国家/地区", "否", "系统中的国家名称"),
            (
                "x_supplier_status",
                "供应商状态",
                "是",
                "prospective / approved / active / blocked / inactive",
            ),
            (
                "x_supplier_channel_type",
                "进货渠道",
                "是",
                "manufacturer / wholesaler / agent / importer / other",
            ),
            ("x_supplier_preferred", "首选供应商", "否", "1 / 0"),
            ("x_supplier_lead_days", "预计交期（天）", "否", "整数"),
            ("x_supplier_default_warehouse_id", "默认收货仓库", "否", "仓库名称"),
            ("property_supplier_payment_term_id", "供应商付款条款", "否", "付款条款名称"),
            ("x_supplier_notes", "供应商备注", "否", "文本或 HTML"),
        ]

    @api.model
    def _generate_partner_channel_import_template_xlsx(self, channel):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as error:
            raise ImportError(_("生成导入模板需要安装 openpyxl。")) from error

        if channel == "distributor":
            columns = self._get_distributor_import_columns()
            import_sheet_name = "经销商导入"
            title = "经销商导入模板"
        elif channel == "supplier":
            columns = self._get_supplier_import_columns()
            import_sheet_name = "供应商导入"
            title = "供应商导入模板"
        else:
            raise ValidationError(_("Unknown partner import channel."))

        workbook = Workbook()
        import_sheet = workbook.active
        import_sheet.title = import_sheet_name
        import_sheet.append([field_name for field_name, _label, _required, _accepted in columns])
        if channel == "distributor":
            import_sheet.append([label for _field_name, label, _required, _accepted in columns])

        instruction_sheet = workbook.create_sheet("字段说明")
        instruction_sheet.append(["字段", "中文说明", "必填", "填写规则"])
        for field_name, label, required, accepted in columns:
            instruction_sheet.append([field_name, label, required, accepted])

        note_sheet = workbook.create_sheet(
            "导入规则" if channel == "distributor" else "使用说明"
        )
        note_sheet.append([title])
        if channel == "distributor":
            note_sheet.append(["1. 第 1 行是系统字段名，请勿修改；第 2 行是中文标题，仅供阅读。"])
            note_sheet.append(["2. 请从第 3 行开始填写经销商数据；执行导入时忽略第 2 行中文标题。"])
            note_sheet.append(["3. 导入前请先执行测试；选填字段可以留空，必填字段不得为空。"])
            note_sheet.append(["4. 字段允许值及填写方法请查看“字段说明”工作表。"])
            note_sheet.append(["5. 更新已有记录时，建议使用 Odoo 导出的数据库 ID 或外部 ID。"])
            note_sheet.append(["6. 未结清账单、欠款、预计到货等计算字段不参与导入，由系统自动汇总。"])
        else:
            note_sheet.append(["1. 请在第一个工作表中从第 2 行开始填写数据，不要修改第 1 行字段名。"])
            note_sheet.append(["2. 导入前可使用测试导入表验证；空白选填字段可以保留。"])
            note_sheet.append(["3. 已有记录需要更新时，建议使用 Odoo 导出的数据库 ID 或外部 ID。"])
            note_sheet.append(["4. 计算字段（未结清账单、欠款、预计到货）不参与导入，由系统自动汇总。"])

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")
        section_fill = PatternFill("solid", fgColor="D9EAF7")
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.sheet_view.showGridLines = False
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(vertical="center")
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                width = min(max(max_length + 2, 14), 52)
                sheet.column_dimensions[
                    get_column_letter(column_cells[0].column)
                ].width = width
        if channel == "distributor":
            import_sheet.freeze_panes = "A3"
            for cell in import_sheet[2]:
                cell.font = Font(bold=True, color="1F1F1F")
                cell.fill = section_fill
                cell.alignment = Alignment(vertical="center")
        instruction_sheet["A2"].fill = section_fill

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @api.model
    def _partner_channel_export_value(self, partner, field_name):
        value = partner[field_name]
        field = partner._fields[field_name]
        if field.type == "many2one":
            if not value:
                return ""
            if field.comodel_name == "res.currency":
                return value.name
            return value.display_name
        if field.type == "boolean":
            return 1 if value else 0
        if field.type in ("integer", "float", "monetary"):
            return value or 0
        if field.type in ("date", "datetime"):
            return value or ""
        return str(value or "")

    def _generate_partner_channel_export_xlsx(self, channel):
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise ImportError(_("生成导出文件需要安装 openpyxl。")) from error

        content = self._generate_partner_channel_import_template_xlsx(channel)
        workbook = load_workbook(BytesIO(content))
        worksheet = workbook.worksheets[0]
        if channel == "distributor":
            columns = self._get_distributor_import_columns()
        elif channel == "supplier":
            columns = self._get_supplier_import_columns()
        else:
            raise ValidationError(_("Unknown partner import channel."))

        for partner in self.sorted(lambda record: (record.name or "", record.id)):
            worksheet.append([
                self._partner_channel_export_value(partner, field_name)
                for field_name, _label, _required, _accepted in columns
            ])

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @api.model
    def get_import_templates(self):
        channel = self.env.context.get("partner_channel_import_type")
        channel_templates = {
            "distributor": {
                "label": "经销商导入模板",
                "template": DISTRIBUTOR_IMPORT_TEMPLATE_ROUTE,
            },
            "supplier": {
                "label": "供应商导入模板",
                "template": SUPPLIER_IMPORT_TEMPLATE_ROUTE,
            },
        }
        if channel in channel_templates:
            return [channel_templates[channel]]
        return list(channel_templates.values())

    @api.model
    def _open_partner_channel_import(self, channel, name):
        context = dict(self.env.context, partner_channel_import_type=channel)
        if channel == "distributor":
            context.update({
                "default_x_is_distributor": True,
                "default_is_company": True,
                "default_company_type": "company",
                "default_customer_rank": 1,
            })
        return {
            "type": "ir.actions.client",
            "name": name,
            "tag": "import",
            "target": "current",
            "params": {
                "model": "res.partner",
                "active_model": "res.partner",
                "context": context,
            },
            "context": context,
        }

    @api.model
    def action_open_distributor_import(self, _selected_ids=None):
        return self._open_partner_channel_import(
            "distributor",
            _("Import Distributors"),
        )

    @api.model
    def action_open_supplier_import(self, _selected_ids=None):
        return self._open_partner_channel_import(
            "supplier",
            _("Import Suppliers"),
        )
