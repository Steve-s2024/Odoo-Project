from io import BytesIO
import hashlib
import logging
import re
from calendar import timegm
from dateutil.relativedelta import relativedelta
from urllib.parse import urlencode

from odoo import _, api, fields, models
from odoo.addons.base.models.ir_mail_server import MailDeliveryException
from odoo.exceptions import UserError
from odoo.tools import float_compare


SALE_ORDER_IMPORT_TEMPLATE_ROUTE = "/stock_subwarehouse_hierarchy/import_template/sale_order.xlsx"
SALE_IMPORT_FIELD_LABELS_ZH = {
    "name": "订单号",
    "partner_id": "客户",
    "date_order": "下单时间",
    "validity_date": "有效期",
    "payment_term_id": "付款条款",
    "user_id": "销售员",
    "team_id": "销售团队",
    "x_platform": "平台",
    "x_channel": "渠道",
    "x_sale_nature": "销售性质",
    "x_processing_fee": "手续费",
    "x_amount_received": "实收",
    "x_is_external_order": "外部订单",
    "x_finance_remark": "财务备注",
    "order_line/product_id": "产品ID",
    "order_line/product_id/.id": "产品数据库ID",
    "order_line/product_id/id": "产品外部ID",
    "order_line/product_uom_qty": "数量",
    "order_line/price_unit": "单价",
    "order_line/tax_ids": "税",
}
_logger = logging.getLogger(__name__)


class SaleNature(models.Model):
    _name = "stock.subwarehouse.sale.nature"
    _description = "销售性质"
    _order = "sequence, name, id"

    name = fields.Char(string="销售性质", required=True, translate=True)
    code = fields.Char(string="代码", required=True, index=True)
    sequence = fields.Integer(string="顺序", default=10)
    active = fields.Boolean(default=True)

    _code_unique = models.Constraint(
        "unique(code)",
        "销售性质代码必须唯一。",
    )

    @api.model_create_multi
    def create(self, vals_list):
        for values in vals_list:
            if not values.get("code") and values.get("name"):
                values["code"] = self._next_import_code(values["name"])
        return super().create(vals_list)

    @api.model
    def name_create(self, name):
        """Enable Odoo import's “Create new values” fallback safely."""
        created = self.create({"name": name})
        return created.id, created.display_name

    @api.model
    def _next_import_code(self, name):
        base = re.sub(r"[^0-9a-z]+", "_", (name or "").casefold()).strip("_")
        if not base:
            digest = hashlib.sha1((name or "").encode("utf-8")).hexdigest()[:12]
            base = f"nature_{digest}"
        candidate = base[:56]
        suffix = 1
        while self.with_context(active_test=False).search_count([("code", "=", candidate)], limit=1):
            suffix += 1
            candidate = f"{base[:50]}_{suffix}"
        return candidate


class SaleOrder(models.Model):
    _inherit = "sale.order"

    active = fields.Boolean(string="启用", default=True, index=True)

    @api.model
    def load(self, import_fields, import_data):
        # ``x_official_total`` is intentionally present in both import and
        # export formats so users can reconcile the official listed amount.
        # It is authoritative ERP data, however, and must never be trusted
        # from an incoming spreadsheet.
        if "x_official_total" in import_fields:
            official_total_index = import_fields.index("x_official_total")
            import_fields = [
                field_name
                for index, field_name in enumerate(import_fields)
                if index != official_total_index
            ]
            import_data = [
                [value for index, value in enumerate(row) if index != official_total_index]
                for row in import_data
            ]
        filtered = self._filter_skipped_sale_import_rows(import_fields, import_data)
        if filtered["data"]:
            # Nested field paths such as ``order_line/product_id`` cannot be
            # checked by the core post-conversion skip logic (the converted
            # record only contains ``order_line``). Remove only the paths this
            # pre-filter already handled and leave unrelated skip policies to
            # Odoo.
            remaining_skip_fields = [
                field_name
                for field_name in (self.env.context.get("import_skip_records") or [])
                if field_name not in filtered["handled_skip_fields"]
            ]
            import_model = self.with_context(import_skip_records=remaining_skip_fields)
            result = super(SaleOrder, import_model).load(
                filtered["fields"], filtered["data"]
            )
        else:
            result = {"ids": [], "messages": [], "nextrow": 0}

        result["x_sale_import_skipped_rows"] = filtered["skipped_rows"]
        result["x_business_import_failures"] = filtered["duplicate_failures"]
        imported_orders = self.browse(result.get("ids", [])).exists()
        imported_orders.filtered("x_is_external_order")._finalize_external_orders()
        self._record_sale_import_results(result, filtered)
        if filtered["has_more_source_rows"]:
            # Base import adds the current source offset to this value. Keep
            # pagination based on source rows, not on the smaller accepted set.
            result["nextrow"] = filtered["source_window_size"]
        return result

    @api.model
    def _filter_skipped_sale_import_rows(self, import_fields, import_data):
        """Remove expected rejected rows before Odoo's relational conversion.

        Odoo normally converts relational values before it can isolate a bad
        nested product value. A missing or unknown product can therefore emit a
        blocking conversion error and roll back the valid orders as well. Sales
        imports always use partial-success semantics for the mapped product ID:
        invalid product rows are checked and removed before ORM conversion,
        without requiring the user to select a conflict-handling option.
        """
        result = {
            "fields": import_fields,
            "data": import_data,
            "skipped_rows": [],
            "handled_skip_fields": set(),
            "source_window_size": len(import_data),
            "has_more_source_rows": False,
            "accepted_orders": [],
            "duplicate_failures": [],
        }
        if not self.env.context.get("import_file"):
            return result

        skipped_fields = set(self.env.context.get("import_skip_records") or [])
        import_limit = self.env.context.get("_import_limit")
        source_window_size = min(len(import_data), import_limit) if import_limit else len(import_data)
        source_data = import_data[:source_window_size]
        result.update({
            "data": source_data,
            "source_window_size": source_window_size,
            "has_more_source_rows": source_window_size < len(import_data),
        })

        field_indexes = {field_name: index for index, field_name in enumerate(import_fields)}
        selected_skip_fields = skipped_fields.intersection(field_indexes)

        product_field = next((
            field_name
            for field_name in (
                "order_line/product_id",
                "order_line/product_id/.id",
                "order_line/product_id/id",
            )
            if field_name in field_indexes
        ), None)
        # Odoo's post-conversion skip check cannot address nested paths. Product
        # IDs are always handled here so a single mismatch never blocks the
        # remaining valid sales rows.
        result["handled_skip_fields"] = {product_field} if product_field else set()
        if (
            "name" not in field_indexes
            and not product_field
            and "partner_id" not in field_indexes
            and not selected_skip_fields
        ):
            return result

        source_offset = self.env.context.get("sale_import_source_offset", 0)
        header_offset = 2 if self.env.context.get("sale_import_has_headers") else 1
        if "name" not in field_indexes:
            result["data"] = []
            result["duplicate_failures"] = [
                {
                    "source_row": source_offset + source_index + header_offset,
                    "identifier": "",
                    "reason": "必须映射销售订单号（name）字段。",
                }
                for source_index, _row in enumerate(source_data)
            ]
            return result

        self.env.cr.execute(
            "SELECT pg_advisory_xact_lock(hashtext(%s))",
            ["stock_subwarehouse_hierarchy.sale_name_import"],
        )
        self.flush_model(["name"])
        self.env.cr.execute("""
            SELECT lower(btrim(name))
              FROM sale_order
             WHERE name IS NOT NULL AND btrim(name) != ''
        """)
        existing_order_names = {row[0] for row in self.env.cr.fetchall()}
        accepted_order_names = set()

        validate_product_existence = bool(product_field)
        valid_product_values = self._get_valid_sale_import_product_values(
            product_field if validate_product_existence else None,
            source_data,
            field_indexes,
        )

        root_indexes = [
            index for index, field_name in enumerate(import_fields)
            if "/" not in field_name
        ]
        groups = []
        current_group = []
        for source_index, row in enumerate(source_data):
            starts_new_order = any(
                not self._is_empty_import_value(row[index])
                for index in root_indexes
            )
            if current_group and starts_new_order:
                groups.append(current_group)
                current_group = []
            current_group.append((source_index, row))
        if current_group:
            groups.append(current_group)

        accepted_rows = []
        accepted_orders = []
        rejected_rows = []
        duplicate_failures = []
        for group in groups:
            first_source_index, first_row = group[0]
            display_order_name = str(first_row[field_indexes["name"]] or "").strip()
            normalized_order_name = display_order_name.casefold()
            if not normalized_order_name:
                duplicate_failures.append({
                    "source_row": source_offset + first_source_index + header_offset,
                    "identifier": display_order_name,
                    "reason": "销售订单号不能为空。",
                })
                continue
            if normalized_order_name in existing_order_names:
                duplicate_failures.append({
                    "source_row": source_offset + first_source_index + header_offset,
                    "identifier": display_order_name,
                    "reason": "销售订单号已存在于 ERP。",
                })
                continue
            if normalized_order_name in accepted_order_names:
                duplicate_failures.append({
                    "source_row": source_offset + first_source_index + header_offset,
                    "identifier": display_order_name,
                    "reason": "销售订单号在本次导入文件中重复。",
                })
                continue
            required_root_fields = set(
                field_name for field_name in selected_skip_fields if "/" not in field_name
            )
            if "partner_id" in field_indexes:
                required_root_fields.add("partner_id")
            root_failure = next((
                field_name
                for field_name in required_root_fields
                if "/" not in field_name
                and self._is_empty_import_value(first_row[field_indexes[field_name]])
            ), None)
            if root_failure:
                for source_index, row in group:
                    rejected_rows.append(self._sale_import_rejection(
                        source_index,
                        row,
                        product_field,
                        field_indexes,
                        self._sale_import_missing_value_reason(root_failure),
                    ))
                continue

            accepted_group_rows = []
            for source_index, row in group:
                required_line_fields = {product_field} if product_field else set()
                missing_line_field = next((
                    field_name for field_name in required_line_fields
                    if self._is_empty_import_value(row[field_indexes[field_name]])
                ), None)
                if missing_line_field:
                    reason = (
                        "产品ID为空。"
                        if missing_line_field == product_field
                        else self._sale_import_missing_value_reason(missing_line_field)
                    )
                    rejected_rows.append(self._sale_import_rejection(
                        source_index, row, product_field, field_indexes, reason
                    ))
                    continue

                if validate_product_existence:
                    product_value = self._normalize_sale_import_value(
                        row[field_indexes[product_field]]
                    )
                    if product_value not in valid_product_values:
                        rejected_rows.append(self._sale_import_rejection(
                            source_index,
                            row,
                            product_field,
                            field_indexes,
                            "产品ID在ERP中不存在。",
                        ))
                        continue
                accepted_group_rows.append((source_index, list(row)))

            if accepted_group_rows:
                # If the group's first line was rejected, copy the parent order
                # cells to its first valid continuation line.
                first_valid_row = accepted_group_rows[0][1]
                for index in root_indexes:
                    if self._is_empty_import_value(first_valid_row[index]):
                        first_valid_row[index] = first_row[index]
                accepted_rows.extend(row for _source_index, row in accepted_group_rows)
                accepted_order_names.add(normalized_order_name)
                accepted_orders.append({
                    "source_index": first_source_index,
                    "identifier": display_order_name,
                })

        result["data"] = accepted_rows
        result["skipped_rows"] = rejected_rows
        result["accepted_orders"] = accepted_orders
        result["duplicate_failures"] = duplicate_failures
        return result

    @api.model
    def _record_sale_import_results(self, result, filtered):
        batch_id = self.env.context.get("business_import_result_batch_id")
        if not batch_id:
            return
        source_offset = self.env.context.get("business_import_source_offset", 0)
        header_offset = 2 if self.env.context.get("business_import_has_headers") else 1
        values = []
        for accepted, record_id in zip(filtered["accepted_orders"], result.get("ids") or []):
            values.append({
                "batch_id": batch_id,
                "source_row": source_offset + accepted["source_index"] + header_offset,
                "status": "success",
                "identifier": accepted["identifier"],
                "record_ref": f"sale.order,{record_id}",
                "reason": "导入成功。",
            })
        values.extend({
            "batch_id": batch_id,
            "source_row": failure["source_row"],
            "status": "failed",
            "identifier": failure["identifier"],
            "reason": failure["reason"],
        } for failure in filtered["duplicate_failures"])
        if values:
            self.env["stock.subwarehouse.business.import.result.line"].create(values)

    @api.model
    def _get_valid_sale_import_product_values(self, product_field, source_data, field_indexes):
        if not product_field:
            return set()

        value_index = field_indexes[product_field]
        raw_values = {
            self._normalize_sale_import_value(row[value_index])
            for row in source_data
            if not self._is_empty_import_value(row[value_index])
        }
        if not raw_values:
            return set()

        product_model = self.env["product.product"].with_context(active_test=False)
        if product_field.endswith("/.id"):
            numeric_ids = {int(value) for value in raw_values if value.isdigit()}
            return {str(product.id) for product in product_model.browse(numeric_ids).exists()}
        if product_field.endswith("/id"):
            xml_ids = self.env["ir.model.data"].sudo().search([
                ("model", "=", "product.product"),
                ("complete_name", "in", list(raw_values)),
            ]).mapped("complete_name")
            return set(xml_ids)

        products = product_model.search([("default_code", "in", list(raw_values))])
        return {
            self._normalize_sale_import_value(product.default_code)
            for product in products
            if product.default_code
        }

    @api.model
    def _sale_import_rejection(self, source_index, row, product_field, field_indexes, reason):
        source_offset = self.env.context.get("sale_import_source_offset", 0)
        header_offset = 2 if self.env.context.get("sale_import_has_headers") else 1
        product_value = ""
        if product_field:
            product_value = str(row[field_indexes[product_field]] or "").strip()
        return {
            "source_row": source_offset + source_index + header_offset,
            "product_id": product_value,
            "reason": reason,
        }

    @api.model
    def _sale_import_missing_value_reason(self, field_name):
        field_label = SALE_IMPORT_FIELD_LABELS_ZH.get(field_name)
        if not field_label:
            root_field_name = field_name.split("/", 1)[0]
            field = self._fields.get(root_field_name)
            field_label = field.string if field else field_name
        return f"必填字段“{field_label}”缺少值。"

    @api.model
    def _normalize_sale_import_value(self, value):
        return str(value or "").strip()

    @api.model
    def _is_empty_import_value(self, value):
        return value is None or (isinstance(value, str) and not value.strip())

    x_platform = fields.Char(string="平台")
    x_channel = fields.Char(string="渠道")
    x_sale_nature = fields.Many2one(
        "stock.subwarehouse.sale.nature",
        string="性质",
        ondelete="restrict",
    )
    x_finance_remark = fields.Char(string="备注")
    x_official_total = fields.Monetary(
        string="总金额",
        compute="_compute_official_total",
        store=True,
        currency_field="currency_id",
        help="按产品当前官方销售标价和订单数量计算；导入值不会覆盖系统计算结果。",
    )
    x_processing_fee = fields.Monetary(
        string="手续费",
        currency_field="currency_id",
        default=0.0,
    )
    x_amount_received = fields.Monetary(
        string="实收",
        currency_field="currency_id",
        default=0.0,
    )
    x_is_external_order = fields.Boolean(
        string="外部订单",
        default=False,
        copy=False,
        index=True,
        tracking=True,
        help="外部订单只参加独立销售统计，不预留或扣减ERP库存，也不生成出库或发票流程。",
    )
    x_external_order_status = fields.Selection(
        [("complete", "已完成")],
        string="外部订单状态",
        compute="_compute_external_order_status",
        store=True,
    )
    x_external_completed_at = fields.Datetime(
        string="外部订单完成时间",
        copy=False,
        readonly=True,
    )
    state = fields.Selection(
        selection_add=[("external_done", "外部订单已完成")],
        ondelete={"external_done": "set null"},
    )
    x_website_checkout_language = fields.Selection(
        [("zh_CN", "中文"), ("en_US", "English")],
        string="网站结账语言",
        copy=False,
    )
    x_website_chinese_pricelist_id = fields.Many2one(
        "product.pricelist",
        string="中文网站价目表",
        copy=False,
        readonly=True,
    )
    x_website_stock_reserved_at = fields.Datetime(
        string="Website Stock Reserved At",
        copy=False,
        readonly=True,
    )
    x_website_stock_reserved_until = fields.Datetime(
        string="网站库存保留至",
        compute="_compute_website_stock_reservation_expiry",
    )
    x_website_stock_reservation_expiry_epoch = fields.Integer(
        string="网站库存保留到期时间戳",
        compute="_compute_website_stock_reservation_expiry",
    )
    x_website_payment_ids = fields.Many2many(
        "account.payment",
        string="网站收款",
        compute="_compute_website_payment_details",
    )
    x_website_payment_count = fields.Integer(
        string="网站收款数量",
        compute="_compute_website_payment_details",
    )
    x_website_payment_state = fields.Selection(
        [
            ("unpaid", "未支付"),
            ("pending", "支付处理中"),
            ("paid", "已支付"),
            ("expired", "支付已过期"),
            ("error", "支付失败"),
        ],
        string="网站支付状态",
        compute="_compute_website_payment_details",
    )
    x_website_payment_expired_at = fields.Datetime(
        string="支付订单过期时间",
        copy=False,
        readonly=True,
        index=True,
    )
    x_website_delivery_state = fields.Selection(
        [
            ("awaiting_delivery", "待发货"),
            ("delivering", "配送中"),
            ("delivered", "已送达"),
        ],
        string="网站配送状态",
        copy=False,
        tracking=True,
        index=True,
    )
    x_website_delivery_started_at = fields.Datetime(
        string="开始配送时间",
        copy=False,
        readonly=True,
    )
    x_website_delivered_at = fields.Datetime(
        string="确认送达时间",
        copy=False,
        readonly=True,
    )
    x_pending_website_delivery_count = fields.Integer(
        string="待发货",
        compute="_compute_pending_website_delivery_count",
    )
    x_website_payment_reference = fields.Char(
        string="支付交易号",
        compute="_compute_website_payment_details",
    )
    x_website_refund_request_ids = fields.One2many(
        "stock.subwarehouse.website.refund.request", "order_id", string="网站退款申请"
    )
    x_website_refund_request_count = fields.Integer(
        compute="_compute_website_refund_request_count", string="退款申请数量"
    )
    x_pending_website_refund_request_count = fields.Integer(
        compute="_compute_website_refund_request_count", string="待处理退款"
    )

    @api.depends(
        "order_line.display_type",
        "order_line.product_id",
        "order_line.product_uom_qty",
        "currency_id",
        "company_id",
        "date_order",
    )
    def _compute_official_total(self):
        for order in self:
            conversion_date = order.date_order or fields.Datetime.now()
            company_currency = order.company_id.currency_id
            total = 0.0
            for line in order.order_line.filtered(
                lambda sale_line: not sale_line.display_type and sale_line.product_id
            ):
                listed_price = company_currency._convert(
                    line.product_id.lst_price,
                    order.currency_id,
                    order.company_id,
                    conversion_date,
                    round=False,
                )
                total += listed_price * line.product_uom_qty
            order.x_official_total = order.currency_id.round(total)

    @api.depends("x_is_external_order")
    def _compute_external_order_status(self):
        for order in self:
            order.x_external_order_status = "complete" if order.x_is_external_order else False

    @api.model_create_multi
    def create(self, vals_list):
        orders = super().create(vals_list)
        if not self.env.context.get("skip_external_order_finalization"):
            orders.filtered("x_is_external_order")._finalize_external_orders()
        return orders

    def write(self, vals):
        previous_delivery_states = {
            order.id: order.x_website_delivery_state for order in self
        }
        result = super().write(vals)
        if (
            not self.env.context.get("skip_external_order_finalization")
            and ({"x_is_external_order", "order_line"} & set(vals))
        ):
            self.filtered("x_is_external_order")._finalize_external_orders()
        if (
            "x_website_delivery_state" in vals
            and not self.env.context.get("skip_website_delivery_notifications")
        ):
            entered_queue = self.filtered(
                lambda order: order.x_website_delivery_state == "awaiting_delivery"
                and previous_delivery_states.get(order.id) != "awaiting_delivery"
            )
            left_queue = self.filtered(
                lambda order: previous_delivery_states.get(order.id) == "awaiting_delivery"
                and order.x_website_delivery_state != "awaiting_delivery"
            )
            entered_queue._notify_website_delivery_reviewers()
            left_queue._complete_website_delivery_activities()
        return result

    def _finalize_external_orders(self):
        """Finish external sales without invoking stock/account workflows."""
        external_orders = self.filtered("x_is_external_order")
        if not external_orders:
            return True
        now = fields.Datetime.now()
        for order in external_orders:
            values = {"state": "external_done"}
            if not order.x_external_completed_at:
                values["x_external_completed_at"] = now
            order.with_context(
                skip_external_order_finalization=True,
                skip_procurement=True,
                tracking_disable=True,
            ).write(values)
            order.order_line.with_context(skip_procurement=True).write({
                "x_source_location_id": False,
                "x_website_stock_reserved_until": False,
            })
            for line in order.order_line.filtered(lambda sale_line: not sale_line.display_type):
                # No picking is deliberately generated for an external order,
                # so delivered quantity must be recorded explicitly.
                line.with_context(skip_procurement=True).qty_delivered = line.product_uom_qty
        return True

    @api.depends("picking_ids", "picking_ids.state", "x_is_external_order")
    def _compute_delivery_status(self):
        external_orders = self.filtered("x_is_external_order")
        regular_orders = self - external_orders
        if regular_orders:
            super(SaleOrder, regular_orders)._compute_delivery_status()
        external_orders.delivery_status = "full"

    @api.depends("x_website_refund_request_ids.review_state")
    def _compute_website_refund_request_count(self):
        for order in self:
            order.x_website_refund_request_count = len(order.x_website_refund_request_ids)
            order.x_pending_website_refund_request_count = len(
                order.x_website_refund_request_ids.filtered(
                    lambda request: request.review_state == "requested"
                )
            )

    @api.depends("order_line.x_website_stock_reserved_until")
    def _compute_website_stock_reservation_expiry(self):
        for order in self:
            expiries = [
                expiry
                for expiry in order.order_line.mapped("x_website_stock_reserved_until")
                if expiry
            ]
            expiry = max(expiries) if expiries else False
            order.x_website_stock_reserved_until = expiry
            order.x_website_stock_reservation_expiry_epoch = (
                timegm(expiry.timetuple()) * 1000 if expiry else 0
            )

    @api.depends(
        "transaction_ids.state",
        "transaction_ids.provider_reference",
        "transaction_ids.payment_id",
        "x_website_payment_expired_at",
    )
    def _compute_website_payment_details(self):
        for order in self:
            payments = order.transaction_ids.mapped("payment_id")
            latest_tx = order.transaction_ids.sorted("id")[-1:]
            order.x_website_payment_ids = payments
            order.x_website_payment_count = len(payments)
            order.x_website_payment_reference = (
                latest_tx.provider_reference or latest_tx.reference
            ) if latest_tx else False
            if order.transaction_ids.filtered(lambda tx: tx.state == "done"):
                order.x_website_payment_state = "paid"
            elif order.x_website_payment_expired_at:
                order.x_website_payment_state = "expired"
            elif order.transaction_ids.filtered(lambda tx: tx.state in ("draft", "pending", "authorized")):
                order.x_website_payment_state = "pending"
            elif order.transaction_ids:
                order.x_website_payment_state = "error"
            else:
                order.x_website_payment_state = "unpaid"

    @api.depends("x_website_delivery_state")
    def _compute_pending_website_delivery_count(self):
        for order in self:
            order.x_pending_website_delivery_count = int(
                order.x_website_delivery_state == "awaiting_delivery"
            )

    def _website_delivery_reviewer_users(self):
        stock_managers = self.env.ref(
            "stock.group_stock_manager", raise_if_not_found=False,
        )
        users = stock_managers.all_user_ids if stock_managers else self.env["res.users"]
        users = users.filtered(lambda user: user.active and not user.share)
        if users:
            return users
        sales_managers = self.env.ref(
            "sales_team.group_sale_manager", raise_if_not_found=False,
        )
        users = sales_managers.all_user_ids if sales_managers else self.env["res.users"]
        users = users.filtered(lambda user: user.active and not user.share)
        if users:
            return users
        administrator = self.env.ref("base.user_admin", raise_if_not_found=False)
        return administrator if administrator and administrator.active else self.env.user

    def _notify_website_delivery_reviewers(self):
        activity_type = self.env.ref("mail.mail_activity_data_todo", raise_if_not_found=False)
        if not activity_type:
            return
        summary = _("已支付订单待发货")
        for order in self.filtered(
            lambda item: item.x_website_delivery_state == "awaiting_delivery"
        ):
            reviewers = order._website_delivery_reviewer_users().filtered(
                lambda user: not user.company_ids or order.company_id in user.company_ids
            )
            for reviewer in reviewers:
                existing = order.activity_ids.filtered(
                    lambda activity: activity.activity_type_id == activity_type
                    and activity.user_id == reviewer
                    and activity.summary == summary
                )
                if not existing:
                    order.activity_schedule(
                        act_type_xmlid="mail.mail_activity_data_todo",
                        user_id=reviewer.id,
                        summary=summary,
                        note=_(
                            "订单 %(order)s 已支付，正在等待发货。请按下单时间顺序处理。",
                            order=order.name,
                        ),
                    )
            order.message_post(
                body=_(
                    "订单已完成支付并加入待发货队列。库存将在开始配送时扣减。"
                ),
                message_type="comment",
                subtype_xmlid="mail.mt_note",
                partner_ids=reviewers.partner_id.ids,
            )

    def _complete_website_delivery_activities(self):
        summary = _("已支付订单待发货")
        activities = self.mapped("activity_ids").filtered(
            lambda activity: activity.summary == summary
        )
        if activities:
            activities.sudo().action_feedback(feedback=_("订单已开始配送。"))

    def _queue_paid_website_delivery(self):
        """Idempotently place paid, physical website orders in the FIFO queue."""
        for order in self:
            if (
                order.x_is_external_order
                or not order.website_id
                or order.x_website_payment_state != "paid"
                or not order.order_line.filtered(
                    lambda line: not line.display_type
                    and not line.is_delivery
                    and line.product_id
                    and line.is_storable
                    and line.product_uom_qty > 0
                )
            ):
                continue
            if order.x_website_delivery_state:
                continue
            outgoing = order.picking_ids.filtered(
                lambda picking: picking.picking_type_code == "outgoing"
                and picking.state != "cancel"
            )
            values = {"x_website_delivery_state": "awaiting_delivery"}
            if outgoing and all(picking.state == "done" for picking in outgoing):
                completed_dates = [
                    date_done for date_done in outgoing.mapped("date_done") if date_done
                ]
                values = {
                    "x_website_delivery_state": "delivered",
                    "x_website_delivery_started_at": min(completed_dates)
                    if completed_dates else fields.Datetime.now(),
                    "x_website_delivered_at": max(completed_dates)
                    if completed_dates else fields.Datetime.now(),
                }
            order.write(values)
        return True

    def _website_payment_deadline_is_expired(self, now=None):
        self.ensure_one()
        now = now or fields.Datetime.now()
        deadline = self.x_website_stock_reserved_until
        return bool(
            self.x_website_payment_state != "paid"
            and (self.x_website_payment_expired_at or (deadline and deadline <= now))
        )

    def _assert_website_payment_reservation_active(self):
        for order in self:
            if order._website_payment_deadline_is_expired():
                raise UserError(_("该订单的15分钟库存预留已过期，不能继续支付。"))
            if not order.x_website_stock_reserved_until:
                raise UserError(_("该订单没有有效的库存预留，不能继续支付。"))
        return True

    def _expire_website_payment(self):
        """Expire unpaid orders without releasing or renewing them optimistically."""
        now = fields.Datetime.now()
        for order in self:
            self.env.cr.execute(
                "SELECT id FROM sale_order WHERE id = %s FOR UPDATE",
                [order.id],
            )
            order.invalidate_recordset()
            if (
                order.x_website_payment_state == "paid"
                or order.x_website_payment_expired_at
                or not order.x_website_stock_reserved_until
                or order.x_website_stock_reserved_until > now
            ):
                continue
            pending_transactions = order.transaction_ids.filtered(
                lambda transaction: transaction.state in ("draft", "pending", "authorized")
            )
            if pending_transactions:
                pending_transactions._set_canceled(
                    state_message=_("15分钟支付期限已过，库存预留已释放。")
                )
            order.with_context(
                skip_website_delivery_notifications=True,
                shop_api_skip_event=True,
            ).write({"x_website_payment_expired_at": now})
            if order.state in ("draft", "sent"):
                order.with_context(shop_api_skip_event=True).action_cancel()
            order.message_post(
                body=_("订单因15分钟支付期限届满而过期，库存预留已自动释放。"),
                message_type="comment",
                subtype_xmlid="mail.mt_note",
            )
        return True

    def action_start_website_delivery(self):
        """Deduct stock at dispatch and move the order into the in-transit state."""
        for order in self:
            if order.x_website_payment_state != "paid":
                raise UserError(_("只有已支付订单可以开始配送。"))
            if order.x_website_delivery_state != "awaiting_delivery":
                raise UserError(_("只有待发货订单可以开始配送。"))
            if order.state in ("draft", "sent"):
                order.action_confirm()
            outgoing = order.picking_ids.filtered(
                lambda picking: picking.picking_type_code == "outgoing"
                and picking.state != "cancel"
            )
            if not outgoing:
                raise UserError(_("该订单没有可执行的客户发货单。"))
            order.write({
                "x_website_delivery_state": "delivering",
                "x_website_delivery_started_at": fields.Datetime.now(),
            })
            for picking in outgoing.filtered(lambda item: item.state != "done"):
                picking.action_assign()
                if picking.state != "assigned":
                    raise UserError(_("发货库存尚未完整预留，无法开始配送。"))
                for move in picking.move_ids.filtered(
                    lambda stock_move: stock_move.product_id.tracking == "none"
                ):
                    move.quantity = move.product_uom_qty
                    move.picked = True
                picking.with_context(shop_api_skip_event=True).button_validate()
                if picking.state != "done":
                    raise UserError(_("发货单尚未完成，请先处理发货单中的待确认步骤。"))
            order.message_post(
                body=_("配送已开始；发货库存已经扣减，订单当前处于运输途中。"),
                message_type="comment",
                subtype_xmlid="mail.mt_note",
            )
        return {"type": "ir.actions.client", "tag": "reload"}

    def action_mark_website_delivery_delivered(self):
        for order in self:
            if order.x_website_delivery_state != "delivering":
                raise UserError(_("只有配送中的订单可以确认送达。"))
            incomplete = order.picking_ids.filtered(
                lambda picking: picking.picking_type_code == "outgoing"
                and picking.state not in ("done", "cancel")
            )
            if incomplete:
                raise UserError(_("仍有未完成的发货单，不能确认订单已送达。"))
            order.write({
                "x_website_delivery_state": "delivered",
                "x_website_delivered_at": fields.Datetime.now(),
            })
            order.message_post(
                body=_("订单已确认送达客户。"),
                message_type="comment",
                subtype_xmlid="mail.mt_note",
            )
        return {"type": "ir.actions.client", "tag": "reload"}

    def action_view_pending_website_deliveries(self):
        return self.env["ir.actions.actions"]._for_xml_id(
            "stock_subwarehouse_hierarchy.action_pending_website_deliveries"
        )

    def _get_website_payment_receipt(self):
        self.ensure_one()
        return self.transaction_ids.filtered(
            lambda tx: tx.state == "done" and tx.payment_id
        ).sorted("id")[-1:].payment_id

    @api.model
    def _is_website_checkout_country_allowed(self, country, is_english):
        chinese_region_codes = {"CN", "HK", "MO", "TW"}
        country_code = (country.code or "").upper()
        return bool(country_code) and (
            country_code not in chinese_region_codes if is_english else country_code in chinese_region_codes
        )

    @api.model
    def _website_checkout_country_message(self, is_english):
        return (
            _("Please switch the website language to Chinese; this delivery country or region is not supported.")
            if is_english
            else _("请更换网站语言至英文，否则不支持此收货国家/地域。")
        )

    def _get_website_usd_pricelist(self):
        self.ensure_one()
        usd = self.env.ref("base.USD")
        if usd.symbol != "$" or usd.position != "before" or not usd.active:
            usd.write({"symbol": "$", "position": "before", "active": True})
        pricelist = self.env["product.pricelist"].sudo().search([
            ("website_id", "=", self.website_id.id),
            ("name", "=", "SUN Website USD Checkout"),
        ], limit=1)
        if not pricelist:
            pricelist = self.env["product.pricelist"].sudo().create({
                "name": "SUN Website USD Checkout",
                "currency_id": usd.id,
                "company_id": self.company_id.id,
                "website_id": self.website_id.id,
                "selectable": False,
            })
        elif pricelist.currency_id != usd or pricelist.selectable:
            pricelist.write({"currency_id": usd.id, "selectable": False})
        return pricelist

    def _apply_website_checkout_language(self, is_english):
        for order in self.filtered("website_id"):
            target_language = "en_US" if is_english else "zh_CN"
            if is_english:
                missing_mappings = order.order_line.filtered(
                    lambda line: (
                        not line.display_type
                        and not line.is_delivery
                        and line.product_id
                        and not line.product_id.product_tmpl_id.x_website_code_mapping_id
                    )
                )
                if missing_mappings:
                    raise UserError(_(
                        "以下产品没有英文网站编号价格规则，无法使用英文结账：%s"
                    ) % ", ".join(missing_mappings.mapped("product_id.display_name")))
                if not order.x_website_chinese_pricelist_id:
                    order.x_website_chinese_pricelist_id = order.pricelist_id
                usd_pricelist = order._get_website_usd_pricelist()
                if order.pricelist_id != usd_pricelist:
                    order.pricelist_id = usd_pricelist
                for line in order.order_line.filtered(
                    lambda line: not line.display_type and not line.is_delivery and line.product_id
                ):
                    template = line.product_id.product_tmpl_id
                    line.write({
                        "name": template._get_website_display_name(True),
                        "price_unit": template.x_website_usd_price,
                    })
            else:
                chinese_pricelist = order.x_website_chinese_pricelist_id
                if chinese_pricelist and order.pricelist_id != chinese_pricelist:
                    order.pricelist_id = chinese_pricelist
                # The price recomputation hook checks this flag to decide whether
                # it must enforce USD mapping values.
                order.x_website_checkout_language = target_language
                order._recompute_prices()
                for line in order.order_line.filtered(
                    lambda line: not line.display_type and not line.is_delivery and line.product_id
                ):
                    line.name = line.product_id.with_context(lang="zh_CN").get_product_multiline_description_sale()
            order.x_website_checkout_language = target_language

    def _recompute_prices(self):
        """Keep code-mapped USD prices after Odoo refreshes a website cart."""
        super()._recompute_prices()
        for order in self.filtered(
            lambda order: order.website_id and order.x_website_checkout_language == "en_US"
        ):
            for line in order.order_line.filtered(
                lambda line: not line.display_type and not line.is_delivery and line.product_id
            ):
                template = line.product_id.product_tmpl_id
                line.write({
                    "name": template._get_website_display_name(True),
                    "price_unit": template.x_website_usd_price,
                })

    def action_view_website_payments(self):
        self.ensure_one()
        action = self.env["ir.actions.actions"]._for_xml_id("account.action_account_payments")
        payments = self.x_website_payment_ids
        if len(payments) == 1:
            action.update({
                "view_mode": "form",
                "res_id": payments.id,
                "views": [(False, "form")],
            })
        else:
            action["domain"] = [("id", "in", payments.ids)]
        return action

    def action_view_website_refund_requests(self):
        self.ensure_one()
        action = self.env["ir.actions.actions"]._for_xml_id(
            "stock_subwarehouse_hierarchy.action_website_refund_requests"
        )
        action["domain"] = [("order_id", "=", self.id)]
        return action

    def action_refund_website_payment(self):
        self.ensure_one()
        payment = self._get_website_payment_receipt()
        if not payment:
            raise UserError(_("该网站订单没有可退款的已完成支付。"))
        transaction = payment.payment_transaction_id
        if (
            transaction.provider_code not in ("wechatpay", "alipay")
            or not transaction.provider_id.support_refund
        ):
            raise UserError(_("该网站订单的支付方式暂不支持从此处退款。"))
        return {
            "type": "ir.actions.act_window",
            "name": _("支付退款"),
            "res_model": "stock.subwarehouse.website.payment.refund.wizard",
            "view_mode": "form",
            "target": "new",
            "context": {
                "default_order_id": self.id,
                "default_transaction_id": transaction.id,
            },
        }

    @api.model
    def get_import_templates(self):
        return [{
            "label": _("报价单导入模板（产品ID）"),
            "template": SALE_ORDER_IMPORT_TEMPLATE_ROUTE,
        }]

    def _get_sale_order_import_template_columns(self):
        return [
            ("Order Reference", "订单号"),
            ("Customer*", "客户"),
            ("Order Date", "下单时间"),
            ("x_platform", "平台"),
            ("x_channel", "渠道"),
            ("Salesperson", "销售人员"),
            ("x_sale_nature", "性质"),
            ("Order Lines/Products*", "产品ID"),
            ("Order Lines/x_import_product_name", "品名"),
            ("Order Lines/x_color", "颜色"),
            ("Order Lines/x_size", "尺码"),
            ("Order Lines/x_flex", "款型"),
            ("Order Lines/Quantity", "数量"),
            ("Order Lines/Unit Price", "单价"),
            ("Order Lines/x_source_location_id", "发货仓库"),
            ("x_official_total", "总金额（系统按官方标价计算）"),
            ("x_processing_fee", "手续费"),
            ("x_amount_received", "实收"),
            ("x_is_external_order", "外部订单（True/False）"),
            ("x_finance_remark", "备注"),
        ]

    def _generate_sale_order_import_template_xlsx(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as error:
            raise ImportError(_("生成导入模板需要安装 openpyxl。")) from error

        workbook = Workbook()
        import_sheet = workbook.active
        import_sheet.title = "报价单导入"
        columns = self._get_sale_order_import_template_columns()
        import_sheet.append([field_name for field_name, _label in columns])
        import_sheet.append([label for _field_name, label in columns])
        import_sheet.append([
            "S00001",
            self.env.user.partner_id.display_name,
            fields.Date.today().strftime("%Y-%m-%d"),
            "有赞",
            "凌动雪具",
            self.env.user.display_name,
            "零售",
            "152410Yb-MK000-H001150",
            "双板鞋",
            "黑色",
            "260",
            "硬度100",
            10,
            111,
            "张家口/Stock",
            "",
            10,
            1100,
            "False",
            "",
        ])
        import_sheet.append([
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "072409Y-MA000-G001##S",
            "滑雪服",
            "绿色",
            "S",
            "",
            1,
            4000,
            "",
            "",
            "",
            "",
            "",
            "",
        ])

        field_sheet = workbook.create_sheet("导入字段")
        field_sheet.append(["字段", "中文说明"])
        for field_name, label in columns:
            field_sheet.append([field_name, label])

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A3"
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
            if sheet.max_row >= 2:
                for cell in sheet[2]:
                    cell.font = Font(italic=True)
                    cell.fill = PatternFill("solid", fgColor="E2F0D9")
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 55)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def action_export_import_template_format(self):
        ids = ",".join(str(record_id) for record_id in self.ids)
        return {
            "type": "ir.actions.act_url",
            "url": f"/stock_subwarehouse_hierarchy/export/sale_order.xlsx?{urlencode({'ids': ids})}",
            "target": "self",
        }

    def _generate_sale_order_export_xlsx(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as error:
            raise ImportError(_("生成导出文件需要安装 openpyxl。")) from error

        workbook = Workbook()
        export_sheet = workbook.active
        export_sheet.title = "报价单导出"
        columns = self._get_sale_order_import_template_columns()
        export_sheet.append([field_name for field_name, _label in columns])
        export_sheet.append([label for _field_name, label in columns])

        for order in self:
            order_lines = order.order_line.filtered(lambda line: not line.display_type)
            if not order_lines:
                export_sheet.append(self._sale_order_export_row(order, self.env["sale.order.line"], include_order=True))
                continue
            for index, line in enumerate(order_lines):
                export_sheet.append(self._sale_order_export_row(order, line, include_order=index == 0))

        self._format_sale_order_export_workbook(workbook)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _sale_order_export_row(self, order, line, include_order=True):
        product = line.product_id if line else self.env["product.product"]
        return [
            order.name if include_order else "",
            order.partner_id.display_name if include_order else "",
            order.date_order.strftime("%Y-%m-%d") if include_order and order.date_order else "",
            order.x_platform if include_order else "",
            order.x_channel if include_order else "",
            order.user_id.display_name if include_order and order.user_id else "",
            order.x_sale_nature.display_name if include_order and order.x_sale_nature else "",
            product.default_code or product.display_name or "",
            line.x_import_product_name if line else "",
            line.x_color if line else "",
            line.x_size if line else "",
            line.x_flex if line else "",
            line.product_uom_qty if line else "",
            line.price_unit if line else "",
            line.x_source_location_id.display_name if line and line.x_source_location_id else "",
            order.x_official_total if include_order else "",
            order.x_processing_fee if include_order else "",
            order.x_amount_received if include_order else "",
            order.x_is_external_order if include_order else "",
            order.x_finance_remark if include_order else "",
        ]

    def _format_sale_order_export_workbook(self, workbook):
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A3"
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
            if sheet.max_row >= 2:
                for cell in sheet[2]:
                    cell.font = Font(italic=True)
                    cell.fill = PatternFill("solid", fgColor="E2F0D9")
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 55)

    def _check_source_inventory_availability(self):
        precision = self.env["decimal.precision"].precision_get("Product Unit")
        demands = {}
        for line in self.filtered(lambda order: not order.x_is_external_order).order_line:
            if (
                line.display_type
                or not line.is_storable
                or not line.x_source_location_id
            ):
                continue
            key = (line.product_id, line.x_source_location_id)
            demands.setdefault(key, {
                "line": line,
                "requested": 0.0,
            })
            demands[key]["requested"] += line.product_uom_id._compute_quantity(
                line.product_uom_qty,
                line.product_id.uom_id,
            )

        shortages = []
        for line in self.order_line:
            if (
                line.display_type
                or not line.is_storable
                or float_compare(line.product_uom_qty, 0.0, precision_digits=precision) <= 0
                or line.x_source_location_id
            ):
                continue
            shortages.append(_(
                "%(product)s：请选择有足够现货的来源库存。",
                product=line.product_id.display_name,
            ))
        for (product, location), demand in demands.items():
            line = demand["line"]
            available = self.env["stock.quant"]._get_available_quantity(product, location, strict=True)
            if float_compare(
                demand["requested"],
                available,
                precision_digits=precision,
            ) > 0:
                shortages.append(_(
                    "%(product)s 来自 %(location)s：需要 %(requested)s %(uom)s，可用 %(available)s %(uom)s",
                    product=product.display_name,
                    location=location.display_name,
                    requested=demand["requested"],
                    available=available,
                    uom=product.uom_id.display_name,
                ))
        if shortages:
            raise UserError(_("所选来源库存无法满足此报价单：\n%s") % "\n".join(shortages))

    def _get_website_reserved_qty_for_source_location(self, product, location, exclude_order=False):
        now = fields.Datetime.now()
        domain = [
            ("product_id", "=", product.id),
            ("x_source_location_id", "=", location.id),
            ("x_website_stock_reserved_until", ">", now),
            ("order_id.state", "in", ["draft", "sent"]),
        ]
        if exclude_order:
            domain.append(("order_id", "!=", exclude_order.id))
        reserved_qty = 0.0
        reserved_lines = self.env["sale.order.line"].sudo().search(domain)
        for line in reserved_lines:
            reserved_qty += line.product_uom_id._compute_quantity(
                line.product_uom_qty,
                product.uom_id,
            )
        return reserved_qty

    def _get_available_qty_for_source_location(self, product, location, exclude_order=False):
        physical_qty = self.env["stock.quant"].sudo()._get_available_quantity(
            product,
            location,
            strict=True,
        )
        reserved_qty = self._get_website_reserved_qty_for_source_location(
            product,
            location,
            exclude_order=exclude_order,
        )
        return max(physical_qty - reserved_qty, 0.0)

    def _auto_assign_website_source_locations(self):
        precision = self.env["decimal.precision"].precision_get("Product Unit")
        StockLocation = self.env["stock.location"]
        for order in self:
            planned_demands = {}
            lines = order.order_line.filtered(
                lambda line: (
                    not line.display_type
                    and line.product_id
                    and line.is_storable
                    and float_compare(line.product_uom_qty, 0.0, precision_digits=precision) > 0
                )
            )
            for line in lines:
                product = line.product_id
                required_qty = line._get_required_qty_in_product_uom()
                current_location = line.x_source_location_id
                if current_location:
                    planned_key = (product.id, current_location.id)
                    available_qty = order._get_available_qty_for_source_location(
                        product,
                        current_location,
                        exclude_order=order,
                    ) - planned_demands.get(planned_key, 0.0)
                    if float_compare(available_qty, required_qty, precision_digits=precision) >= 0:
                        planned_demands[planned_key] = planned_demands.get(planned_key, 0.0) + required_qty
                        continue

                best_location = StockLocation
                best_available_qty = False
                for location in line._get_source_location_candidates():
                    planned_key = (product.id, location.id)
                    available_qty = order._get_available_qty_for_source_location(
                        product,
                        location,
                        exclude_order=order,
                    ) - planned_demands.get(planned_key, 0.0)
                    if float_compare(available_qty, required_qty, precision_digits=precision) < 0:
                        continue
                    if best_available_qty is False or available_qty < best_available_qty:
                        best_location = location
                        best_available_qty = available_qty

                line.x_source_location_id = best_location
                if best_location:
                    planned_key = (product.id, best_location.id)
                    planned_demands[planned_key] = planned_demands.get(planned_key, 0.0) + required_qty

    def _get_source_inventory_shortages(self):
        precision = self.env["decimal.precision"].precision_get("Product Unit")
        demands = {}
        shortages = []
        for line in self.order_line:
            if (
                line.display_type
                or not line.is_storable
                or float_compare(line.product_uom_qty, 0.0, precision_digits=precision) <= 0
            ):
                continue
            if not line.x_source_location_id:
                shortages.append(_("%(product)s：没有可满足数量的发货仓库。", product=line.product_id.display_name))
                continue
            key = (line.product_id, line.x_source_location_id)
            demands.setdefault(key, {
                "line": line,
                "requested": 0.0,
            })
            demands[key]["requested"] += line.product_uom_id._compute_quantity(
                line.product_uom_qty,
                line.product_id.uom_id,
            )

        for (product, location), demand in demands.items():
            available = self._get_available_qty_for_source_location(
                product,
                location,
                exclude_order=self,
            )
            if float_compare(demand["requested"], available, precision_digits=precision) > 0:
                shortages.append(_(
                    "%(product)s 来自 %(location)s：需要 %(requested)s %(uom)s，可用 %(available)s %(uom)s",
                    product=product.display_name,
                    location=location.display_name,
                    requested=demand["requested"],
                    available=available,
                    uom=product.uom_id.display_name,
                ))
        return shortages

    def _get_source_inventory_shortage_lines(self):
        """Return cart lines whose full quantity cannot be sourced from one location."""
        precision = self.env["decimal.precision"].precision_get("Product Unit")
        shortage_lines = self.env["sale.order.line"]
        for order in self:
            planned_demands = {}
            lines = order.order_line.filtered(
                lambda line: (
                    not line.display_type
                    and line.product_id
                    and line.is_storable
                    and float_compare(line.product_uom_qty, 0.0, precision_digits=precision) > 0
                )
            )
            for line in lines:
                product = line.product_id
                required_qty = line._get_required_qty_in_product_uom()
                candidates = line._get_source_location_candidates()
                current_location = line.x_source_location_id
                if current_location and current_location in candidates:
                    candidates = current_location | (candidates - current_location)

                selected_location = self.env["stock.location"]
                selected_available_qty = False
                for location in candidates:
                    planned_key = (product.id, location.id)
                    available_qty = order._get_available_qty_for_source_location(
                        product,
                        location,
                        exclude_order=order,
                    ) - planned_demands.get(planned_key, 0.0)
                    if float_compare(available_qty, required_qty, precision_digits=precision) < 0:
                        continue
                    if selected_available_qty is False or available_qty < selected_available_qty:
                        selected_location = location
                        selected_available_qty = available_qty

                if not selected_location:
                    shortage_lines |= line
                    continue
                planned_key = (product.id, selected_location.id)
                planned_demands[planned_key] = (
                    planned_demands.get(planned_key, 0.0) + required_qty
                )
        return shortage_lines

    def _check_source_inventory_availability(self):
        shortages = self._get_source_inventory_shortages()
        if shortages:
            raise UserError(_("所选来源库存无法满足此报价单：\n%s") % "\n".join(shortages))

    def _prepare_website_stock_for_payment(self, hold_minutes=15):
        regular_orders = self.filtered(lambda order: not order.x_is_external_order)
        regular_orders._auto_assign_website_source_locations()
        regular_orders._check_source_inventory_availability()
        now = fields.Datetime.now()
        for order in regular_orders:
            reservable_lines = order.order_line.filtered(
                lambda line: (
                    not line.display_type
                    and line.product_id
                    and line.is_storable
                    and line.x_source_location_id
                    and line.product_uom_qty > 0
                )
            )
            previous_deadlines = reservable_lines.mapped("x_website_stock_reserved_until")
            if any(deadline and deadline <= now for deadline in previous_deadlines):
                raise UserError(_("该订单的15分钟库存预留已过期，请重新创建订单。"))
            active_deadlines = [deadline for deadline in previous_deadlines if deadline and deadline > now]
            if active_deadlines:
                reserved_until = min(active_deadlines)
            else:
                reserved_until = now + relativedelta(minutes=hold_minutes)
                reservable_lines.write({"x_website_stock_reserved_until": reserved_until})
                order.x_website_stock_reserved_at = now
            if order.x_website_payment_expired_at:
                raise UserError(_("该订单已经过期，不能刷新库存预留。"))

    def action_bulk_confirm(self):
        """Confirm selected quotations independently and report partial failures."""
        eligible_orders = self.filtered(lambda order: order.state in ("draft", "sent"))
        ignored_count = len(self) - len(eligible_orders)
        confirmed_count = 0
        failures = []

        for order in eligible_orders:
            try:
                with self.env.cr.savepoint():
                    order.action_confirm()
                confirmed_count += 1
            except UserError as error:
                failures.append((order.display_name, str(error)))

        message_parts = [f"已成功确认 {confirmed_count} 个销售订单。"]
        if ignored_count:
            message_parts.append(f"已忽略 {ignored_count} 个非待确认状态的订单。")
        if failures:
            message_parts.append(f"有 {len(failures)} 个订单确认失败：")
            message_parts.extend(
                f"{order_name}：{reason}"
                for order_name, reason in failures[:10]
            )
            if len(failures) > 10:
                message_parts.append(f"另有 {len(failures) - 10} 个失败订单未在此处展开。")

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": "批量确认完成" if not failures else "批量确认完成（部分失败）",
                "message": "\n".join(message_parts),
                "type": "success" if not failures else "warning",
                "sticky": bool(failures),
                "next": {"type": "ir.actions.client", "tag": "reload"},
            },
        }

    def action_bulk_archive(self):
        """Archive selected sales independently and report partial failures."""
        eligible_orders = self.filtered("active")
        ignored_count = len(self) - len(eligible_orders)
        archived_count = 0
        failures = []

        for order in eligible_orders:
            order_name = order.display_name
            try:
                with self.env.cr.savepoint():
                    order.action_archive()
                archived_count += 1
            except UserError as error:
                failures.append((order_name, str(error)))

        message_parts = [f"已成功归档 {archived_count} 个销售订单。"]
        if ignored_count:
            message_parts.append(f"已忽略 {ignored_count} 个已归档订单。")
        if failures:
            message_parts.append(f"有 {len(failures)} 个订单归档失败：")
            message_parts.extend(
                f"{order_name}：{reason}"
                for order_name, reason in failures[:10]
            )
            if len(failures) > 10:
                message_parts.append(f"另有 {len(failures) - 10} 个失败订单未在此处展开。")

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": "批量归档完成" if not failures else "批量归档完成（部分失败）",
                "message": "\n".join(message_parts),
                "type": "success" if not failures else "warning",
                "sticky": bool(failures),
                "next": {"type": "ir.actions.client", "tag": "reload"},
            },
        }

    def action_confirm(self):
        external_orders = self.filtered("x_is_external_order")
        external_orders._finalize_external_orders()
        regular_orders_to_confirm = self - external_orders
        regular_orders_to_confirm._check_source_inventory_availability()
        website_orders = regular_orders_to_confirm.filtered(
            lambda order: order.website_id and self.env.context.get("send_email")
        )
        regular_orders = regular_orders_to_confirm - website_orders
        result = True
        if regular_orders:
            result = super(SaleOrder, regular_orders).action_confirm()
        if website_orders:
            result = super(SaleOrder, website_orders.with_context(send_email=False)).action_confirm()
            try:
                with self.env.cr.savepoint():
                    website_orders._send_order_confirmation_mail()
            except (UserError, MailDeliveryException):
                _logger.exception(
                    "Website orders %s were confirmed after payment, but their confirmation email failed.",
                    website_orders.mapped("name"),
                )
        regular_orders_to_confirm._queue_paid_website_delivery()
        return result

    def _create_invoices(self, grouped=False, final=False, date=None):
        if self.filtered("x_is_external_order"):
            raise UserError("外部订单不参与ERP开票流程。")
        return super()._create_invoices(grouped=grouped, final=final, date=date)


class SaleOrderLine(models.Model):
    _inherit = "sale.order.line"

    def _action_launch_stock_rule(self, *, previous_product_uom_qty=False):
        internal_lines = self.filtered(lambda line: not line.order_id.x_is_external_order)
        if not internal_lines:
            return True
        return super(SaleOrderLine, internal_lines)._action_launch_stock_rule(
            previous_product_uom_qty=previous_product_uom_qty
        )

    @api.depends("product_id", "is_expense", "order_id.x_is_external_order")
    def _compute_qty_delivered_method(self):
        external_lines = self.filtered(lambda line: line.order_id.x_is_external_order)
        regular_lines = self - external_lines
        if regular_lines:
            super(SaleOrderLine, regular_lines)._compute_qty_delivered_method()
        external_lines.qty_delivered_method = "manual"

    @api.depends(
        "qty_delivered_method",
        "analytic_line_ids.so_line",
        "analytic_line_ids.unit_amount",
        "analytic_line_ids.product_uom_id",
        "order_id.x_is_external_order",
        "product_uom_qty",
    )
    def _compute_qty_delivered(self):
        external_lines = self.filtered(lambda line: line.order_id.x_is_external_order)
        regular_lines = self - external_lines
        if regular_lines:
            super(SaleOrderLine, regular_lines)._compute_qty_delivered()
        for line in external_lines:
            line.qty_delivered = line.product_uom_qty

    @api.depends(
        "qty_invoiced",
        "qty_delivered",
        "product_uom_qty",
        "state",
        "order_id.x_is_external_order",
    )
    def _compute_qty_to_invoice(self):
        external_lines = self.filtered(lambda line: line.order_id.x_is_external_order)
        regular_lines = self - external_lines
        if regular_lines:
            super(SaleOrderLine, regular_lines)._compute_qty_to_invoice()
        external_lines.qty_to_invoice = 0.0

    @api.depends("state", "product_uom_qty", "qty_delivered", "qty_to_invoice", "qty_invoiced", "order_id.x_is_external_order")
    def _compute_invoice_status(self):
        external_lines = self.filtered(lambda line: line.order_id.x_is_external_order)
        regular_lines = self - external_lines
        if regular_lines:
            super(SaleOrderLine, regular_lines)._compute_invoice_status()
        external_lines.invoice_status = "no"

    @api.depends("product_id.display_name", "order_id.x_website_checkout_language")
    def _compute_name_short(self):
        super()._compute_name_short()
        for line in self.filtered(
            lambda line: (
                line.order_id.website_id
                and line.order_id.x_website_checkout_language == "en_US"
                and line.product_id
                and not line.is_delivery
            )
        ):
            line.name_short = line.product_id.product_tmpl_id._get_website_display_name(True)

    def _get_line_header(self):
        self.ensure_one()
        if (
            self.order_id.website_id
            and self.order_id.x_website_checkout_language == "en_US"
            and self.product_id
            and not self.is_delivery
        ):
            return self.product_id.product_tmpl_id._get_website_display_name(True)
        return super()._get_line_header()
