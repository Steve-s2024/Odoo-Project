import base64
import json
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from odoo import Command
from odoo.exceptions import UserError
from odoo.tests import TransactionCase, tagged

from odoo.addons.stock_subwarehouse_hierarchy.controllers.website import _amap_search_url
from odoo.addons.stock_subwarehouse_hierarchy.controllers.purchase_history import WebsitePurchaseHistory


@tagged("post_install", "-at_install")
class TestDescendantInventoryTotals(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.StockLocation = cls.env["stock.location"]
        cls.StockQuant = cls.env["stock.quant"]

        cls.warehouse = cls.env["stock.warehouse"].create({
            "name": "Descendant Test Warehouse",
            "code": "DTW",
        })
        cls.product_a = cls.env["product.product"].create({
            "name": "Descendant Product A",
            "is_storable": True,
        })
        cls.product_b = cls.env["product.product"].create({
            "name": "Descendant Product B",
            "is_storable": True,
        })

        cls.subwarehouse = cls.StockLocation.create({
            "name": "Subwarehouse A",
            "usage": "view",
            "location_id": cls.warehouse.view_location_id.id,
        })
        cls.bin_a = cls.StockLocation.create({
            "name": "Subwarehouse A / Bin A",
            "usage": "internal",
            "location_id": cls.subwarehouse.id,
        })
        cls.bin_b = cls.StockLocation.create({
            "name": "Subwarehouse A / Bin B",
            "usage": "internal",
            "location_id": cls.subwarehouse.id,
        })
        cls.supplier_location = cls.env.ref("stock.stock_location_suppliers")
        cls.customer = cls.env["res.partner"].create({
            "name": "Subwarehouse Test Customer",
        })
        cls.cny_pricelist = cls.env["product.pricelist"].create({
            "name": "Descendant Test CNY Pricelist",
            "currency_id": cls.env.ref("base.CNY").id,
        })

    def test_descendant_inventory_totals_sum_child_locations(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        self.StockQuant._update_available_quantity(self.product_a, self.bin_b, 7.0)
        self.StockQuant._update_available_quantity(self.product_b, self.bin_b, 3.0)
        self.StockQuant._update_available_quantity(self.product_a, self.supplier_location, 11.0)

        totals = self.warehouse._get_descendant_inventory_totals()

        self.assertEqual(totals[self.product_a.id]["quantity"], 12.0)
        self.assertEqual(totals[self.product_a.id]["available_quantity"], 12.0)
        self.assertEqual(totals[self.product_b.id]["quantity"], 3.0)
        self.assertNotIn(
            self.supplier_location.id,
            self.StockQuant.search([
                ("location_id", "child_of", self.warehouse.view_location_id.id),
                ("location_id.usage", "=", "internal"),
            ]).mapped("location_id").ids,
        )

    def test_sale_import_rejects_duplicate_order_ids_with_partial_success(self):
        existing = self.env["sale.order"].create({
            "name": "IMPORT-SALE-EXISTING",
            "partner_id": self.customer.id,
        })
        fields_list = ["name", "partner_id"]
        rows = [
            [existing.name.lower(), self.customer.name],
            ["IMPORT-SALE-UNIQUE", self.customer.name],
            [" import-sale-unique ", self.customer.name],
            ["", self.customer.name],
        ]

        result = self.env["sale.order"].with_context(
            import_file=True,
            sale_import_source_offset=0,
            sale_import_has_headers=True,
        ).load(fields_list, rows)

        imported = self.env["sale.order"].browse(result["ids"])
        self.assertEqual(imported.mapped("name"), ["IMPORT-SALE-UNIQUE"])
        self.assertEqual(len(result["x_business_import_failures"]), 3)
        self.assertEqual(
            {failure["source_row"] for failure in result["x_business_import_failures"]},
            {2, 4, 5},
        )

    def test_mrp_import_rejects_duplicate_ids_with_partial_success(self):
        existing = self.env["mrp.production"].create({
            "name": "IMPORT-MO-EXISTING",
            "product_id": self.product_a.id,
            "product_qty": 1.0,
            "product_uom_id": self.product_a.uom_id.id,
            "location_src_id": self.bin_a.id,
            "location_dest_id": self.bin_b.id,
        })
        fields_list = [
            "name", "product_id", "product_qty", "product_uom_id",
            "location_src_id", "location_dest_id",
        ]
        valid_row = [
            "IMPORT-MO-UNIQUE", self.product_a.display_name, 1,
            self.product_a.uom_id.name, self.bin_a.complete_name, self.bin_b.complete_name,
        ]
        rows = [
            [existing.name.lower(), *valid_row[1:]],
            valid_row,
            [" import-mo-unique ", *valid_row[1:]],
            ["", *valid_row[1:]],
        ]

        result = self.env["mrp.production"].with_context(
            import_file=True,
            business_import_source_offset=0,
            business_import_has_headers=True,
        ).load(fields_list, rows)

        imported = self.env["mrp.production"].browse(result["ids"])
        self.assertEqual(imported.mapped("name"), ["IMPORT-MO-UNIQUE"])
        self.assertEqual(len(result["x_business_import_failures"]), 2)

    def test_mrp_import_groups_blank_name_product_rows_as_finished_products(self):
        self.product_a.default_code = "IMPORT-MO-FINISHED-A"
        self.product_b.default_code = "IMPORT-MO-FINISHED-B"
        fields_list = [
            "name", "product_id", "product_qty", "product_uom_id",
            "origin", "location_src_id", "location_dest_id",
        ]
        rows = [
            [
                "IMPORT-MO-MULTI-PRODUCT",
                self.product_a.default_code,
                2,
                self.product_a.uom_id.name,
                "Grouped manufacturing import",
                self.bin_a.complete_name,
                self.bin_b.complete_name,
            ],
            [
                "",
                self.product_b.default_code,
                3,
                self.product_b.uom_id.name,
                "",
                "",
                "",
            ],
        ]

        result = self.env["mrp.production"].with_context(
            import_file=True,
            business_import_source_offset=0,
            business_import_has_headers=True,
        ).load(fields_list, rows)

        self.assertFalse(result["messages"])
        self.assertFalse(result["x_business_import_failures"])
        self.assertEqual(len(result["ids"]), 1)
        production = self.env["mrp.production"].browse(result["ids"])
        self.assertEqual(production.name, "IMPORT-MO-MULTI-PRODUCT")
        self.assertEqual(production.product_id, self.product_a)
        self.assertEqual(production.product_qty, 2)
        byproduct = production.move_byproduct_ids.filtered(
            lambda move: move.product_id == self.product_b
        )
        self.assertEqual(len(byproduct), 1)
        self.assertEqual(byproduct.product_uom_qty, 3)

        production.action_confirm()
        production.qty_producing = 2
        production.button_mark_done()
        self.assertEqual(
            self.StockQuant._get_available_quantity(self.product_a, self.bin_b),
            2,
        )
        self.assertEqual(
            self.StockQuant._get_available_quantity(self.product_b, self.bin_b),
            3,
        )

    def test_mrp_export_keeps_one_name_and_multiple_finished_product_rows(self):
        from openpyxl import load_workbook

        self.product_a.default_code = "EXPORT-MO-FINISHED-A"
        self.product_b.default_code = "EXPORT-MO-FINISHED-B"
        production = self.env["mrp.production"].create({
            "name": "EXPORT-MO-MULTI-PRODUCT",
            "product_id": self.product_a.id,
            "product_qty": 2,
            "product_uom_id": self.product_a.uom_id.id,
            "location_src_id": self.bin_a.id,
            "location_dest_id": self.bin_b.id,
            "move_byproduct_ids": [Command.create({
                "product_id": self.product_b.id,
                "product_uom_qty": 3,
                "product_uom": self.product_b.uom_id.id,
            })],
        })

        workbook = load_workbook(
            BytesIO(production._generate_dynamic_export_xlsx()),
            read_only=True,
        )
        rows = list(workbook["制造单导出"].iter_rows(values_only=True))
        headers = rows[0]
        self.assertEqual(rows[2][headers.index("name")], "EXPORT-MO-MULTI-PRODUCT")
        self.assertEqual(rows[2][headers.index("product_id")], "EXPORT-MO-FINISHED-A")
        self.assertIsNone(rows[3][headers.index("name")])
        self.assertEqual(rows[3][headers.index("product_id")], "EXPORT-MO-FINISHED-B")
        self.assertEqual(rows[3][headers.index("product_qty")], 3)

    def test_mrp_import_preview_maps_readonly_name_as_manufacturing_number(self):
        import_job = self.env["base_import.import"].create({
            "res_model": "mrp.production",
        })

        fields_tree = import_job.get_fields_tree("mrp.production")
        name_fields = [field for field in fields_tree if field.get("name") == "name"]

        self.assertEqual(len(name_fields), 1)
        self.assertEqual(name_fields[0]["string"], "制造单号")
        self.assertTrue(name_fields[0]["required"])
        for header in ("name", "制造单号"):
            suggestion = import_job._get_mapping_suggestion(
                header, fields_tree, ["char"], {},
            )
            self.assertEqual(suggestion["field_path"], ["name"])

    def test_product_page_family_title_and_out_of_stock_overlay_are_presentation_only(self):
        family_title_arch = self.env.ref(
            "stock_subwarehouse_hierarchy.product_page_family_title"
        ).arch_db
        unavailable_button_arch = self.env.ref(
            "stock_subwarehouse_hierarchy.product_page_replace_add_to_cart_button"
        ).arch_db

        self.assertIn("双板商品(SKI PRODUCTS)", family_title_arch)
        self.assertIn("font-size: 16px", family_title_arch)
        self.assertIn("x_shop_product_family_title", family_title_arch)
        self.assertIn('disabled="disabled"', unavailable_button_arch)
        self.assertIn("x_shop_stock_banner", unavailable_button_arch)
        self.assertIn("Out of stock", unavailable_button_arch)
        self.assertIn("缺货", unavailable_button_arch)
        self.assertNotIn("fa-times", unavailable_button_arch)

    def test_item_search_and_product_cards_reuse_title_hover_motion(self):
        item_page_arch = self.env.ref(
            "stock_subwarehouse_hierarchy.item_pages_product_card_section"
        ).arch_db
        family_page_arch = self.env.ref(
            "stock_subwarehouse_hierarchy.custom_product_family_page"
        ).arch_db

        for arch, search_class, card_class in (
            (item_page_arch, "x_item_product_search", "x_item_product_card"),
            (family_page_arch, "x_custom_family_search", "x_custom_product_card"),
        ):
            self.assertIn(f".{search_class}:focus-within", arch)
            self.assertIn(f".{card_class}:focus-visible", arch)
            self.assertIn("transition: transform 180ms ease", arch)
            self.assertIn("transition: transform 220ms ease", arch)
            self.assertIn("transform: translateY(-1px)", arch)
            self.assertIn("transform: scaleX(1)", arch)
            self.assertIn("prefers-reduced-motion: reduce", arch)

    def test_quant_import_rejects_duplicate_identity_with_partial_success(self):
        self.product_a.default_code = "IMPORT-QUANT-A"
        self.product_b.default_code = "IMPORT-QUANT-B"
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 1.0)
        fields_list = [
            "product_id/.id",
            "location_id/.id",
            "inventory_quantity_auto_apply",
        ]
        rows = [
            [str(self.product_a.id), str(self.bin_a.id), 2],
            [str(self.product_b.id), str(self.bin_b.id), 3],
            [str(self.product_b.id), str(self.bin_b.id), 4],
            ["", str(self.bin_b.id), 5],
        ]

        result = self.StockQuant.with_context(
            import_file=True,
            inventory_mode=True,
            business_import_source_offset=0,
            business_import_has_headers=True,
        ).load(fields_list, rows)

        imported = self.StockQuant.browse(result["ids"])
        self.assertEqual(imported.product_id, self.product_b)
        self.assertEqual(imported.location_id, self.bin_b)
        self.assertEqual(len(result["x_business_import_failures"]), 3)

    def test_store_map_url_opens_amap_search(self):
        map_url = _amap_search_url("SUN Beijing Flagship Store Test Address")

        self.assertTrue(map_url.startswith("https://uri.amap.com/search?"))
        self.assertIn("keyword=SUN+Beijing+Flagship+Store+Test+Address", map_url)
        self.assertIn("view=map", map_url)
        self.assertIn("callnative=1", map_url)

    def test_webclient_bootstrap_uses_dashboard_home_instead_of_discuss(self):
        user = self.env.user.sudo()
        dashboard = self.env.ref("spreadsheet_dashboard.ir_actions_dashboard_action")
        discuss = self.env.ref("mail.action_discuss")
        user.action_id = discuss.id

        user._on_webclient_bootstrap()

        self.assertEqual(user.action_id.id, dashboard.id)

    def test_website_pages_and_menus_are_made_public(self):
        view = self.env["ir.ui.view"].create({
            "name": "Private website page test",
            "type": "qweb",
            "key": "stock_subwarehouse_hierarchy.private_website_page_test",
            "arch": "<t t-call='website.layout'><main>Public access test</main></t>",
            "visibility": "connected",
        })
        page = self.env["website.page"].create({
            "name": "Private website page test",
            "url": "/private-website-page-test",
            "view_id": view.id,
            "is_published": False,
        })
        menu = self.env["website.menu"].create({
            "name": "Private website page test",
            "url": page.url,
            "page_id": page.id,
            "group_ids": [Command.set([self.env.ref("base.group_user").id])],
        })

        self.env["website.page"].action_make_all_pages_public()

        self.assertTrue(page.website_published)
        self.assertFalse(page.visibility)
        self.assertFalse(page.group_ids)
        self.assertFalse(menu.group_ids)

    def test_apply_sun_logo_updates_company_logo(self):
        expected_logo = self.env["res.company"]._get_sun_logo_binary()

        self.env["res.company"].action_apply_sun_logo()

        self.assertEqual(self.env.company.logo, expected_logo)

    def test_chinese_language_is_default_for_users_and_partners(self):
        self.env["res.lang"].action_use_chinese_by_default()

        zh_cn = self.env["res.lang"].with_context(active_test=False).search([("code", "=", "zh_CN")])
        en_us = self.env["res.lang"].with_context(active_test=False).search([("code", "=", "en_US")])
        self.assertTrue(zh_cn.active)
        self.assertEqual(en_us.name, "English")
        self.assertEqual(self.env["ir.default"]._get("res.partner", "lang"), "zh_CN")
        self.assertEqual(self.env.user.lang, "zh_CN")
        self.assertEqual(self.env.user.partner_id.lang, "zh_CN")

    def test_descendant_inventory_action_filters_internal_descendants(self):
        action = self.warehouse.action_view_descendant_inventory_totals()

        self.assertEqual(action["res_model"], "stock.quant")
        self.assertEqual(action["view_mode"], "list,pivot,graph,form")
        self.assertIn(("location_id", "child_of", self.warehouse.view_location_id.id), action["domain"])
        self.assertIn(("location_id", "!=", self.warehouse.view_location_id.id), action["domain"])
        self.assertIn(("location_id.usage", "=", "internal"), action["domain"])
        self.assertEqual(action["context"]["search_default_productgroup"], 1)
        self.assertEqual(action["context"]["search_default_locationgroup"], 1)
        self.assertEqual(action["context"]["descendant_inventory_warehouse_id"], self.warehouse.id)
        self.assertEqual(action["context"]["descendant_inventory_root_location_id"], self.warehouse.view_location_id.id)

    def test_descendant_product_totals_action_lists_products(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        self.StockQuant._update_available_quantity(self.product_a, self.bin_b, 7.0)
        self.StockQuant._update_available_quantity(self.product_b, self.bin_b, 3.0)

        action = self.warehouse.action_view_descendant_inventory_product_totals()
        totals = self.env["stock.subwarehouse.inventory.total"].search(action["domain"])
        totals_by_product = {
            total.product_id: total
            for total in totals
        }

        self.assertEqual(action["res_model"], "stock.subwarehouse.inventory.total")
        self.assertEqual(action["views"][0][1], "list")
        self.assertEqual(totals_by_product[self.product_a].quantity, 12.0)
        self.assertEqual(totals_by_product[self.product_a].available_quantity, 12.0)
        self.assertEqual(totals_by_product[self.product_b].quantity, 3.0)

    def test_descendant_product_totals_transfer_selected_opens_destination_wizard(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        self.StockQuant._update_available_quantity(self.product_a, self.bin_b, 7.0)

        action = self.warehouse.action_view_descendant_inventory_product_totals()
        total = self.env["stock.subwarehouse.inventory.total"].search([
            *action["domain"],
            ("product_id", "=", self.product_a.id),
        ])
        wizard_action = total.action_transfer_selected_out_of_current_warehouse()
        wizard = self.env[wizard_action["res_model"]].browse(wizard_action["res_id"])
        destination_warehouse = self.env["stock.warehouse"].create({
            "name": "Totals Transfer Destination",
            "code": "TTD",
        })
        wizard.destination_warehouse_id = destination_warehouse
        transfer_action = wizard.action_create_internal_transfer()
        picking = self.env["stock.picking"].browse(transfer_action["res_id"])

        self.assertEqual(wizard_action["res_model"], "stock.subwarehouse.internal.transfer.wizard")
        self.assertEqual(wizard.selected_inventory_count, 2)
        self.assertEqual(wizard.selected_product_count, 1)
        self.assertEqual(transfer_action["res_model"], "stock.picking")
        self.assertEqual(picking.picking_type_id, self.warehouse.int_type_id)
        self.assertEqual(picking.location_dest_id, destination_warehouse.lot_stock_id)
        self.assertEqual(picking.origin, "下级库存内部调拨")
        self.assertEqual(sum(picking.move_ids.mapped("product_uom_qty")), 12.0)
        self.assertEqual(set(picking.move_ids.mapped("location_id").ids), set((self.bin_a | self.bin_b).ids))
        self.assertTrue(all(move.product_id == self.product_a for move in picking.move_ids))

    def test_quant_descendant_transfer_button_uses_selected_destination_warehouse(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        self.StockQuant._update_available_quantity(self.product_b, self.bin_b, 3.0)
        quants = self.StockQuant.search([
            ("product_id", "in", (self.product_a | self.product_b).ids),
            ("location_id", "in", (self.bin_a | self.bin_b).ids),
        ])

        wizard_action = quants.with_context(
            descendant_inventory_warehouse_id=self.warehouse.id,
            descendant_inventory_root_location_id=self.warehouse.view_location_id.id,
        ).action_transfer_selected_out_of_descendant_inventory()
        wizard = self.env[wizard_action["res_model"]].browse(wizard_action["res_id"])
        destination_warehouse = self.env["stock.warehouse"].create({
            "name": "Quant Transfer Destination",
            "code": "QTD",
        })
        wizard.destination_warehouse_id = destination_warehouse
        transfer_action = wizard.action_create_internal_transfer()
        picking = self.env["stock.picking"].browse(transfer_action["res_id"])

        self.assertEqual(wizard_action["res_model"], "stock.subwarehouse.internal.transfer.wizard")
        self.assertEqual(set(wizard.quant_ids.ids), set(quants.ids))
        self.assertEqual(picking.picking_type_id, self.warehouse.int_type_id)
        self.assertEqual(picking.location_dest_id, destination_warehouse.lot_stock_id)
        self.assertEqual(picking.state, "draft")
        self.assertEqual(
            {move.product_id.id: move.product_uom_qty for move in picking.move_ids},
            {self.product_a.id: 5.0, self.product_b.id: 3.0},
        )

    def test_location_descendant_inventory_action_filters_location_subtree(self):
        action = self.subwarehouse.action_view_descendant_inventory_totals()

        self.assertEqual(action["res_model"], "stock.quant")
        self.assertEqual(action["view_mode"], "list,pivot,graph,form")
        self.assertIn(("location_id", "child_of", self.subwarehouse.id), action["domain"])
        self.assertIn(("location_id.usage", "=", "internal"), action["domain"])
        self.assertEqual(action["context"]["search_default_productgroup"], 1)
        self.assertEqual(action["context"]["search_default_locationgroup"], 1)

    def test_location_inventory_in_out_action_filters_source_or_destination(self):
        action = self.subwarehouse.action_view_inventory_in_out()

        self.assertEqual(action["res_model"], "stock.move")
        self.assertEqual(action["name"], "库存出入历史")
        self.assertEqual(action["view_mode"], "list,form")
        self.assertIn(("state", "=", "done"), action["domain"])
        self.assertIn(("location_id", "child_of", self.subwarehouse.id), action["domain"])
        self.assertIn(("location_dest_id", "child_of", self.subwarehouse.id), action["domain"])
        self.assertEqual(action["context"]["search_default_by_product"], 1)
        self.assertEqual(action["context"]["search_default_groupby_location_id"], 1)
        self.assertEqual(action["context"]["search_default_groupby_dest_location_id"], 1)
        self.assertFalse(action["context"]["create"])

    def test_location_internal_transfer_action_prefills_current_location(self):
        action = self.subwarehouse.action_create_internal_transfer()

        self.assertEqual(action["res_model"], "stock.picking")
        self.assertEqual(action["view_mode"], "form")
        self.assertEqual(action["context"]["restricted_picking_type_code"], "internal")
        self.assertEqual(action["context"]["default_location_id"], self.subwarehouse.id)
        self.assertEqual(action["context"]["default_location_dest_id"], self.subwarehouse.id)
        self.assertEqual(
            action["context"]["default_picking_type_id"],
            self.warehouse.int_type_id.id,
        )

    def test_location_load_remove_inventory_action_filters_current_location(self):
        action = self.subwarehouse.action_load_remove_inventory()

        self.assertEqual(action["res_model"], "stock.quant")
        self.assertEqual(action["view_mode"], "list")
        self.assertEqual(action["name"], "装入/移除产品")
        self.assertIn(("location_id", "child_of", self.subwarehouse.id), action["domain"])
        self.assertIn(("location_id.usage", "in", ["internal", "transit"]), action["domain"])
        self.assertTrue(action["context"]["inventory_mode"])
        self.assertEqual(action["context"]["default_location_id"], self.subwarehouse.id)

    def test_location_manufacture_product_action_prefills_internal_location(self):
        action = self.bin_a.action_manufacture_product()

        self.assertEqual(action["res_model"], "mrp.production")
        self.assertEqual(action["view_mode"], "form")
        self.assertEqual(action["context"]["subwarehouse_manufacturing_location_id"], self.bin_a.id)
        self.assertEqual(action["context"]["default_location_src_id"], self.bin_a.id)
        self.assertEqual(action["context"]["default_location_dest_id"], self.bin_a.id)
        self.assertEqual(
            action["context"]["default_picking_type_id"],
            self.warehouse.manu_type_id.id,
        )

    def test_view_location_manufacture_product_action_uses_internal_child(self):
        action = self.subwarehouse.action_manufacture_product()

        self.assertEqual(action["res_model"], "mrp.production")
        self.assertIn(
            action["context"]["default_location_dest_id"],
            (self.bin_a | self.bin_b).ids,
        )

    def test_warehouse_create_subwarehouse_creates_internal_location(self):
        action = self.warehouse.action_create_subwarehouse()
        location = self.StockLocation.browse(action["res_id"])

        self.assertEqual(location.usage, "internal")
        self.assertEqual(location.location_id, self.warehouse.view_location_id)
        self.assertTrue(location.x_is_subwarehouse)

    def _create_sale_order_line(self, product, quantity, source_location=False):
        order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "warehouse_id": self.warehouse.id,
        })
        line_values = {
            "order_id": order.id,
            "product_id": product.id,
            "product_uom_qty": quantity,
            "product_uom_id": product.uom_id.id,
        }
        if source_location:
            line_values["x_source_location_id"] = source_location.id
        line = self.env["sale.order.line"].create(line_values)
        return order, line

    def test_website_payment_auto_assigns_exact_stocked_subwarehouse(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 2.0)
        self.StockQuant._update_available_quantity(self.product_a, self.bin_b, 5.0)
        order, line = self._create_sale_order_line(self.product_a, 4.0)
        line.x_source_location_id = False

        order._prepare_website_stock_for_payment()

        self.assertEqual(line.x_source_location_id, self.bin_b)
        self.assertTrue(line.x_website_stock_reserved_until)
        self.assertTrue(order.x_website_stock_reserved_at)
        self.assertEqual(order.x_website_stock_reserved_until, line.x_website_stock_reserved_until)
        self.assertGreater(order.x_website_stock_reservation_expiry_epoch, 0)

    def test_website_payment_hold_blocks_second_quotation(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        first_order, first_line = self._create_sale_order_line(self.product_a, 4.0)
        first_line.x_source_location_id = False
        first_order._prepare_website_stock_for_payment()

        second_order, second_line = self._create_sale_order_line(self.product_a, 2.0)
        second_line.x_source_location_id = False

        with self.assertRaises(UserError):
            second_order._prepare_website_stock_for_payment()

    def test_cart_shortage_line_clears_after_quantity_is_reduced(self):
        shortage_product = self.env["product.product"].create({
            "name": "Cart Shortage Isolated Product",
            "is_storable": True,
        })
        self.StockQuant._update_available_quantity(shortage_product, self.bin_a, 2.0)
        order, line = self._create_sale_order_line(shortage_product, 3.0)
        line.x_source_location_id = False

        shortage_lines = order._get_source_inventory_shortage_lines()
        candidate_availability = [
            (
                location.complete_name,
                order._get_available_qty_for_source_location(
                    shortage_product, location, exclude_order=order,
                ),
            )
            for location in line._get_source_location_candidates()
        ]
        self.assertTrue(line.is_storable)
        self.assertFalse(line.display_type)
        self.assertLess(max(quantity for _location, quantity in candidate_availability), 3.0)
        self.assertIn(
            line,
            shortage_lines,
            f"quantity={line.product_uom_qty}, required={line._get_required_qty_in_product_uom()}, "
            f"candidates={candidate_availability}",
        )
        shortage_html = str(self.env["ir.ui.view"]._render_template(
            "website_sale.cart_lines",
            {
                "website_sale_order": order,
                "is_view_active": lambda _xml_id: True,
            },
        ))
        self.assertIn("x_stock_shortage_row", shortage_html)

        line.product_uom_qty = 2.0

        self.assertNotIn(line, order._get_source_inventory_shortage_lines())
        available_html = str(self.env["ir.ui.view"]._render_template(
            "website_sale.cart_lines",
            {
                "website_sale_order": order,
                "is_view_active": lambda _xml_id: True,
            },
        ))
        self.assertNotIn("x_stock_shortage_row", available_html)

    def _create_simulated_wechat_transaction(self, order, reference):
        provider = self.env.ref("payment_wechatpay.payment_provider_wechatpay")
        provider.write({
            "state": "test",
            "wechatpay_simulation_mode": True,
        })
        return self.env["payment.transaction"].create({
            "provider_id": provider.id,
            "payment_method_id": self.env.ref("payment_wechatpay.payment_method_wechatpay").id,
            "reference": reference,
            "amount": order.amount_total,
            "currency_id": order.currency_id.id,
            "partner_id": order.partner_id.id,
            "operation": "online_redirect",
            "sale_order_ids": [Command.set(order.ids)],
        })

    def test_simulated_wechat_payment_confirms_stocked_website_order(self):
        self.env["ir.config_parameter"].sudo().set_param("sale.automatic_invoice", "True")
        self.product_a.list_price = 120.0
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        order, line = self._create_sale_order_line(self.product_a, 2.0)
        order.pricelist_id = self.cny_pricelist
        order.website_id = self.env["website"].get_current_website()
        line.x_source_location_id = False
        order._prepare_website_stock_for_payment()
        tx = self._create_simulated_wechat_transaction(order, "WX-WEBSITE-SUCCESS")
        tx._wechatpay_ensure_native_order()

        tx._process("wechatpay", {
            "reference": tx.reference,
            "out_trade_no": tx.wechatpay_out_trade_no,
            "transaction_id": f"SIM-{tx.reference}",
            "trade_state": "SUCCESS",
        })
        with patch.object(
            self.env.registry["sale.order"],
            "_send_order_confirmation_mail",
            side_effect=UserError("Simulated PDF/email failure"),
        ):
            tx.with_context(skip_sale_auto_invoice_send=True)._post_process()

        self.assertEqual(tx.state, "done")
        self.assertEqual(order.state, "sale")
        self.assertTrue(tx.payment_id)
        self.assertEqual(tx.payment_id.state, "paid")
        self.assertTrue(tx.invoice_ids)
        self.assertTrue(tx.invoice_ids.filtered(lambda invoice: invoice.state == "posted"))
        self.assertEqual(order.x_website_payment_state, "paid")
        self.assertEqual(order.x_website_payment_reference, f"SIM-{tx.reference}")
        self.assertEqual(order._get_website_payment_receipt(), tx.payment_id)
        payment_action = order.action_view_website_payments()
        self.assertEqual(payment_action["res_id"], tx.payment_id.id)
        order._portal_ensure_token()
        receipt_html = str(self.env["ir.ui.view"]._render_template(
            "stock_subwarehouse_hierarchy.website_payment_receipt",
            {"order": order},
        ))
        self.assertIn("支付收据", receipt_html)
        self.assertIn(tx.payment_id.name, receipt_html)
        self.assertIn(f"/shop/payment/receipt/{order.id}", receipt_html)
        report_html, _report_type = self.env["ir.actions.report"]._render_qweb_html(
            "stock_subwarehouse_hierarchy.action_report_website_payment_receipt",
            tx.payment_id.ids,
        )
        report_html = report_html.decode()
        self.assertIn(order.name, report_html)
        self.assertIn(self.product_a.display_name, report_html)
        self.assertIn("不作为增值税发票", report_html)

        refund_wizard = self.env["stock.subwarehouse.website.payment.refund.wizard"].create({
            "order_id": order.id,
            "transaction_id": tx.id,
            "amount_to_refund": tx.amount,
        })
        refund_wizard.action_submit_refund()
        refund_tx = tx.child_transaction_ids.filtered(lambda child: child.operation == "refund")
        self.assertEqual(refund_tx.state, "done")
        self.assertTrue(refund_tx.wechatpay_out_refund_no)
        self.assertTrue(refund_tx.provider_reference.startswith("SIM-"))
        self.assertEqual(line.x_source_location_id, self.bin_a)
        delivery_move = line.move_ids.filtered(lambda move: move.product_id == self.product_a)
        self.assertTrue(delivery_move)
        self.assertEqual(delivery_move[:1].location_id, self.bin_a)

    def test_simulated_wechat_payment_does_not_charge_after_stock_disappears(self):
        self.product_a.list_price = 120.0
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 2.0)
        order, line = self._create_sale_order_line(self.product_a, 2.0)
        order.pricelist_id = self.cny_pricelist
        line.x_source_location_id = False
        order._prepare_website_stock_for_payment()
        tx = self._create_simulated_wechat_transaction(order, "WX-WEBSITE-NO-STOCK")
        tx._wechatpay_ensure_native_order()
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, -2.0)

        tx._process("wechatpay", {
            "reference": tx.reference,
            "out_trade_no": tx.wechatpay_out_trade_no,
            "transaction_id": f"SIM-{tx.reference}",
            "trade_state": "SUCCESS",
        })

        self.assertEqual(tx.state, "error")
        self.assertEqual(order.state, "draft")
        self.assertIn("库存复核失败", tx.state_message)

    def test_shop_groups_published_same_name_products_by_representative(self):
        product_1, product_2, product_3 = self.env["product.template"].create([
            {"name": "Shop Group Test", "sale_ok": True},
            {"name": "Shop Group Test", "sale_ok": True},
            {"name": "Different Shop Group", "sale_ok": True},
        ])
        (product_1 | product_2 | product_3).action_publish_to_shop()

        grouped_products = (product_1 | product_2 | product_3)._get_shop_grouped_products()

        self.assertEqual(grouped_products, product_1 | product_3)
        self.assertEqual(product_1.x_shop_group_variant_count, 2)

    def test_shop_group_cover_selects_named_product_and_can_restore_default(self):
        product_1, product_2 = self.env["product.template"].create([
            {"name": "Shop Cover Test", "sale_ok": True},
            {"name": "Shop Cover Test", "sale_ok": True},
        ])
        (product_1 | product_2).action_publish_to_shop()
        product_2.image_1920 = (
            b"iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8A"
            b"AQUBAScY42YAAAAASUVORK5CYII="
        )

        product_2.action_set_shop_group_cover()

        self.assertFalse(product_1.x_shop_group_cover)
        self.assertTrue(product_2.x_shop_group_cover)
        self.assertEqual(
            (product_1 | product_2)._get_shop_grouped_products(),
            product_2,
        )

        product_2.action_clear_shop_group_cover()

        self.assertFalse(product_2.x_shop_group_cover)
        self.assertEqual(
            (product_1 | product_2)._get_shop_grouped_products(),
            product_1,
        )

    def test_shop_group_cover_without_image_falls_back_to_default(self):
        product_1, product_2 = self.env["product.template"].create([
            {"name": "Missing Cover Image Test", "sale_ok": True},
            {
                "name": "Missing Cover Image Test",
                "sale_ok": True,
                "x_shop_group_cover": True,
            },
        ])

        self.assertEqual(
            (product_1 | product_2)._get_shop_grouped_products(),
            product_1,
        )

    def test_product_bulk_update_applies_shared_images_and_general_fields(self):
        products = self.env["product.template"].create([
            {"name": "Bulk Product A", "sale_ok": True},
            {"name": "Bulk Product B", "sale_ok": True},
        ])
        category = self.env["product.category"].create({"name": "Bulk Category"})
        image_data = (
            b"iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8A"
            b"AQUBAScY42YAAAAASUVORK5CYII="
        )
        wizard = self.env["stock.subwarehouse.product.bulk.update.wizard"].create({
            "selected_product_ids": [Command.set(products.ids)],
            "apply_main_image": True,
            "main_image": image_data,
            "gallery_mode": "append",
            "gallery_image_ids": [Command.create({
                "name": "Shared gallery image",
                "image_1920": image_data,
            })],
            "apply_description_zh": True,
            "description_zh": "<p>共享中文产品描述</p>",
            "apply_description_en": True,
            "description_en": "<p>Shared English product description</p>",
            "apply_english_name": True,
            "website_english_name": "Shared English Name",
            "apply_category": True,
            "categ_id": category.id,
            "apply_material_type": True,
            "material_type": "finished",
            "apply_list_price": True,
            "list_price": 799.0,
        })

        wizard.action_apply()

        for product in products:
            self.assertTrue(product.image_1920)
            self.assertIn("共享中文产品描述", product.x_website_description_zh)
            self.assertIn("Shared English product description", product.x_website_description_en)
            self.assertEqual(product.x_website_english_name, "Shared English Name")
            self.assertEqual(product.categ_id, category)
            self.assertEqual(product.x_material_type, "finished")
            self.assertEqual(product.list_price, 799.0)
            self.assertEqual(len(product.product_template_image_ids), 1)

        wizard.action_apply()
        self.assertEqual(
            products.mapped("product_template_image_ids").mapped("product_tmpl_id"),
            products,
        )

    def test_shop_product_family_filter_splits_ski_snowboard_and_other(self):
        ski, bracketed_ski, snowboard, other = self.env["product.template"].create([
            {"name": "零售滑雪双板", "default_code": "062410X-MA006-W001170", "sale_ok": True},
            {"name": "滑雪杖", "default_code": "[012410Z-MA000-HR01130]", "sale_ok": True},
            {"name": "儿童单板刻滑滑雪板", "default_code": "052411Dc-MK787-HX02135", "sale_ok": True},
            {"name": "滑雪手套", "default_code": "112411T1-MA000-P001##L", "sale_ok": True},
        ])

        products = ski | bracketed_ski | snowboard | other

        self.assertEqual(products._filter_shop_products_by_family("ski"), ski | bracketed_ski)
        self.assertEqual(products._filter_shop_products_by_family("snowboard"), snowboard)
        self.assertEqual(products._filter_shop_products_by_family("other"), other)

    def test_managed_custom_attributes_are_hidden_from_website_lines(self):
        custom_attribute = self.env["product.attribute"].create({
            "name": "测试自定义属性",
            "x_apply_to_all_products": True,
        })
        visible_attribute = self.env["product.attribute"].create({
            "name": "公开规格",
        })
        visible_value = self.env["product.attribute.value"].create({
            "name": "公开值",
            "attribute_id": visible_attribute.id,
        })
        product = self.env["product.template"].create({
            "name": "Website Attribute Test",
            "sale_ok": True,
            "attribute_line_ids": [
                Command.create({
                    "attribute_id": visible_attribute.id,
                    "value_ids": [Command.set(visible_value.ids)],
                }),
            ],
        })

        visible_lines = product._get_visible_website_attribute_lines()
        single_values = product.valid_product_template_attribute_line_ids._prepare_single_value_for_display()

        self.assertNotIn(custom_attribute, visible_lines.attribute_id)
        self.assertIn(visible_attribute, visible_lines.attribute_id)
        self.assertNotIn(custom_attribute, single_values)
        self.assertIn(visible_attribute, single_values)

    def test_shop_group_variant_values_use_custom_attributes(self):
        color_attribute = self.env["product.attribute"].create({
            "name": "颜色",
            "x_apply_to_all_products": True,
        })
        size_attribute = self.env["product.attribute"].create({
            "name": "尺码",
            "x_apply_to_all_products": True,
        })
        product = self.env["product.template"].create({
            "name": "Variant Value Test",
            "default_code": "SKU-001",
            "sale_ok": True,
        })
        product.x_custom_attribute_value_ids.filtered(
            lambda value: value.attribute_id == color_attribute
        ).value_text = "黑色"
        product.x_custom_attribute_value_ids.filtered(
            lambda value: value.attribute_id == size_attribute
        ).value_text = "260"

        variant_values = product._get_shop_variant_display_values()

        self.assertEqual(variant_values["default_code"], "SKU-001")
        self.assertEqual(variant_values["color"], "黑色")
        self.assertEqual(variant_values["size"], "260")

    def test_shop_variant_values_decode_missing_values_from_product_id(self):
        product = self.env["product.template"].create({
            "name": "Decoded Variant Test",
            "default_code": "152410Yb-MK000-H001150",
            "sale_ok": True,
        })

        variant_values = product._get_shop_variant_display_values()

        self.assertEqual(variant_values["color"], "黑")
        self.assertEqual(variant_values["size"], "150")
        self.assertEqual(variant_values["flex"], "无硬度")
        self.assertEqual(variant_values["audience"], "儿童/青少年")

    def test_shop_variant_values_decode_mixed_color_and_letter_size(self):
        product = self.env["product.template"].create({
            "name": "Decoded Letter Size Test",
            "default_code": "072409Y-MA000-G001##S",
            "sale_ok": True,
        })

        variant_values = product._get_shop_variant_display_values()

        self.assertEqual(variant_values["color"], "绿")
        self.assertEqual(variant_values["size"], "S")
        self.assertEqual(variant_values["flex"], "无硬度")
        self.assertEqual(variant_values["audience"], "成人")

    def test_shop_variant_values_decode_hardness_and_multiple_colors(self):
        product = self.env["product.template"].create({
            "name": "Decoded Flex Test",
            "default_code": "152410Y-MK787-HW02130",
            "sale_ok": True,
        })

        variant_values = product._get_shop_variant_display_values()

        self.assertEqual(variant_values["color"], "黑白")
        self.assertEqual(variant_values["size"], "130")
        self.assertEqual(variant_values["flex"], "787")
        self.assertEqual(variant_values["audience"], "儿童/青少年")

    def test_shop_group_variant_option_groups_use_same_name_products(self):
        product_1, product_2, product_3 = self.env["product.template"].create([
            {
                "name": "Shop Selector Test",
                "default_code": "152410Yb-MK000-H001150",
                "sale_ok": True,
            },
            {
                "name": "Shop Selector Test",
                "default_code": "152410Yb-MK000-W001155",
                "sale_ok": True,
            },
            {
                "name": "Shop Selector Test",
                "default_code": "152410Yb-MK100-H001150",
                "sale_ok": True,
            },
        ])
        (product_1 | product_2 | product_3).action_publish_to_shop()

        option_groups = product_1._get_shop_group_variant_option_groups()
        options_by_key = {
            group["key"]: group["values"]
            for group in option_groups
        }
        groups_by_key = {
            group["key"]: group
            for group in option_groups
        }

        self.assertEqual(options_by_key["type_color"], ["YB · 白", "YB · 黑"])
        self.assertEqual(options_by_key["size"], ["150", "155"])
        self.assertEqual(options_by_key["flex"], ["100", "无硬度"])
        self.assertEqual(
            groups_by_key["type_color"]["image_products"]["YB · 黑"],
            product_1,
        )
        self.assertEqual(
            groups_by_key["type_color"]["image_products"]["YB · 白"],
            product_2,
        )
        self.assertFalse(groups_by_key["size"]["image_products"])

    def test_shop_picture_options_distinguish_product_type_and_color_code(self):
        t1_glove, t5_glove = self.env["product.template"].create([
            {
                "name": "Type Color Selector Test Glove",
                "default_code": "112411T1-MA000-H001##S",
                "sale_ok": True,
            },
            {
                "name": "Type Color Selector Test Glove",
                "default_code": "112411T5-MA000-H001##S",
                "sale_ok": True,
            },
        ])
        (t1_glove | t5_glove).action_publish_to_shop()

        values = {
            group["key"]: group["values"]
            for group in t1_glove._get_shop_group_variant_option_groups()
        }

        self.assertEqual(values["type_color"], ["T1 · 黑", "T5 · 黑"])
        self.assertEqual(t1_glove._get_shop_variant_display_values()["color_code"], "H001")
        self.assertEqual(t5_glove._get_shop_variant_display_values()["type_code"], "T5")

    def test_shop_variant_sizes_are_sorted_numerically(self):
        products = self.env["product.template"].create([
            {
                "name": "Sorted Ski Boot",
                "default_code": "012307S2-MA100-H001275",
                "sale_ok": True,
            },
            {
                "name": "Sorted Ski Boot",
                "default_code": "012307S2-MA100-H001220",
                "sale_ok": True,
            },
            {
                "name": "Sorted Ski Boot",
                "default_code": "012307S2-MA100-H001245",
                "sale_ok": True,
            },
        ])
        products.action_publish_to_shop()

        groups = products[0]._get_shop_group_variant_option_groups()
        sizes = next(group["values"] for group in groups if group["key"] == "size")

        self.assertEqual(sizes, ["220", "245", "275"])

    def test_shop_variant_option_groups_hide_flex_when_all_products_have_no_flex(self):
        products = self.env["product.template"].create([
            {
                "name": "No Flex Glove Selector Test",
                "default_code": "072409Y-MA000-G001##S",
                "sale_ok": True,
            },
            {
                "name": "No Flex Glove Selector Test",
                "default_code": "072409Y-MA000-H001##M",
                "sale_ok": True,
            },
        ])
        products.action_publish_to_shop()

        groups = products[0]._get_shop_group_variant_option_groups()

        self.assertNotIn("flex", [group["key"] for group in groups])

    def test_shop_groups_use_stable_chinese_name_across_languages(self):
        products = self.env["product.template"].create([
            {
                "name": "Legacy Ski Boot",
                "default_code": "012307S2-MA100-H001245",
                "sale_ok": True,
            },
            {
                "name": "Edited English Ski Boot",
                "default_code": "012307S2-MA100-H001250",
                "sale_ok": True,
            },
        ])
        products.with_context(lang="zh_CN").write({"name": "分组语言稳定性测试雪鞋"})
        products.action_publish_to_shop()

        english_products = products.with_context(lang="en_US")

        self.assertEqual(len(english_products._get_shop_grouped_products()), 1)
        self.assertEqual(
            english_products[0]._get_shop_group_siblings(),
            english_products,
        )

    def test_shop_boot_product_detection(self):
        ski_boot = self.env["product.template"].create({"name": "双板滑雪鞋"})
        snowboard = self.env["product.template"].create({"name": "Snowboard Boot Pro"})
        ski = self.env["product.template"].create({"name": "双板滑雪板"})

        self.assertTrue(ski_boot._is_shop_boot_product())
        self.assertTrue(snowboard._is_shop_boot_product())
        self.assertFalse(ski._is_shop_boot_product())

    def test_shop_availability_uses_current_product_on_hand(self):
        product = self.env["product.template"].create({
            "name": "Shop Stock Test",
            "is_storable": True,
            "sale_ok": True,
        })

        self.assertFalse(product._is_shop_available())

        self.StockQuant._update_available_quantity(
            product.product_variant_id,
            self.warehouse.lot_stock_id,
            2.0,
        )

        self.assertTrue(product._is_shop_available())
        self.assertEqual(product._get_shop_available_quantity(), 2.0)

    def test_shop_publish_actions_toggle_website_visibility(self):
        product = self.env["product.template"].create({
            "name": "Publish Action Test",
            "sale_ok": False,
            "website_published": False,
        })

        product.action_publish_to_shop()

        self.assertTrue(product.sale_ok)
        self.assertTrue(product.website_published)

        product.action_unpublish_from_shop()

        self.assertFalse(product.website_published)

    def test_source_location_check_does_not_use_descendant_stock(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        order, _line = self._create_sale_order_line(self.product_a, 1.0, self.subwarehouse)

        with self.assertRaises(UserError):
            order._check_source_inventory_availability()

    def test_view_location_manufacture_product_action_creates_internal_child_when_missing(self):
        view_location = self.StockLocation.create({
            "name": "Empty Manufacturing Subwarehouse",
            "usage": "view",
            "location_id": self.warehouse.view_location_id.id,
        })

        action = view_location.action_manufacture_product()
        internal_child = self.StockLocation.browse(action["context"]["default_location_dest_id"])

        self.assertEqual(internal_child.usage, "internal")
        self.assertEqual(internal_child.location_id, view_location)
        self.assertTrue(internal_child.x_is_subwarehouse)

    def test_manufacturing_order_defaults_keep_subwarehouse_location(self):
        action = self.bin_a.action_manufacture_product()
        defaults = self.env["mrp.production"].with_context(
            action["context"],
        ).default_get(["location_src_id", "location_dest_id"])

        self.assertEqual(defaults["location_src_id"], self.bin_a.id)
        self.assertEqual(defaults["location_dest_id"], self.bin_a.id)
        self.assertNotEqual(defaults["location_dest_id"], self.warehouse.lot_stock_id.id)

    def test_manufacturing_order_create_keeps_subwarehouse_location(self):
        action = self.bin_a.action_manufacture_product()
        production = self.env["mrp.production"].with_context(action["context"]).create({
            "product_id": self.product_a.id,
            "product_qty": 1.0,
            "product_uom_id": self.product_a.uom_id.id,
            "picking_type_id": self.warehouse.manu_type_id.id,
        })

        self.assertEqual(production.location_src_id, self.bin_a)
        self.assertEqual(production.location_dest_id, self.bin_a)
        self.assertNotEqual(production.location_dest_id, self.warehouse.lot_stock_id)

    def test_manufacturing_completion_adds_stock_to_subwarehouse_inventory(self):
        action = self.bin_a.action_manufacture_product()
        production = self.env["mrp.production"].with_context(action["context"]).create({
            "product_id": self.product_a.id,
            "product_qty": 2.0,
            "product_uom_id": self.product_a.uom_id.id,
            "picking_type_id": self.warehouse.manu_type_id.id,
        })

        production.action_confirm()
        production.qty_producing = 2.0
        production.button_mark_done()

        quantity = self.StockQuant._get_available_quantity(self.product_a, self.bin_a)
        self.assertEqual(quantity, 2.0)

    def test_location_import_manufacturing_sheet_action_uses_builtin_import(self):
        action = self.subwarehouse.action_import_manufacturing_sheet()

        self.assertEqual(action["type"], "ir.actions.client")
        self.assertEqual(action["tag"], "import")
        self.assertEqual(action["target"], "current")
        self.assertEqual(action["params"]["model"], "mrp.production")
        self.assertEqual(action["params"]["active_model"], "mrp.production")
        self.assertIn(
            action["params"]["context"]["subwarehouse_manufacturing_location_id"],
            (self.bin_a | self.bin_b).ids,
        )
        self.assertEqual(
            action["params"]["context"]["default_location_src_id"],
            action["params"]["context"]["subwarehouse_manufacturing_location_id"],
        )
        self.assertEqual(
            action["params"]["context"]["default_location_dest_id"],
            action["params"]["context"]["subwarehouse_manufacturing_location_id"],
        )

    def test_location_manufacturing_history_action_filters_current_inventory(self):
        action = self.subwarehouse.action_view_manufacturing_history()

        self.assertEqual(action["res_model"], "mrp.production")
        self.assertIn(("location_src_id", "child_of", action["context"]["subwarehouse_manufacturing_location_id"]), action["domain"])
        self.assertIn(("location_dest_id", "child_of", action["context"]["subwarehouse_manufacturing_location_id"]), action["domain"])
        self.assertIn(
            action["context"]["subwarehouse_manufacturing_location_id"],
            (self.bin_a | self.bin_b).ids,
        )

    def test_sale_line_source_inventory_shows_available_quantity(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "warehouse_id": self.warehouse.id,
            "order_line": [(0, 0, {
                "product_id": self.product_a.id,
                "product_uom_qty": 3.0,
                "x_source_location_id": self.bin_a.id,
            })],
        })
        line = order.order_line

        self.assertIn(line.x_source_location_id._origin, line.x_eligible_source_location_ids)
        self.assertEqual(line.x_source_available_qty, 5.0)
        self.assertTrue(line.x_source_can_fulfill)

    def test_sale_line_source_inventory_parent_location_shows_child_quantity(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "warehouse_id": self.warehouse.id,
            "order_line": [(0, 0, {
                "product_id": self.product_a.id,
                "product_uom_qty": 3.0,
                "x_source_location_id": self.subwarehouse.id,
            })],
        })
        line = order.order_line

        self.assertEqual(line.x_source_location_id, self.subwarehouse)
        self.assertEqual(line.x_source_available_qty, 0.0)
        self.assertFalse(line.x_source_can_fulfill)
        with self.assertRaises(UserError):
            order.action_confirm()

    def test_sale_line_source_inventory_options_only_include_locations_that_can_fulfill(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        self.StockQuant._update_available_quantity(self.product_a, self.bin_b, 2.0)
        other_warehouse = self.env["stock.warehouse"].create({
            "name": "Other Source Warehouse",
            "code": "OSW",
        })
        self.StockQuant._update_available_quantity(self.product_a, other_warehouse.lot_stock_id, 4.0)
        order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "warehouse_id": self.warehouse.id,
            "order_line": [(0, 0, {
                "product_id": self.product_a.id,
                "product_uom_qty": 3.0,
            })],
        })
        line = order.order_line

        self.assertIn(self.bin_a, line.x_eligible_source_location_ids)
        self.assertNotIn(self.bin_b, line.x_eligible_source_location_ids)
        self.assertIn(other_warehouse.lot_stock_id, line.x_eligible_source_location_ids)
        self.assertIn(line.x_source_location_id._origin, line.x_eligible_source_location_ids)
        onchange_result = line._onchange_x_source_location_domain()
        self.assertEqual(
            onchange_result["domain"]["x_source_location_id"],
            [("id", "in", line.x_eligible_source_location_ids.ids)],
        )
        dropdown_options = self.StockLocation.with_context(
            sale_source_inventory_filter=True,
            sale_source_product_id=self.product_a.id,
            sale_source_product_uom_id=self.product_a.uom_id.id,
            sale_source_product_uom_qty=3.0,
            sale_source_warehouse_id=self.warehouse.id,
        ).name_search(limit=100)
        dropdown_location_ids = {location_id for location_id, _name in dropdown_options}
        self.assertIn(self.bin_a.id, dropdown_location_ids)
        self.assertNotIn(self.bin_b.id, dropdown_location_ids)
        self.assertIn(other_warehouse.lot_stock_id.id, dropdown_location_ids)

    def test_sale_source_inventory_name_search_can_return_more_than_default_dropdown_limit(self):
        stocked_locations = self.env["stock.location"]
        for index in range(9):
            location = self.StockLocation.create({
                "name": f"Overflow Source {index}",
                "usage": "internal",
                "location_id": self.warehouse.view_location_id.id,
            })
            stocked_locations |= location
            self.StockQuant._update_available_quantity(self.product_a, location, 5.0)

        dropdown_options = self.StockLocation.with_context(
            sale_source_inventory_filter=True,
            sale_source_product_id=self.product_a.id,
            sale_source_product_uom_id=self.product_a.uom_id.id,
            sale_source_product_uom_qty=3.0,
            sale_source_warehouse_id=self.warehouse.id,
        ).name_search(limit=10)
        dropdown_location_ids = {location_id for location_id, _name in dropdown_options}

        self.assertTrue(set(stocked_locations.ids).issubset(dropdown_location_ids))

    def test_sale_line_source_inventory_options_are_empty_for_zero_or_negative_quantity(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "warehouse_id": self.warehouse.id,
            "order_line": [(0, 0, {
                "product_id": self.product_a.id,
                "product_uom_qty": 0.0,
            })],
        })
        line = order.order_line

        self.assertFalse(line.x_eligible_source_location_ids)
        self.assertFalse(line.x_source_location_id)

        line.product_uom_qty = -1.0
        line._compute_x_eligible_source_location_ids()
        line._compute_x_source_location_id()
        self.assertFalse(line.x_eligible_source_location_ids)
        self.assertFalse(line.x_source_location_id)
        dropdown_options = self.StockLocation.with_context(
            sale_source_inventory_filter=True,
            sale_source_product_id=self.product_a.id,
            sale_source_product_uom_id=self.product_a.uom_id.id,
            sale_source_product_uom_qty=0.0,
            sale_source_warehouse_id=self.warehouse.id,
        ).name_search(limit=100)
        self.assertFalse(dropdown_options)

    def test_sale_confirmation_blocks_aggregate_source_shortage(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "warehouse_id": self.warehouse.id,
            "order_line": [
                (0, 0, {
                    "product_id": self.product_a.id,
                    "product_uom_qty": 3.0,
                    "x_source_location_id": self.bin_a.id,
                }),
                (0, 0, {
                    "product_id": self.product_a.id,
                    "product_uom_qty": 3.0,
                    "x_source_location_id": self.bin_a.id,
                }),
            ],
        })

        with self.assertRaises(UserError):
            order.action_confirm()

    def test_bulk_sale_confirmation_confirms_eligible_orders_with_partial_failure(self):
        service_product = self.env["product.product"].create({
            "name": "Bulk Confirmation Service",
            "type": "service",
        })
        good_order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "order_line": [Command.create({
                "product_id": service_product.id,
                "product_uom_qty": 1,
            })],
        })
        already_confirmed_order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "order_line": [Command.create({
                "product_id": service_product.id,
                "product_uom_qty": 1,
            })],
        })
        already_confirmed_order.action_confirm()
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 1.0)
        failing_order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "warehouse_id": self.warehouse.id,
            "order_line": [Command.create({
                "product_id": self.product_a.id,
                "product_uom_qty": 2,
                "x_source_location_id": self.bin_a.id,
            })],
        })

        result = (good_order | failing_order | already_confirmed_order).action_bulk_confirm()

        self.assertEqual(good_order.state, "sale")
        self.assertEqual(failing_order.state, "draft")
        self.assertEqual(already_confirmed_order.state, "sale")
        self.assertEqual(result["tag"], "display_notification")
        self.assertEqual(result["params"]["type"], "warning")
        self.assertIn("已成功确认 1 个销售订单", result["params"]["message"])
        self.assertIn("已忽略 1 个非待确认状态的订单", result["params"]["message"])
        self.assertIn("有 1 个订单确认失败", result["params"]["message"])

        server_action = self.env.ref(
            "stock_subwarehouse_hierarchy.action_sale_order_bulk_confirm"
        )
        self.assertEqual(server_action.binding_model_id.model, "sale.order")
        self.assertEqual(server_action.binding_view_types, "list")

    def test_bulk_sale_archiving_archives_active_orders(self):
        active_order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
        })
        already_archived_order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "active": False,
        })

        result = (active_order | already_archived_order).action_bulk_archive()

        self.assertFalse(active_order.active)
        self.assertFalse(already_archived_order.active)
        self.assertEqual(result["tag"], "display_notification")
        self.assertEqual(result["params"]["type"], "success")
        self.assertIn("已成功归档 1 个销售订单", result["params"]["message"])
        self.assertIn("已忽略 1 个已归档订单", result["params"]["message"])

        server_action = self.env.ref(
            "stock_subwarehouse_hierarchy.action_sale_order_bulk_archive"
        )
        self.assertEqual(server_action.binding_model_id.model, "sale.order")
        self.assertEqual(server_action.binding_view_types, "list")

    def test_external_sale_is_completed_without_inventory_or_delivery_documents(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        self.product_a.lst_price = 325.0

        order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "warehouse_id": self.warehouse.id,
            "x_is_external_order": True,
            "x_processing_fee": 25.0,
            "x_amount_received": 600.0,
            "order_line": [Command.create({
                "product_id": self.product_a.id,
                "product_uom_qty": 2.0,
                "price_unit": 300.0,
                "x_source_location_id": self.bin_a.id,
            })],
        })

        self.assertEqual(order.state, "external_done")
        self.assertEqual(order.delivery_status, "full")
        self.assertEqual(order.order_line.qty_delivered, 2.0)
        self.assertEqual(order.order_line.qty_to_invoice, 0.0)
        self.assertEqual(order.order_line.invoice_status, "no")
        self.assertFalse(order.order_line.x_source_location_id)
        self.assertFalse(order.picking_ids)
        self.assertEqual(order.x_official_total, 650.0)
        self.product_a.lst_price = 999.0
        self.assertEqual(order.x_official_total, 650.0)
        self.assertEqual(
            self.StockQuant._get_available_quantity(self.product_a, self.bin_a),
            5.0,
        )
        with self.assertRaisesRegex(UserError, "外部订单不参与ERP开票流程"):
            order._create_invoices()

    def test_external_sale_report_is_separate_and_allocates_order_finance_once(self):
        self.product_a.lst_price = 100.0
        external_order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "x_is_external_order": True,
            "x_processing_fee": 15.0,
            "x_amount_received": 250.0,
            "order_line": [
                Command.create({
                    "product_id": self.product_a.id,
                    "product_uom_qty": 1.0,
                    "price_unit": 100.0,
                }),
                Command.create({
                    "product_id": self.product_b.id,
                    "product_uom_qty": 2.0,
                    "price_unit": 100.0,
                }),
            ],
        })
        service_product = self.env["product.product"].create({
            "name": "External Report Internal Control",
            "type": "service",
            "list_price": 100.0,
        })
        internal_order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "order_line": [Command.create({
                "product_id": service_product.id,
                "product_uom_qty": 1.0,
                "price_unit": 100.0,
            })],
        })
        internal_order.with_context(skip_procurement=True).action_confirm()

        external_report = self.env["sale.report"].search([
            ("name", "=", external_order.name),
            ("x_is_external_order", "=", True),
        ])
        internal_report = self.env["sale.report"].search([
            ("name", "=", internal_order.name),
            ("x_is_external_order", "=", False),
        ])

        self.assertEqual(len(external_report), 2)
        self.assertAlmostEqual(sum(external_report.mapped("x_processing_fee")), 15.0)
        self.assertAlmostEqual(sum(external_report.mapped("x_amount_received")), 250.0)
        self.assertTrue(internal_report)

        internal_action = self.env.ref("sale.action_order_report_all")
        external_action = self.env.ref(
            "stock_subwarehouse_hierarchy.action_external_sale_report"
        )
        self.assertIn("('x_is_external_order', '=', False)", internal_action.domain)
        self.assertIn("('x_is_external_order', '=', True)", external_action.domain)

    def test_internal_and_external_sales_dashboards_use_separate_domains(self):
        internal = self.env.ref(
            "spreadsheet_dashboard_sale.spreadsheet_dashboard_sales"
        )
        external = self.env.ref(
            "stock_subwarehouse_hierarchy.spreadsheet_dashboard_external_sales"
        )
        self.env["spreadsheet.dashboard"]._configure_separate_sales_dashboards()

        internal_snapshot = json.loads(internal.spreadsheet_data)
        external_snapshot = json.loads(external.spreadsheet_data)

        internal_domains = [
            data["domain"]
            for data in internal_snapshot.get("lists", {}).values()
            if data.get("model") == "sale.order"
        ] + [
            data["domain"]
            for data in internal_snapshot.get("pivots", {}).values()
            if data.get("model") == "sale.report"
        ]
        external_domains = [
            data["domain"]
            for data in external_snapshot.get("lists", {}).values()
            if data.get("model") == "sale.order"
        ] + [
            data["domain"]
            for data in external_snapshot.get("pivots", {}).values()
            if data.get("model") == "sale.report"
        ]

        self.assertTrue(internal_domains)
        self.assertTrue(external_domains)
        self.assertTrue(all(
            '"x_is_external_order", "=", false' in json.dumps(domain)
            for domain in internal_domains
        ))
        self.assertTrue(all(
            '"x_is_external_order", "=", true' in json.dumps(domain)
            for domain in external_domains
        ))
        self.assertIn('"x_amount_received"', external.spreadsheet_data)

    def test_sale_delivery_move_uses_selected_source_inventory(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "warehouse_id": self.warehouse.id,
            "order_line": [(0, 0, {
                "product_id": self.product_a.id,
                "product_uom_qty": 3.0,
                "x_source_location_id": self.bin_a.id,
            })],
        })

        order.action_confirm()

        move = order.order_line.move_ids.filtered(lambda stock_move: stock_move.product_id == self.product_a)
        self.assertTrue(move)
        self.assertEqual(move[:1].location_id, self.bin_a)

    def test_internal_transfer_from_parent_location_reserves_child_stock(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        picking = self.env["stock.picking"].create({
            "picking_type_id": self.warehouse.int_type_id.id,
            "location_id": self.subwarehouse.id,
            "location_dest_id": self.warehouse.lot_stock_id.id,
            "move_ids": [(0, 0, {
                "description_picking": self.product_a.display_name,
                "product_id": self.product_a.id,
                "product_uom_qty": 3.0,
                "product_uom": self.product_a.uom_id.id,
                "location_id": self.subwarehouse.id,
                "location_dest_id": self.warehouse.lot_stock_id.id,
            })],
        })

        picking.action_confirm()
        picking.action_assign()

        move = picking.move_ids.filtered(lambda stock_move: stock_move.product_id == self.product_a)
        self.assertEqual(move.state, "assigned")
        self.assertTrue(move.move_line_ids.filtered(lambda line: line.location_id == self.bin_a))

    def test_internal_transfer_product_selector_uses_source_descendant_available_stock(self):
        self.StockQuant._update_available_quantity(self.product_a, self.bin_a, 5.0)
        self.StockQuant._update_available_quantity(self.product_b, self.bin_b, 2.0)
        self.StockQuant._update_reserved_quantity(
            self.product_b,
            self.bin_b,
            2.0,
            strict=True,
        )
        other_warehouse = self.env["stock.warehouse"].create({
            "name": "Internal Selector Other Warehouse",
            "code": "ISOW",
        })
        self.StockQuant._update_available_quantity(
            self.product_b,
            other_warehouse.lot_stock_id,
            8.0,
        )

        options = self.env["product.product"].with_context(
            internal_transfer_stock_filter=True,
            internal_transfer_source_location_id=self.subwarehouse.id,
            internal_transfer_company_id=self.warehouse.company_id.id,
        )
        autocomplete_ids = {
            product_id for product_id, _label in options.name_search(limit=100)
        }
        search_more_ids = set(options.search([]).ids)

        self.assertIn(self.product_a.id, autocomplete_ids)
        self.assertNotIn(self.product_b.id, autocomplete_ids)
        self.assertIn(self.product_a.id, search_more_ids)
        self.assertNotIn(self.product_b.id, search_more_ids)

    def test_product_attribute_apply_wizard_adds_attribute_to_all_products(self):
        templates = self.env["product.template"].search([])
        wizard = self.env["stock.subwarehouse.product.attribute.apply.wizard"].create({
            "attribute_name": "Global Test Attribute",
            "value_name": "Default Test Value",
        })

        action = wizard.action_apply()

        attribute = self.env["product.attribute"].search([
            ("name", "=", "Global Test Attribute"),
        ], limit=1)
        value = self.env["product.attribute.value"].search([
            ("attribute_id", "=", attribute.id),
            ("name", "=", "Default Test Value"),
        ], limit=1)
        self.assertTrue(attribute)
        self.assertEqual(attribute.create_variant, "no_variant")
        self.assertTrue(attribute.x_apply_to_all_products)
        self.assertEqual(attribute.x_default_custom_value, "Default Test Value")
        self.assertTrue(value)
        self.assertTrue(value.is_custom)
        for template in templates:
            line = template.attribute_line_ids.filtered(
                lambda attribute_line: attribute_line.attribute_id == attribute
            )
            custom_value = template.x_custom_attribute_value_ids.filtered(
                lambda record: record.attribute_id == attribute
            )
            self.assertTrue(line)
            self.assertIn(value, line.value_ids)
            self.assertTrue(custom_value)
            self.assertEqual(custom_value.value_text, "Default Test Value")
        self.assertEqual(action["res_model"], "product.template")

    def test_new_product_gets_global_custom_attributes_as_free_text(self):
        wizard = self.env["stock.subwarehouse.product.attribute.apply.wizard"].create({
            "attribute_name": "Future Product Attribute",
            "value_name": "Future Any Text 42.5 / 批次A",
        })
        wizard.action_apply()

        product = self.env["product.template"].create({
            "name": "Product Created After Global Attribute",
        })
        attribute = self.env["product.attribute"].search([
            ("name", "=", "Future Product Attribute"),
        ], limit=1)
        value = self.env["product.attribute.value"].search([
            ("attribute_id", "=", attribute.id),
            ("name", "=", "Future Any Text 42.5 / 批次A"),
        ], limit=1)
        line = product.attribute_line_ids.filtered(
            lambda attribute_line: attribute_line.attribute_id == attribute
        )
        custom_value = product.x_custom_attribute_value_ids.filtered(
            lambda record: record.attribute_id == attribute
        )

        self.assertTrue(line)
        self.assertIn(value, line.value_ids)
        self.assertTrue(custom_value)
        self.assertEqual(custom_value.value_text, "Future Any Text 42.5 / 批次A")

    def test_product_form_has_visible_custom_attributes_page(self):
        view = self.env.ref(
            "stock_subwarehouse_hierarchy.product_template_form_custom_attributes_visible"
        )

        self.assertIn('name="custom_attributes"', view.arch_db)
        self.assertIn('name="x_custom_attribute_value_ids"', view.arch_db)
        self.assertIn('name="value_text"', view.arch_db)

    def test_product_views_show_material_type(self):
        form_view = self.env.ref(
            "stock_subwarehouse_hierarchy.product_template_form_material_type_visible"
        )
        list_view = self.env.ref(
            "stock_subwarehouse_hierarchy.product_template_list_material_type_visible"
        )
        search_view = self.env.ref(
            "stock_subwarehouse_hierarchy.product_template_search_material_type_filters"
        )

        self.assertIn('name="x_material_type"', form_view.arch_db)
        self.assertIn('name="x_material_type"', list_view.arch_db)
        self.assertIn("material_component", search_view.arch_db)
        self.assertIn("group_by_material_type", search_view.arch_db)

    def test_inventory_material_type_actions_filter_finished_and_components(self):
        self.product_a.product_tmpl_id.x_material_type = "finished"
        self.product_b.product_tmpl_id.x_material_type = "component"

        finished_action = self.env.ref(
            "stock_subwarehouse_hierarchy.action_finished_product_inventory"
        )
        component_action = self.env.ref(
            "stock_subwarehouse_hierarchy.action_component_inventory"
        )
        search_view = self.env.ref(
            "stock_subwarehouse_hierarchy.stock_quant_search_material_type_filters"
        )

        self.assertIn("x_material_type", finished_action.domain)
        self.assertIn("finished", finished_action.domain)
        self.assertIn("x_material_type", component_action.domain)
        self.assertIn("component", component_action.domain)
        self.assertIn("material_finished_inventory", search_view.arch_db)
        self.assertIn("material_component_inventory", search_view.arch_db)

    def test_remove_global_custom_attribute_only_removes_managed_attribute(self):
        wizard = self.env["stock.subwarehouse.product.attribute.apply.wizard"].create({
            "attribute_name": "Removable Custom Attribute",
            "value_name": "Remove Me",
        })
        wizard.action_apply()
        attribute = self.env["product.attribute"].search([
            ("name", "=", "Removable Custom Attribute"),
        ], limit=1)
        normal_attribute = self.env["product.attribute"].create({
            "name": "Normal Variant Attribute",
            "create_variant": "no_variant",
        })
        normal_value = self.env["product.attribute.value"].create({
            "name": "Normal Value",
            "attribute_id": normal_attribute.id,
        })
        self.product_a.product_tmpl_id.write({
            "attribute_line_ids": [(0, 0, {
                "attribute_id": normal_attribute.id,
                "value_ids": [(6, 0, normal_value.ids)],
            })],
        })

        remove_wizard = self.env["stock.subwarehouse.product.attribute.remove.wizard"].create({
            "attribute_id": attribute.id,
        })
        remove_wizard.action_remove()

        self.assertFalse(attribute.x_apply_to_all_products)
        self.assertFalse(self.env["product.template.custom.attribute.value"].search([
            ("attribute_id", "=", attribute.id),
        ]))
        self.assertTrue(self.product_a.product_tmpl_id.attribute_line_ids.filtered(
            lambda line: line.attribute_id == normal_attribute
        ))

    def test_product_import_template_is_available_on_import_page(self):
        templates = self.env["product.template"].get_import_templates()

        self.assertEqual(
            [template["template"] for template in templates],
            ["/stock_subwarehouse_hierarchy/import_template/product_template.xlsx"],
        )

    def test_product_import_template_auto_matches_global_custom_attributes(self):
        from openpyxl import load_workbook

        wizard = self.env["stock.subwarehouse.product.attribute.apply.wizard"].create({
            "attribute_name": "Import Matched Attribute",
            "value_name": "Import Matched Value",
        })
        wizard.action_apply()

        content = self.env["product.template"]._generate_dynamic_product_import_template_xlsx()
        workbook = load_workbook(BytesIO(content), read_only=True)

        import_rows = list(workbook["产品导入"].iter_rows(values_only=True))
        import_headers = import_rows[0]
        chinese_labels = import_rows[1]
        self.assertIn("name", import_headers)
        self.assertIn("x_material_type", import_headers)
        self.assertIn("x_component_material", import_headers)
        self.assertIn("x_component_specification", import_headers)
        self.assertIn("x_component_color", import_headers)
        self.assertIn("is_storable", import_headers)
        self.assertIn("x_import_custom_attribute_value_1", import_headers)
        self.assertNotIn("x_import_custom_attribute_1", import_headers)
        self.assertEqual(
            chinese_labels[import_headers.index("x_material_type")],
            "\u7269\u6599\u7c7b\u578b",
        )
        self.assertEqual(chinese_labels[import_headers.index("x_component_material")], "\u6750\u6599")
        self.assertEqual(chinese_labels[import_headers.index("x_component_specification")], "\u5c3a\u5bf8\u89c4\u683c")
        self.assertEqual(chinese_labels[import_headers.index("x_component_color")], "\u989c\u8272")
        self.assertIn("产品名称", chinese_labels)
        self.assertIn("Import Matched Attribute", chinese_labels)

        custom_attribute_rows = list(workbook["自定义属性列表"].iter_rows(values_only=True))
        flattened_attribute_rows = "\n".join(
            " ".join(str(value or "") for value in row)
            for row in custom_attribute_rows
        )
        self.assertIn("Import Matched Attribute", flattened_attribute_rows)
        self.assertIn("Import Matched Value", flattened_attribute_rows)

    def test_product_import_preserves_multiple_custom_attribute_values_with_one_column_per_attribute(self):
        self.env["stock.subwarehouse.product.attribute.apply.wizard"].create({
            "attribute_name": "测试属性",
            "value_name": "default",
        }).action_apply()
        self.env["stock.subwarehouse.product.attribute.apply.wizard"].create({
            "attribute_name": "测试属性2号",
            "value_name": "default",
        }).action_apply()
        self.env["product.attribute"].search([
            ("name", "=", "测试属性"),
        ], limit=1).sequence = -100
        self.env["product.attribute"].search([
            ("name", "=", "测试属性2号"),
        ], limit=1).sequence = -99

        result = self.env["product.template"].load(
            [
                "name",
                "default_code",
                "x_import_custom_attribute_value_1",
                "x_import_custom_attribute_value_2",
            ],
            [[
                "Imported Multi Custom Attribute Product",
                "IMPORTED-MULTI-CUSTOM-ATTRIBUTE",
                "全场9九折 这个商品非常好",
                "第二个属性也应该显示导入值",
            ]],
        )

        self.assertFalse(result["messages"])
        product = self.env["product.template"].browse(result["ids"][0])
        self.assertTrue(product.is_storable)
        values_by_attribute = {
            value.attribute_id.name: value.value_text
            for value in product.x_custom_attribute_value_ids
        }
        self.assertEqual(values_by_attribute["测试属性"], "全场9九折 这个商品非常好")
        self.assertEqual(values_by_attribute["测试属性2号"], "第二个属性也应该显示导入值")

    def test_product_import_without_is_storable_still_tracks_inventory(self):
        result = self.env["product.template"].load(
            ["name", "default_code", "type", "x_material_type"],
            [["Imported Stock Product", "IMPORTED-STOCK-PRODUCT", "consu", "component"]],
        )

        self.assertFalse(result["messages"])
        product_template = self.env["product.template"].browse(result["ids"][0])
        self.assertTrue(product_template.is_storable)
        self.assertEqual(product_template.x_material_type, "component")

        product = product_template.product_variant_id
        action = self.bin_a.action_manufacture_product()
        production = self.env["mrp.production"].with_context(action["context"]).create({
            "product_id": product.id,
            "product_qty": 1.0,
            "product_uom_id": product.uom_id.id,
            "picking_type_id": self.warehouse.manu_type_id.id,
        })
        production.action_confirm()
        production.qty_producing = 1.0
        production.button_mark_done()

        quantity = self.StockQuant._get_available_quantity(product, self.bin_a)
        self.assertEqual(quantity, 1.0)

    def test_component_import_preserves_material_specification_and_color(self):
        result = self.env["product.template"].load(
            [
                "name",
                "default_code",
                "type",
                "x_material_type",
                "x_component_material",
                "x_component_specification",
                "x_component_color",
            ],
            [[
                "\u5f39\u7c27\u5957\u7ba1",
                "DBX-19",
                "consu",
                "component",
                "\u4e0d\u9508\u94a2",
                "\u03c68.5*8.1*52",
                "\u672c\u8272",
            ]],
        )

        self.assertFalse(result["messages"])
        product_template = self.env["product.template"].browse(result["ids"][0])
        self.assertEqual(product_template.x_material_type, "component")
        self.assertEqual(product_template.x_component_material, "\u4e0d\u9508\u94a2")
        self.assertEqual(product_template.x_component_specification, "\u03c68.5*8.1*52")
        self.assertEqual(product_template.x_component_color, "\u672c\u8272")

    def test_product_import_skips_existing_and_in_file_duplicate_codes(self):
        existing_code = "IMPORT-UNIQUE-EXISTING"
        self.env["product.template"].create({
            "name": "Existing import uniqueness product",
            "default_code": existing_code,
        })
        batch = self.env["stock.subwarehouse.product.import.result.batch"].create({
            "import_job_id": 987654,
        })

        result = self.env["product.template"].with_context(
            import_file=True,
            product_import_result_batch_id=batch.id,
            product_import_has_headers=True,
        ).load(
            ["name", "default_code"],
            [
                ["Unique import product", "IMPORT-UNIQUE-NEW"],
                ["Existing duplicate", existing_code.lower()],
                ["In-file duplicate", "import-unique-new"],
                ["Missing code", ""],
            ],
        )

        self.assertEqual(len(result["ids"]), 1, result)
        self.assertEqual(len(result["x_product_import_failures"]), 3)
        self.assertEqual(batch.line_ids.mapped("status").count("success"), 1)
        self.assertEqual(batch.line_ids.mapped("status").count("failed"), 3)
        self.assertEqual(batch.line_ids.mapped("source_row"), [2, 3, 4, 5])
        self.assertTrue(self.env["product.template"].search([
            ("default_code", "=", "IMPORT-UNIQUE-NEW"),
        ]))
        self.assertFalse(self.env["product.template"].search([
            ("name", "=", "In-file duplicate"),
        ]))

    def test_product_file_import_requires_product_code_mapping(self):
        result = self.env["product.template"].with_context(
            import_file=True,
            product_import_has_headers=True,
        ).load(
            ["name"],
            [["Missing product code column"]],
        )

        self.assertEqual(result["ids"], [])
        self.assertEqual(len(result["x_product_import_failures"]), 1)
        self.assertFalse(self.env["product.template"].search([
            ("name", "=", "Missing product code column"),
        ]))

    def test_product_import_preserves_legacy_repeated_custom_attribute_columns(self):
        self.env["stock.subwarehouse.product.attribute.apply.wizard"].create({
            "attribute_name": "Legacy Attribute A",
            "value_name": "default",
        }).action_apply()
        self.env["stock.subwarehouse.product.attribute.apply.wizard"].create({
            "attribute_name": "Legacy Attribute B",
            "value_name": "default",
        }).action_apply()

        result = self.env["product.template"].load(
            [
                "name",
                "default_code",
                "x_custom_attribute_value_ids/attribute_id",
                "x_custom_attribute_value_ids/value_text",
                "x_custom_attribute_value_ids/attribute_id",
                "x_custom_attribute_value_ids/value_text",
            ],
            [[
                "Imported Legacy Custom Attribute Product",
                "IMPORTED-LEGACY-CUSTOM-ATTRIBUTE",
                "Legacy Attribute A",
                "Legacy Value A",
                "Legacy Attribute B",
                "Legacy Value B",
            ]],
        )

        self.assertFalse(result["messages"])
        product = self.env["product.template"].browse(result["ids"][0])
        values_by_attribute = {
            value.attribute_id.name: value.value_text
            for value in product.x_custom_attribute_value_ids
        }
        self.assertEqual(values_by_attribute["Legacy Attribute A"], "Legacy Value A")
        self.assertEqual(values_by_attribute["Legacy Attribute B"], "Legacy Value B")

    def test_mrp_production_import_template_is_available_on_import_page(self):
        templates = self.env["mrp.production"].get_import_templates()

        self.assertEqual(
            [template["template"] for template in templates],
            ["/stock_subwarehouse_hierarchy/import_template/mrp_production.xlsx"],
        )

    def test_mrp_production_import_template_uses_current_product_attributes(self):
        from openpyxl import load_workbook

        attribute = self.env["product.attribute"].create({
            "name": "Template Dynamic Attribute",
            "create_variant": "no_variant",
        })
        value = self.env["product.attribute.value"].create({
            "name": "Template Dynamic Value",
            "attribute_id": attribute.id,
        })

        content = self.env["mrp.production"]._generate_dynamic_import_template_xlsx()
        workbook = load_workbook(BytesIO(content), read_only=True)

        import_headers = next(workbook["制造单导入"].iter_rows(values_only=True))
        self.assertIn("product_id", import_headers)
        self.assertIn("product_qty", import_headers)
        self.assertIn("never_product_template_attribute_value_ids", import_headers)
        import_rows = list(workbook["制造单导入"].iter_rows(values_only=True))
        self.assertEqual(import_rows[1][import_headers.index("name")], "MO-IMPORT-001")
        self.assertIsNone(import_rows[2][import_headers.index("name")])
        self.assertTrue(import_rows[2][import_headers.index("product_id")])

        attribute_rows = list(workbook["产品属性"].iter_rows(values_only=True))
        self.assertIn(
            (attribute.id, attribute.display_name, attribute.create_variant, value.id, value.display_name),
            attribute_rows,
        )

    def test_bom_import_template_is_available_on_import_page(self):
        templates = self.env["mrp.bom"].get_import_templates()

        self.assertEqual(
            [template["template"] for template in templates],
            ["/stock_subwarehouse_hierarchy/import_template/mrp_bom.xlsx"],
        )
        self.assertTrue(
            getattr(type(self.env["mrp.bom"]).get_import_templates, "_api_model", False),
            "BOM import templates must be callable by the import screen without record ids.",
        )

    def test_product_bom_button_opens_component_bom_or_new_form(self):
        self.env["mrp.bom"].create({
            "product_tmpl_id": self.product_a.product_tmpl_id.id,
            "product_qty": 1,
            "product_uom_id": self.product_a.uom_id.id,
            "type": "normal",
        })
        bom = self.env["mrp.bom"].create({
            "product_tmpl_id": self.product_a.product_tmpl_id.id,
            "product_qty": 1,
            "product_uom_id": self.product_a.uom_id.id,
            "type": "normal",
            "bom_line_ids": [(0, 0, {
                "product_id": self.product_b.id,
                "product_qty": 2,
                "product_uom_id": self.product_b.uom_id.id,
            })],
        })

        existing_action = self.product_a.product_tmpl_id.action_configure_product_bom()
        self.assertEqual(existing_action["res_model"], "mrp.bom")
        self.assertEqual(existing_action["res_id"], bom.id)
        self.assertEqual(existing_action["context"]["default_product_tmpl_id"], self.product_a.product_tmpl_id.id)

        product = self.env["product.product"].create({
            "name": "Finished Without BOM Yet",
            "is_storable": True,
            "x_material_type": "finished",
        })
        new_action = product.product_tmpl_id.action_configure_product_bom()
        self.assertEqual(new_action["res_model"], "mrp.bom")
        self.assertNotIn("res_id", new_action)
        self.assertEqual(new_action["context"]["default_product_tmpl_id"], product.product_tmpl_id.id)

    def test_mrp_production_uses_product_bom_components_by_default(self):
        self.product_a.product_tmpl_id.x_material_type = "finished"
        self.product_b.product_tmpl_id.x_material_type = "component"
        bom = self.env["mrp.bom"].create({
            "product_tmpl_id": self.product_a.product_tmpl_id.id,
            "product_qty": 1,
            "product_uom_id": self.product_a.uom_id.id,
            "type": "normal",
            "bom_line_ids": [(0, 0, {
                "product_id": self.product_b.id,
                "product_qty": 2,
                "product_uom_id": self.product_b.uom_id.id,
            })],
        })

        production = self.env["mrp.production"].create({
            "product_id": self.product_a.id,
            "product_qty": 3,
            "product_uom_id": self.product_a.uom_id.id,
            "location_src_id": self.bin_a.id,
            "location_dest_id": self.bin_a.id,
        })

        self.assertEqual(production.bom_id, bom)
        raw_move = production.move_raw_ids.filtered(lambda move: move.product_id == self.product_b)
        self.assertEqual(len(raw_move), 1)
        self.assertEqual(raw_move.product_uom_qty, 6)

    def test_mrp_production_form_onchange_shows_bom_components(self):
        self.product_a.product_tmpl_id.x_material_type = "finished"
        self.product_b.product_tmpl_id.x_material_type = "component"
        self.env["mrp.bom"].create({
            "product_tmpl_id": self.product_a.product_tmpl_id.id,
            "product_qty": 1,
            "product_uom_id": self.product_a.uom_id.id,
            "type": "normal",
        })
        bom = self.env["mrp.bom"].create({
            "product_tmpl_id": self.product_a.product_tmpl_id.id,
            "product_qty": 1,
            "product_uom_id": self.product_a.uom_id.id,
            "type": "normal",
            "bom_line_ids": [(0, 0, {
                "product_id": self.product_b.id,
                "product_qty": 2,
                "product_uom_id": self.product_b.uom_id.id,
            })],
        })

        production = self.env["mrp.production"].new({
            "product_id": self.product_a.id,
            "product_qty": 3,
            "product_uom_id": self.product_a.uom_id.id,
            "location_src_id": self.bin_a.id,
            "location_dest_id": self.bin_a.id,
            "company_id": self.env.company.id,
        })
        production._onchange_product_id_use_default_bom_components()

        self.assertEqual(production.bom_id, bom)
        raw_move = production.move_raw_ids.filtered(lambda move: move.product_id == self.product_b)
        self.assertEqual(len(raw_move), 1)
        self.assertEqual(raw_move.product_uom_qty, 2)

    def test_bom_import_template_can_create_component_lines(self):
        self.product_b.default_code = "BOM-COMPONENT-B"
        self.product_b.product_tmpl_id.x_material_type = "component"
        product_c = self.env["product.product"].create({
            "name": "BOM Component C",
            "default_code": "BOM-COMPONENT-C",
            "is_storable": True,
        })

        result = self.env["mrp.bom"].load(
            [
                "product_tmpl_id",
                "product_qty",
                "product_uom_id",
                "type",
                "code",
                "x_import_bom_component_product",
                "x_import_bom_component_qty",
                "x_import_bom_component_uom",
            ],
            [
                [
                    self.product_a.product_tmpl_id.display_name,
                    1,
                    self.product_a.uom_id.display_name,
                    "normal",
                    "BOM-IMPORT-TEST",
                    "BOM-COMPONENT-B",
                    2,
                    self.product_b.uom_id.display_name,
                ],
                ["", "", "", "", "", "BOM-COMPONENT-C", 3, product_c.uom_id.display_name],
            ],
        )

        self.assertFalse(result["messages"])
        bom = self.env["mrp.bom"].browse(result["ids"][0])
        self.assertEqual(bom.code, "BOM-IMPORT-TEST")
        self.assertEqual(len(bom.bom_line_ids), 2)
        self.assertEqual(bom.bom_line_ids.filtered(lambda line: line.product_id == self.product_b).product_qty, 2)
        self.assertEqual(bom.bom_line_ids.filtered(lambda line: line.product_id == product_c).product_qty, 3)

    def test_bom_import_supports_more_than_twenty_one_component_rows(self):
        components = self.env["product.product"].create([
            {
                "name": f"Unlimited BOM Component {index}",
                "default_code": f"BOM-UNLIMITED-{index:02d}",
                "is_storable": True,
            }
            for index in range(25)
        ])
        import_fields = [
            "product_tmpl_id",
            "product_qty",
            "product_uom_id",
            "type",
            "code",
            "x_import_bom_component_product",
            "x_import_bom_component_qty",
            "x_import_bom_component_uom",
        ]
        import_data = []
        for index, component in enumerate(components):
            import_data.append([
                self.product_a.product_tmpl_id.display_name if index == 0 else "",
                1 if index == 0 else "",
                self.product_a.uom_id.display_name if index == 0 else "",
                "normal" if index == 0 else "",
                "BOM-UNLIMITED-TEST" if index == 0 else "",
                component.default_code,
                index + 1,
                component.uom_id.display_name,
            ])

        result = self.env["mrp.bom"].load(import_fields, import_data)

        self.assertFalse(result["messages"])
        bom = self.env["mrp.bom"].browse(result["ids"][0])
        self.assertEqual(len(bom.bom_line_ids), 25)
        self.assertEqual(bom.bom_line_ids[-1].product_qty, 25)

    def test_bom_import_fails_when_component_product_is_missing(self):
        with self.assertRaisesRegex(UserError, "MISSING-COMPONENT-REF"):
            self.env["mrp.bom"].load(
                [
                    "product_tmpl_id",
                    "product_qty",
                    "product_uom_id",
                    "type",
                    "code",
                    "x_import_bom_component_product",
                    "x_import_bom_component_qty",
                    "x_import_bom_component_uom",
                ],
                [[
                    self.product_a.product_tmpl_id.display_name,
                    1,
                    self.product_a.uom_id.display_name,
                    "normal",
                    "BOM-MISSING-COMPONENT-TEST",
                    "MISSING-COMPONENT-REF",
                    2,
                    self.product_b.uom_id.display_name,
                ]],
            )

    def test_bom_import_export_template_matches_vertical_component_rows(self):
        from openpyxl import load_workbook

        self.product_b.default_code = "BOM-EXPORT-COMPONENT"
        bom = self.env["mrp.bom"].create({
            "product_tmpl_id": self.product_a.product_tmpl_id.id,
            "product_qty": 1,
            "product_uom_id": self.product_a.uom_id.id,
            "code": "BOM-EXPORT-TEST",
            "bom_line_ids": [(0, 0, {
                "product_id": self.product_b.id,
                "product_qty": 3,
                "product_uom_id": self.product_b.uom_id.id,
            })],
        })

        import_workbook = load_workbook(
            BytesIO(self.env["mrp.bom"]._generate_bom_import_template_xlsx()),
            read_only=True,
        )
        import_rows = list(import_workbook["物料清单导入"].iter_rows(values_only=True))
        self.assertIn("x_import_bom_component_product", import_rows[0])
        self.assertNotIn("x_import_bom_component_product_1", import_rows[0])
        self.assertEqual(
            import_rows[1][import_rows[0].index("x_import_bom_component_product")],
            "组件产品",
        )
        self.assertEqual(import_rows[3][import_rows[0].index("product_tmpl_id")], None)

        export_workbook = load_workbook(
            BytesIO(bom._generate_bom_export_xlsx()),
            read_only=True,
        )
        export_rows = list(export_workbook["物料清单导出"].iter_rows(values_only=True))
        self.assertEqual(
            export_rows[0],
            tuple(field_name for field_name, _label in self.env["mrp.bom"]._get_bom_import_template_columns()),
        )
        self.assertEqual(export_rows[2][export_rows[0].index("code")], "BOM-EXPORT-TEST")
        self.assertEqual(
            export_rows[2][export_rows[0].index("x_import_bom_component_product")],
            "BOM-EXPORT-COMPONENT",
        )
        self.assertEqual(export_rows[2][export_rows[0].index("x_import_bom_component_qty")], 3)

    def test_bom_can_be_saved_and_applied_as_reusable_template(self):
        source_bom = self.env["mrp.bom"].create({
            "product_tmpl_id": self.product_a.product_tmpl_id.id,
            "product_qty": 1,
            "product_uom_id": self.product_a.uom_id.id,
            "code": "BOM-REUSABLE-SOURCE",
            "bom_line_ids": [Command.create({
                "product_id": self.product_b.id,
                "product_qty": 4,
                "product_uom_id": self.product_b.uom_id.id,
            })],
        })
        action = source_bom.action_save_as_reusable_template()
        template = self.env["stock.subwarehouse.bom.template"].browse(action["res_id"])
        target_product = self.env["product.product"].create({
            "name": "Reusable BOM Target",
            "is_storable": True,
        })
        target_bom = self.env["mrp.bom"].create({
            "product_tmpl_id": target_product.product_tmpl_id.id,
            "product_qty": 1,
            "product_uom_id": target_product.uom_id.id,
            "x_reusable_template_id": template.id,
        })

        target_bom.action_apply_reusable_template()

        self.assertEqual(len(template.line_ids), 1)
        self.assertEqual(target_bom.bom_line_ids.product_id, self.product_b)
        self.assertEqual(target_bom.bom_line_ids.product_qty, 4)

    def test_sale_order_import_template_uses_product_internal_reference_and_chinese_labels(self):
        from openpyxl import load_workbook

        templates = self.env["sale.order"].get_import_templates()
        self.assertEqual(
            [template["template"] for template in templates],
            ["/stock_subwarehouse_hierarchy/import_template/sale_order.xlsx"],
        )

        content = self.env["sale.order"]._generate_sale_order_import_template_xlsx()
        workbook = load_workbook(BytesIO(content), read_only=True)
        rows = list(workbook["报价单导入"].iter_rows(values_only=True))

        self.assertEqual(rows[0][0], "Order Reference")
        self.assertEqual(rows[1][0], "订单号")
        self.assertEqual(rows[0][:15], (
            "Order Reference",
            "Customer*",
            "Order Date",
            "x_platform",
            "x_channel",
            "Salesperson",
            "x_sale_nature",
            "Order Lines/Products*",
            "Order Lines/x_import_product_name",
            "Order Lines/x_color",
            "Order Lines/x_size",
            "Order Lines/x_flex",
            "Order Lines/Quantity",
            "Order Lines/Unit Price",
            "Order Lines/x_source_location_id",
        ))
        self.assertIn("Order Lines/Products*", rows[0])
        self.assertNotIn("order_line/product_id/default_code", rows[0])
        self.assertNotIn("order_line/product_id/.id", rows[0])
        self.assertNotIn("order_line/product_id", rows[0])
        self.assertEqual(
            rows[1][rows[0].index("Order Lines/Products*")],
            "产品ID",
        )
        self.assertEqual(rows[1][rows[0].index("Order Lines/x_color")], "颜色")
        self.assertEqual(rows[1][rows[0].index("Order Lines/x_source_location_id")], "发货仓库")

    def test_template_format_exports_match_import_headers(self):
        from openpyxl import load_workbook

        self.product_a.default_code = "EXPORT-PRODUCT-A"
        self.product_a.product_tmpl_id.x_material_type = "component"
        export_attribute_name = "Export Custom Attribute"
        export_attribute_value = "Export Real Custom Value 88"
        self.env["stock.subwarehouse.product.attribute.apply.wizard"].create({
            "attribute_name": export_attribute_name,
            "value_name": export_attribute_value,
        }).action_apply()
        sale_order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "x_platform": "有赞",
            "x_channel": "凌动雪具",
            "x_sale_nature": self.env.ref(
                "stock_subwarehouse_hierarchy.sale_nature_retail"
            ).id,
            "x_finance_remark": "财务备注",
            "order_line": [(
                0,
                0,
                {
                    "product_id": self.product_a.id,
                    "product_uom_qty": 2,
                    "price_unit": 15,
                    "x_import_product_name": "双板鞋",
                    "x_color": "黑色",
                    "x_size": "260",
                    "x_flex": "硬度100",
                    "x_source_location_id": self.bin_a.id,
                },
            )],
        })
        production = self.env["mrp.production"].create({
            "product_id": self.product_a.id,
            "product_qty": 3,
            "product_uom_id": self.product_a.uom_id.id,
            "location_src_id": self.bin_a.id,
            "location_dest_id": self.bin_a.id,
        })

        product_workbook = load_workbook(
            BytesIO(self.product_a.product_tmpl_id._generate_dynamic_product_export_xlsx()),
            read_only=True,
        )
        product_rows = list(product_workbook["产品导出"].iter_rows(values_only=True))
        self.assertEqual(
            product_rows[0],
            tuple(field_name for field_name, _label in self.env["product.template"]._get_dynamic_product_import_columns()),
        )
        self.assertEqual(product_rows[2][product_rows[0].index("default_code")], "EXPORT-PRODUCT-A")
        self.assertEqual(product_rows[2][product_rows[0].index("x_material_type")], "component")
        self.assertNotIn(export_attribute_name, product_rows[2])
        self.assertIn(export_attribute_name, product_rows[1])
        attribute_column_index = product_rows[1].index(export_attribute_name)
        self.assertTrue(product_rows[0][attribute_column_index].startswith("x_import_custom_attribute_value_"))
        self.assertEqual(product_rows[2][attribute_column_index], export_attribute_value)

        manufacturing_workbook = load_workbook(
            BytesIO(production._generate_dynamic_export_xlsx()),
            read_only=True,
        )
        manufacturing_rows = list(manufacturing_workbook["制造单导出"].iter_rows(values_only=True))
        self.assertEqual(
            manufacturing_rows[0],
            tuple(field_name for field_name, _label in self.env["mrp.production"]._get_dynamic_import_template_columns()),
        )
        self.assertEqual(manufacturing_rows[2][manufacturing_rows[0].index("product_qty")], 3)

        sale_workbook = load_workbook(
            BytesIO(sale_order._generate_sale_order_export_xlsx()),
            read_only=True,
        )
        sale_rows = list(sale_workbook["报价单导出"].iter_rows(values_only=True))
        self.assertEqual(
            sale_rows[0],
            tuple(field_name for field_name, _label in self.env["sale.order"]._get_sale_order_import_template_columns()),
        )
        self.assertEqual(sale_rows[1][sale_rows[0].index("Order Lines/Products*")], "产品ID")
        self.assertEqual(sale_rows[2][sale_rows[0].index("Order Lines/Products*")], "EXPORT-PRODUCT-A")
        self.assertEqual(sale_rows[2][sale_rows[0].index("x_platform")], "有赞")
        self.assertEqual(sale_rows[2][sale_rows[0].index("x_sale_nature")], "零售")

        for expected_column in (
            "x_official_total",
            "x_processing_fee",
            "x_amount_received",
            "x_is_external_order",
        ):
            self.assertIn(expected_column, sale_rows[0])

        self.assertEqual(sale_rows[2][sale_rows[0].index("Order Lines/x_import_product_name")], "双板鞋")
        self.assertEqual(sale_rows[2][sale_rows[0].index("Order Lines/x_color")], "黑色")
        self.assertEqual(sale_rows[2][sale_rows[0].index("Order Lines/x_size")], "260")
        self.assertEqual(sale_rows[2][sale_rows[0].index("Order Lines/x_flex")], "硬度100")
        self.assertEqual(sale_rows[2][sale_rows[0].index("Order Lines/x_source_location_id")], self.bin_a.display_name)
        self.assertEqual(sale_rows[2][sale_rows[0].index("x_finance_remark")], "财务备注")

        product_action = self.product_a.product_tmpl_id.action_export_import_template_format()
        manufacturing_action = production.action_export_import_template_format()
        sale_action = sale_order.action_export_import_template_format()
        self.assertIn("/stock_subwarehouse_hierarchy/export/product_template.xlsx", product_action["url"])
        self.assertIn("/stock_subwarehouse_hierarchy/export/mrp_production.xlsx", manufacturing_action["url"])
        self.assertIn("/stock_subwarehouse_hierarchy/export/sale_order.xlsx", sale_action["url"])

    def test_sale_import_ignores_official_total_and_finalizes_external_order(self):
        self.product_a.lst_price = 180.0
        self.product_a.default_code = "EXT-IMPORT-PRODUCT"
        result = self.env["sale.order"].with_context(import_file=True).load(
            [
                "name",
                "partner_id",
                "order_line/product_id",
                "order_line/product_uom_qty",
                "x_official_total",
                "x_processing_fee",
                "x_amount_received",
                "x_is_external_order",
            ],
            [[
                "EXT-IMPORT-001",
                self.customer.display_name,
                self.product_a.default_code,
                2.0,
                1.0,
                8.0,
                333.0,
                "True",
            ]],
        )

        self.assertFalse(result["messages"])
        order = self.env["sale.order"].browse(result["ids"])
        self.assertEqual(order.state, "external_done")
        self.assertEqual(order.x_official_total, 360.0)
        self.assertEqual(order.x_processing_fee, 8.0)
        self.assertEqual(order.x_amount_received, 333.0)
        self.assertFalse(order.picking_ids)

    def test_sale_nature_import_can_create_unmatched_value(self):
        sale_nature_field = self.env["sale.order"]._fields["x_sale_nature"]
        self.assertEqual(sale_nature_field.type, "many2one")
        self.assertEqual(
            sale_nature_field.comodel_name,
            "stock.subwarehouse.sale.nature",
        )
        created_id, created_name = self.env[
            "stock.subwarehouse.sale.nature"
        ].name_create("直播团购")
        created = self.env["stock.subwarehouse.sale.nature"].browse(created_id)
        self.assertEqual(created_name, "直播团购")
        self.assertEqual(created.name, "直播团购")
        self.assertTrue(created.code)

        order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "x_sale_nature": created.id,
        })
        self.assertEqual(order.x_sale_nature, created)

    def test_sale_import_skips_unknown_product_without_blocking_valid_order(self):
        self.product_a.default_code = "SALE-IMPORT-PARTIAL-VALID"
        fields_to_import = [
            "name",
            "partner_id",
            "order_line/product_id",
            "order_line/product_uom_qty",
        ]
        rows = [
            [
                "SALE-IMPORT-GOOD",
                self.customer.display_name,
                self.product_a.default_code,
                2,
            ],
            [
                "SALE-IMPORT-BAD",
                self.customer.display_name,
                "SALE-IMPORT-DOES-NOT-EXIST",
                1,
            ],
        ]

        result = self.env["sale.order"].with_context(
            import_file=True,
            import_skip_records=[],
            sale_import_has_headers=True,
        ).load(fields_to_import, rows)

        self.assertFalse(result["messages"])
        self.assertEqual(len(result["ids"]), 1, result)
        imported_order = self.env["sale.order"].browse(result["ids"])
        self.assertEqual(imported_order.name, "SALE-IMPORT-GOOD")
        self.assertEqual(imported_order.order_line.product_id, self.product_a)
        self.assertEqual(
            result["x_sale_import_skipped_rows"],
            [{
                "source_row": 3,
                "product_id": "SALE-IMPORT-DOES-NOT-EXIST",
                "reason": "产品ID在ERP中不存在。",
            }],
        )

    def test_sale_import_skips_blank_product_and_preserves_continuation_order(self):
        self.product_a.default_code = "SALE-IMPORT-CONTINUATION-VALID"
        fields_to_import = [
            "name",
            "partner_id",
            "order_line/product_id",
            "order_line/product_uom_qty",
        ]
        rows = [
            ["SALE-IMPORT-CONTINUATION", self.customer.display_name, "", 1],
            ["", "", self.product_a.default_code, 3],
        ]

        result = self.env["sale.order"].with_context(
            import_file=True,
            import_skip_records=["order_line/product_id"],
            sale_import_has_headers=True,
        ).load(fields_to_import, rows)

        self.assertFalse(result["messages"])
        self.assertEqual(len(result["ids"]), 1, result)
        imported_order = self.env["sale.order"].browse(result["ids"])
        self.assertEqual(imported_order.name, "SALE-IMPORT-CONTINUATION")
        self.assertEqual(imported_order.partner_id, self.customer)
        self.assertEqual(imported_order.order_line.product_id, self.product_a)
        self.assertEqual(imported_order.order_line.product_uom_qty, 3)
        self.assertEqual(result["x_sale_import_skipped_rows"][0]["source_row"], 2)
        self.assertEqual(result["x_sale_import_skipped_rows"][0]["reason"], "产品ID为空。")

    def test_sale_import_skips_group_missing_required_parent_value(self):
        self.product_a.default_code = "SALE-IMPORT-REQUIRED-PARENT"
        fields_to_import = [
            "name",
            "partner_id",
            "order_line/product_id",
        ]
        rows = [
            ["SALE-IMPORT-NO-CUSTOMER", "", self.product_a.default_code],
            ["SALE-IMPORT-WITH-CUSTOMER", self.customer.display_name, self.product_a.default_code],
        ]

        result = self.env["sale.order"].with_context(
            import_file=True,
            import_skip_records=["partner_id", "order_line/product_id"],
            sale_import_has_headers=True,
        ).load(fields_to_import, rows)

        self.assertFalse(result["messages"])
        self.assertEqual(len(result["ids"]), 1)
        self.assertEqual(self.env["sale.order"].browse(result["ids"]).name, "SALE-IMPORT-WITH-CUSTOMER")
        self.assertEqual(result["x_sale_import_skipped_rows"][0]["source_row"], 2)
        self.assertEqual(
            result["x_sale_import_skipped_rows"][0]["reason"],
            "必填字段“客户”缺少值。",
        )

    def test_sale_import_skip_warning_uses_chinese_field_label_for_salesperson(self):
        self.product_a.default_code = "SALE-IMPORT-CHINESE-WARNING"
        result = self.env["sale.order"].with_context(
            lang="en_US",
            import_file=True,
            import_skip_records=["user_id", "order_line/product_id"],
            sale_import_has_headers=True,
        ).load(
            ["name", "partner_id", "user_id", "order_line/product_id"],
            [[
                "SALE-IMPORT-NO-SALESPERSON",
                self.customer.display_name,
                "",
                self.product_a.default_code,
            ]],
        )

        self.assertFalse(result["ids"])
        self.assertFalse(result["messages"])
        self.assertEqual(
            result["x_sale_import_skipped_rows"][0]["reason"],
            "必填字段“销售员”缺少值。",
        )

    def test_sale_import_skips_blank_required_product_even_without_skip_policy(self):
        self.product_a.default_code = "SALE-IMPORT-BLANK-AUTOMATIC"
        fields_to_import = [
            "name",
            "partner_id",
            "order_line/product_id",
        ]
        rows = [
            ["SALE-IMPORT-BLANK", self.customer.display_name, ""],
            ["SALE-IMPORT-NONBLANK", self.customer.display_name, self.product_a.default_code],
        ]

        result = self.env["sale.order"].with_context(
            import_file=True,
            import_skip_records=[],
            sale_import_has_headers=True,
        ).load(fields_to_import, rows)

        self.assertFalse(result["messages"])
        self.assertEqual(len(result["ids"]), 1, result)
        self.assertEqual(self.env["sale.order"].browse(result["ids"]).name, "SALE-IMPORT-NONBLANK")
        self.assertEqual(result["x_sale_import_skipped_rows"][0]["source_row"], 2)
        self.assertEqual(result["x_sale_import_skipped_rows"][0]["reason"], "产品ID为空。")

    def test_sale_file_import_test_and_formal_import_both_allow_partial_success(self):
        self.product_a.default_code = "SALE-IMPORT-FILE-VALID"
        self.product_a.flush_recordset(["default_code"])
        csv_content = (
            "Order Reference,Customer,Product ID,Quantity\n"
            f"SALE-FILE-GOOD,{self.customer.display_name},{self.product_a.default_code},2\n"
            f"SALE-FILE-UNKNOWN,{self.customer.display_name},SALE-FILE-NOT-IN-ERP,1\n"
            f"SALE-FILE-BLANK,{self.customer.display_name},,1\n"
        ).encode("utf-8")
        importer = self.env["base_import.import"].create({
            "res_model": "sale.order",
            "file": csv_content,
            "file_name": "sale_partial_success.csv",
        })
        fields_to_import = [
            "name",
            "partner_id",
            "order_line/product_id",
            "order_line/product_uom_qty",
        ]
        columns = ["order reference", "customer", "product id", "quantity"]
        options = {
            "advanced": True,
            "has_headers": True,
            "headers": True,
            "skip": 0,
            "limit": 2000,
            "encoding": "utf-8",
            "separator": ",",
            "quoting": '"',
            "date_format": "",
            "datetime_format": "",
            "float_thousand_separator": ",",
            "float_decimal_separator": ".",
            "import_skip_records": [],
            "import_set_empty_fields": [],
            "name_create_enabled_fields": {},
            "fallback_values": {},
        }

        dry_result = importer.execute_import(
            fields_to_import, columns, dict(options), dryrun=True
        )
        formal_result = importer.execute_import(
            fields_to_import, columns, dict(options), dryrun=False
        )

        self.assertFalse(dry_result["messages"])
        self.assertEqual(len(dry_result["ids"]), 1, dry_result)
        self.assertEqual(len(dry_result["x_sale_import_skipped_rows"]), 2)
        self.assertFalse(formal_result["messages"])
        self.assertEqual(len(formal_result["ids"]), 1, formal_result)
        self.assertEqual(len(formal_result["x_sale_import_skipped_rows"]), 2)
        imported_order = self.env["sale.order"].browse(formal_result["ids"])
        self.assertEqual(imported_order.name, "SALE-FILE-GOOD")
        self.assertEqual(imported_order.order_line.product_id, self.product_a)
        self.assertFalse(self.env["sale.order"].search([
            ("name", "in", ["SALE-FILE-UNKNOWN", "SALE-FILE-BLANK"]),
        ]))

    def test_new_products_do_not_get_default_product_taxes(self):
        product = self.env["product.template"].create({
            "name": "No Default Tax Product",
            "is_storable": True,
        })

        self.assertFalse(product.taxes_id)
        self.assertFalse(product.supplier_taxes_id)

    def test_international_product_mapping_import_updates_website_fields(self):
        from openpyxl import Workbook

        product = self.env["product.template"].create({
            "name": "\u6d4b\u8bd5\u53cc\u677f\u978b",
            "is_storable": True,
            "list_price": 599.0,
            "default_code": "012307S2-MA100-H001220",
        })
        second_product = self.env["product.template"].create({
            "name": "\u6d4b\u8bd5\u53cc\u677f\u978b",
            "is_storable": True,
            "list_price": 699.0,
            "default_code": "012307S2-MA120-H001220",
        })
        unmatched_product = self.env["product.template"].create({
            "name": "\u672a\u5339\u914d\u4ea7\u54c1",
            "is_storable": True,
            "default_code": "012307S2-MA100-H0012200",
        })
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Pattern Review"
        worksheet.append(["SUN International Pattern Mapping"])
        worksheet.append([])
        worksheet.append(["\u5e8f\u53f7", "\u4e2d\u6587\u540d\u79f0", "\u82f1\u6587\u540d\u79f0", "USD \u4ef7\u683c", "product_code_pattern"])
        worksheet.append([1, "\u53cc\u677f\u978b", "Generic Ski Boots 100 flex", 545, "******S2-**100-*******"])
        worksheet.append([2, "\u53cc\u677f\u978b", "Ski Boots 100 flex", 595, "******S2-MA100-*******"])
        worksheet.append([3, "\u53cc\u677f\u978b", "Ski Boots 120 flex", 650, "******S2-**120-*******"])
        worksheet.append([4, "\u7a7a\u767d\u6a21\u5f0f", "Ignored", 1, ""])
        content = BytesIO()
        workbook.save(content)

        wizard = self.env[
            "stock.subwarehouse.product.international.mapping.import.wizard"
        ].create({
            "import_file": base64.b64encode(content.getvalue()),
            "import_filename": "international_price_list.xlsx",
        })
        action = wizard.action_import_mapping()

        self.assertEqual(product.x_website_english_name, "Ski Boots 100 flex")
        self.assertEqual(product.x_website_usd_price, 595.0)
        self.assertEqual(product._get_website_display_name(True), "Ski Boots 100 flex")
        self.assertEqual(product.with_context(lang="en_US").name, "\u6d4b\u8bd5\u53cc\u677f\u978b")
        self.assertEqual(product._get_website_display_price_label(True), "$595.00")
        self.assertEqual(product._get_website_display_name(False), "\u6d4b\u8bd5\u53cc\u677f\u978b")
        self.assertEqual(product._get_website_display_price_label(False), "\uffe5599.00")
        self.assertEqual(second_product.x_website_english_name, "Ski Boots 120 flex")
        self.assertEqual(second_product.x_website_usd_price, 650.0)
        self.assertFalse(unmatched_product.x_website_code_mapping_id)
        self.assertFalse(self.env["stock.subwarehouse.product.website.code.mapping"].find_matching_mapping(
            unmatched_product.default_code
        ))
        self.assertEqual(product._get_english_shop_variant_value("color", "\u9ed1\u767d"), "Black / White")
        self.assertEqual(product._get_english_shop_variant_value("size", "\u901a\u7801"), "One size")
        self.assertEqual(product._get_english_shop_variant_value("flex", "\u786c\u5ea6100"), "100 flex")
        self.assertEqual(action["params"]["type"], "success")

        unsafe_product = self.env["product.template"].create({
            "name": "\u5b89\u5168 URL \u56de\u9000\u6d4b\u8bd5",
            "x_website_english_name": "---",
        })
        self.assertEqual(
            unsafe_product.with_context(lang="en_US").name,
            "\u5b89\u5168 URL \u56de\u9000\u6d4b\u8bd5",
        )

    def test_english_website_name_does_not_replace_canonical_chinese_product_name(self):
        product = self.env["product.template"].with_context(lang="zh_CN").create({
            "name": "\u4e2d\u6587 ERP \u4ea7\u54c1\u540d",
        })

        product.write({"x_website_english_name": "English Storefront Name"})

        self.assertEqual(product.with_context(lang="zh_CN").name, "\u4e2d\u6587 ERP \u4ea7\u54c1\u540d")
        self.assertEqual(product.with_context(lang="en_US").name, "\u4e2d\u6587 ERP \u4ea7\u54c1\u540d")
        self.assertEqual(product._get_website_display_name(True), "English Storefront Name")

    def test_website_checkout_language_uses_mapped_usd_price_and_name(self):
        website = self.env["website"].get_current_website()
        product = self.env["product.product"].create({
            "name": "中文测试雪鞋",
            "default_code": "012307S2-MA100-H001220",
            "is_storable": True,
            "list_price": 3999.0,
        })
        mapping = self.env["stock.subwarehouse.product.website.code.mapping"].create({
            "product_code_pattern": "******S2-**100-*******",
            "english_name": "Test Ski Boots 100 flex",
            "usd_price": 545.0,
        })
        product.product_tmpl_id.write({
            "x_website_code_mapping_id": mapping.id,
            "x_website_english_name": mapping.english_name,
            "x_website_usd_price": mapping.usd_price,
        })
        order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "website_id": website.id,
        })
        line = self.env["sale.order.line"].create({
            "order_id": order.id,
            "product_id": product.id,
            "product_uom_qty": 1.0,
        })
        delivery_product = self.env["product.product"].create({
            "name": "Standard delivery",
            "is_storable": False,
            "list_price": 25.0,
        })
        delivery_line = self.env["sale.order.line"].create({
            "order_id": order.id,
            "product_id": delivery_product.id,
            "name": "Standard delivery",
            "product_uom_qty": 1.0,
            "is_delivery": True,
            "price_unit": 25.0,
        })
        chinese_currency = order.currency_id

        order._apply_website_checkout_language(True)

        self.assertEqual(order.x_website_checkout_language, "en_US")
        self.assertEqual(order.currency_id, self.env.ref("base.USD"))
        self.assertEqual(order.currency_id.symbol, "$")
        self.assertEqual(line.name, "Test Ski Boots 100 flex")
        self.assertEqual(line.price_unit, 545.0)
        self.assertEqual(delivery_line.name, "Standard delivery")
        self.assertEqual(delivery_line.price_unit, 25.0)
        self.assertEqual(line._get_line_header(), "Test Ski Boots 100 flex")
        self.assertEqual(line.name_short, "Test Ski Boots 100 flex")

        order._recompute_prices()

        self.assertEqual(line.price_unit, 545.0)
        self.assertEqual(line._get_line_header(), "Test Ski Boots 100 flex")

        order._apply_website_checkout_language(False)

        self.assertEqual(order.x_website_checkout_language, "zh_CN")
        self.assertEqual(order.currency_id, chinese_currency)
        self.assertIn("中文测试雪鞋", line.name)
        self.assertEqual(line.price_unit, 3999.0)

        order._apply_website_checkout_language(True)

        self.assertEqual(order.x_website_checkout_language, "en_US")
        self.assertEqual(order.currency_id, self.env.ref("base.USD"))
        self.assertEqual(line.name, "Test Ski Boots 100 flex")
        self.assertEqual(line.price_unit, 545.0)

    def test_website_checkout_country_language_rules(self):
        SaleOrder = self.env["sale.order"]
        china = self.env.ref("base.cn")
        united_states = self.env.ref("base.us")

        self.assertTrue(SaleOrder._is_website_checkout_country_allowed(china, False))
        self.assertFalse(SaleOrder._is_website_checkout_country_allowed(china, True))
        self.assertFalse(SaleOrder._is_website_checkout_country_allowed(united_states, False))
        self.assertTrue(SaleOrder._is_website_checkout_country_allowed(united_states, True))

    def test_purchase_pages_use_english_product_and_status_labels(self):
        product = self.env["product.product"].create({
            "name": "中文雪鞋",
            "default_code": "012307S2-MA100-H001220",
            "is_storable": True,
        })
        product.product_tmpl_id.write({
            "x_website_english_name": "English Ski Boots",
        })
        order = self.env["sale.order"].create({
            "partner_id": self.customer.id,
            "website_id": self.env["website"].get_current_website().id,
            "x_website_payment_state": "paid",
        })
        line = self.env["sale.order.line"].create({
            "order_id": order.id,
            "product_id": product.id,
            "product_uom_qty": 1.0,
        })

        self.assertEqual(
            WebsitePurchaseHistory._customer_line_name(line, True), "English Ski Boots"
        )
        paid_order = SimpleNamespace(x_website_payment_state="paid")
        self.assertEqual(WebsitePurchaseHistory._payment_status(paid_order, True), "Paid")
        self.assertEqual(WebsitePurchaseHistory._payment_status(paid_order, False), "已支付")

    def test_currency_symbols_use_yuan(self):
        usd = self.env.ref("base.USD")
        cny = self.env.ref("base.CNY")

        self.env["res.currency"].action_use_yuan_symbol_everywhere()

        self.assertTrue(cny.active)
        self.assertEqual(cny.symbol, "￥")
        self.assertTrue(usd.active)
        self.assertEqual(usd.symbol, "$")
        self.assertEqual(self.env.company.currency_id, cny)
