from datetime import date, datetime
from io import BytesIO

from odoo.exceptions import ValidationError
from odoo.tests.common import TransactionCase


class TestDistributorPrototype(TransactionCase):
    def test_distributor_import_template_and_test_row(self):
        from openpyxl import load_workbook

        content = self.env[
            "res.partner"
        ]._generate_partner_channel_import_template_xlsx("distributor")
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook["经销商导入"]
        expected_fields = [
            field_name
            for field_name, _label, _required, _accepted in self.env[
                "res.partner"
            ]._get_distributor_import_columns()
        ]
        self.assertEqual(
            list(next(worksheet.iter_rows(values_only=True))),
            expected_fields,
        )

        import_fields = [
            "name",
            "x_is_distributor",
            "company_type",
            "customer_rank",
            "email",
            "x_distributor_status",
            "x_distributor_level",
            "x_distributor_territory",
            "x_distributor_exclusive",
            "x_distributor_agreement_start",
            "x_distributor_agreement_end",
            "x_distributor_annual_target",
        ]
        result = self.env["res.partner"].load(import_fields, [[
            "Imported Distributor Test",
            "1",
            "company",
            "1",
            "distributor.import@example.com",
            "active",
            "gold",
            "North China",
            "1",
            "2026-01-01",
            "2026-12-31",
            "500000",
        ]])
        self.assertFalse(result["messages"])
        distributor = self.env["res.partner"].browse(result["ids"])
        self.assertTrue(distributor.x_is_distributor)
        self.assertEqual(distributor.x_distributor_status, "active")
        self.assertEqual(distributor.x_distributor_level, "gold")
        self.assertTrue(distributor.x_distributor_code)

        action = self.env["res.partner"].action_open_distributor_import([])
        self.assertEqual(action["tag"], "import")
        self.assertEqual(
            action["context"]["partner_channel_import_type"],
            "distributor",
        )
        templates = self.env["res.partner"].with_context(
            partner_channel_import_type="distributor"
        ).get_import_templates()
        self.assertEqual(
            [template["label"] for template in templates],
            ["经销商导入模板"],
        )

        export_content = distributor._generate_partner_channel_export_xlsx(
            "distributor"
        )
        export_workbook = load_workbook(
            BytesIO(export_content),
            read_only=True,
            data_only=True,
        )
        export_rows = list(
            export_workbook["经销商导入"].iter_rows(values_only=True)
        )
        self.assertEqual(list(export_rows[0]), expected_fields)
        self.assertEqual(export_rows[1][0], "Imported Distributor Test")

    def test_distributor_profile_code_and_document_actions(self):
        distributor = self.env["res.partner"].create({
            "name": "Prototype Distributor",
            "is_company": True,
            "x_is_distributor": True,
            "x_distributor_status": "active",
            "x_distributor_territory": "North China",
        })
        order = self.env["sale.order"].create({
            "partner_id": distributor.id,
        })

        self.assertTrue(distributor.x_distributor_code.startswith("DIST/"))
        self.assertEqual(distributor.x_distributor_sale_order_count, 1)
        self.assertIn(
            ("partner_id", "child_of", distributor.id),
            distributor.action_view_distributor_sale_orders()["domain"],
        )
        self.assertEqual(order.partner_id, distributor)

    def test_distributor_agreement_dates_are_validated(self):
        with self.assertRaises(ValidationError):
            self.env["res.partner"].create({
                "name": "Invalid Distributor Agreement",
                "is_company": True,
                "x_is_distributor": True,
                "x_distributor_agreement_start": date(2026, 8, 2),
                "x_distributor_agreement_end": date(2026, 8, 1),
            })

    def test_distributor_menu_replaces_manufacturing_launcher(self):
        self.assertTrue(
            self.env.ref(
                "stock_subwarehouse_hierarchy.menu_distributor_root"
            ).active
        )
        self.assertFalse(self.env.ref("mrp.menu_mrp_root").active)


class TestSupplierPrototype(TransactionCase):
    def test_supplier_import_template_and_test_row(self):
        from openpyxl import load_workbook

        content = self.env[
            "res.partner"
        ]._generate_partner_channel_import_template_xlsx("supplier")
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook["供应商导入"]
        expected_fields = [
            field_name
            for field_name, _label, _required, _accepted in self.env[
                "res.partner"
            ]._get_supplier_import_columns()
        ]
        self.assertEqual(
            list(next(worksheet.iter_rows(values_only=True))),
            expected_fields,
        )

        import_fields = [
            "name",
            "supplier_rank",
            "company_type",
            "email",
            "x_supplier_status",
            "x_supplier_channel_type",
            "x_supplier_preferred",
            "x_supplier_lead_days",
        ]
        result = self.env["res.partner"].load(import_fields, [[
            "Imported Supplier Test",
            "1",
            "company",
            "supplier.import@example.com",
            "active",
            "manufacturer",
            "1",
            "30",
        ]])
        self.assertFalse(result["messages"])
        supplier = self.env["res.partner"].browse(result["ids"])
        self.assertEqual(supplier.supplier_rank, 1)
        self.assertEqual(supplier.x_supplier_status, "active")
        self.assertEqual(supplier.x_supplier_channel_type, "manufacturer")
        self.assertTrue(supplier.x_supplier_code)

        action = self.env["res.partner"].action_open_supplier_import([])
        self.assertEqual(action["tag"], "import")
        self.assertEqual(
            action["context"]["partner_channel_import_type"],
            "supplier",
        )
        templates = self.env["res.partner"].with_context(
            partner_channel_import_type="supplier"
        ).get_import_templates()
        self.assertEqual(
            [template["label"] for template in templates],
            ["供应商导入模板"],
        )
        all_template_labels = {
            template["label"]
            for template in self.env["res.partner"].get_import_templates()
        }
        self.assertEqual(
            all_template_labels,
            {"经销商导入模板", "供应商导入模板"},
        )

        export_content = supplier._generate_partner_channel_export_xlsx(
            "supplier"
        )
        export_workbook = load_workbook(
            BytesIO(export_content),
            read_only=True,
            data_only=True,
        )
        export_rows = list(
            export_workbook["供应商导入"].iter_rows(values_only=True)
        )
        self.assertEqual(list(export_rows[0]), expected_fields)
        self.assertEqual(export_rows[1][0], "Imported Supplier Test")

    def test_supplier_profile_code_and_expected_receipts(self):
        supplier = self.env["res.partner"].create({
            "name": "Prototype Supplier",
            "is_company": True,
            "supplier_rank": 1,
            "x_supplier_status": "active",
            "x_supplier_channel_type": "manufacturer",
        })
        warehouse = self.env["stock.warehouse"].search([], limit=1)
        scheduled_date = datetime(2026, 8, 15, 10, 0, 0)
        receipt = self.env["stock.picking"].create({
            "partner_id": supplier.id,
            "picking_type_id": warehouse.in_type_id.id,
            "location_id": warehouse.in_type_id.default_location_src_id.id,
            "location_dest_id": warehouse.in_type_id.default_location_dest_id.id,
            "scheduled_date": scheduled_date,
        })

        self.assertTrue(supplier.x_supplier_code.startswith("SUP/"))
        self.assertEqual(supplier.x_supplier_expected_receipt_count, 1)
        self.assertEqual(supplier.x_supplier_next_receipt_date, scheduled_date)
        self.assertIn(
            ("partner_id", "child_of", supplier.id),
            supplier.action_view_supplier_expected_receipts()["domain"],
        )
        self.assertEqual(receipt.partner_id, supplier)

    def test_company_can_be_supplier_and_distributor(self):
        partner = self.env["res.partner"].create({
            "name": "Two-Way Trading Partner",
            "is_company": True,
            "supplier_rank": 1,
            "x_is_distributor": True,
        })

        self.assertTrue(partner.x_supplier_code)
        self.assertTrue(partner.x_distributor_code)

    def test_supplier_menu_is_separate_from_distributor_menu(self):
        supplier_menu = self.env.ref(
            "stock_subwarehouse_hierarchy.menu_supplier_root"
        )
        distributor_menu = self.env.ref(
            "stock_subwarehouse_hierarchy.menu_distributor_root"
        )

        self.assertTrue(supplier_menu.active)
        self.assertNotEqual(supplier_menu, distributor_menu)
